from __future__ import annotations

import argparse
import json
import secrets
import sys
import time
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.bootstrap.container import _build_application_container  # noqa: E402
from app.core.config import Settings  # noqa: E402
from app.core.database import create_database  # noqa: E402
from app.domain.accounts import AccountActor  # noqa: E402
from app.infrastructure.accounts import Pbkdf2PasswordAdapter  # noqa: E402
from app.models.user import User  # noqa: E402


OWNER_ID = "full-cv-import-owner"


def load_texts(paths: list[Path]) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError(f"CV workbook has no data rows: {path}")
        for row_index, row in enumerate(rows[1:], start=2):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if len(values) != 1:
                raise ValueError(f"CV row must contain exactly one raw-text cell: {path}:{row_index}")
            raw_text = values[0]
            source_record_id = f"{path.name}:{row_index}"
            records.append((source_record_id, raw_text))
        workbook.close()
    return records


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import every raw CV from one or more one-column workbooks."
    )
    parser.add_argument("--workbook", action="append", required=True, type=Path)
    parser.add_argument("--wait-seconds", type=int, default=7200)
    parser.add_argument("--poll-seconds", type=float, default=5.0)
    args = parser.parse_args(argv)

    records = load_texts(args.workbook)
    settings = Settings()
    if not settings.CV_EXTRACTION_ENABLED:
        raise ValueError("CV_EXTRACTION_ENABLED must be true")
    database = create_database(settings.DATABASE_URL)
    try:
        with database.session_factory() as session:
            owner = session.get(User, OWNER_ID)
            if owner is None:
                session.add(
                    User(
                        id=OWNER_ID,
                        username=OWNER_ID,
                        hashed_password=Pbkdf2PasswordAdapter().hash(
                            secrets.token_urlsafe(32)
                        ),
                        role="personal_user",
                    )
                )
                session.commit()
            elif owner.role != "personal_user":
                raise ValueError("CV import owner has an incompatible role")

        container = _build_application_container(settings, database)
        actor = AccountActor(OWNER_ID, "personal_user")
        task_ids: list[str] = []
        for source_record_id, raw_text in records:
            imported = container.cv_ingestion.import_and_schedule(
                actor,
                source_record_id=source_record_id,
                raw_text=raw_text,
                source_platform="full_cv_workbook_import",
            )
            task_ids.append(imported.cv_extraction_task_id)

        deadline = time.monotonic() + args.wait_seconds
        completed: dict[str, object] = {}
        while time.monotonic() < deadline:
            completed = {
                task_id: container.cv_ingestion.get(actor, task_id)
                for task_id in task_ids
            }
            if all(task.status in {"succeeded", "failed"} for task in completed.values()):
                break
            time.sleep(args.poll_seconds)
        else:
            raise TimeoutError("CV extraction batch did not finish before the deadline")

        failed = [
            {
                "task_id": task.task_id,
                "error_code": task.last_error_code,
                "error_message": task.last_error_message,
            }
            for task in completed.values()
            if task.status == "failed"
        ]
        if failed:
            raise RuntimeError(json.dumps({"failed": failed}, ensure_ascii=False))

        resume_ids: list[str] = []
        skill_count = 0
        for task in completed.values():
            if not task.resume_id:
                raise RuntimeError(f"CV task completed without Resume: {task.task_id}")
            parse_result = container.resumes.get_parse_result(actor, task.resume_id)
            if parse_result.need_review:
                container.resumes.confirm(actor, task.resume_id)
            skill_count += len(
                container.resumes.generate_skill_profile(actor, task.resume_id)
            )
            resume_ids.append(task.resume_id)

        print(
            json.dumps(
                {
                    "source_cv_count": len(records),
                    "task_count": len(task_ids),
                    "resume_count": len(resume_ids),
                    "resume_skill_count": skill_count,
                },
                ensure_ascii=False,
            )
        )
        return 0
    finally:
        database.dispose()


if __name__ == "__main__":
    raise SystemExit(main())
