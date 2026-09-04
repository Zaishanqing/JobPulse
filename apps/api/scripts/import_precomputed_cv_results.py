from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.config import Settings  # noqa: E402
from app.core.database import create_database  # noqa: E402
from app.models.cv_position_classification import (  # noqa: E402
    CVPositionClassification,
)
from app.models.resume import Resume  # noqa: E402
from app.models.source_cv import SourceCV, SourceCVVersion  # noqa: E402


OWNER_ID = "full-cv-import-owner"


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _workbook_texts(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    texts: list[str] = []
    for row_index, row in enumerate(rows[1:], start=2):
        values = [
            str(value).strip()
            for value in row
            if value is not None and str(value).strip()
        ]
        if len(values) != 1:
            raise ValueError(
                f"CV row must contain exactly one raw-text cell: {path}:{row_index}"
            )
        texts.append(values[0])
    return texts


def _load_precomputed(
    run_dirs: list[Path], workbooks: list[Path]
) -> list[dict[str, Any]]:
    if len(run_dirs) != len(workbooks):
        raise ValueError("--run-dir and --workbook counts must match")
    records: list[dict[str, Any]] = []
    for run_dir, workbook in zip(run_dirs, workbooks, strict=True):
        final = run_dir / "final"
        annotations = _read_json(final / "annotations_nested.json")
        normalized = _read_json(final / "normalized_annotations.json")
        position_features = [
            json.loads(line)
            for line in (final / "match_features.jsonl").read_text(
                encoding="utf-8"
            ).splitlines()
            if line.strip()
        ]
        flags = [
            json.loads(line)
            for line in (final / "review_flags.jsonl").read_text(
                encoding="utf-8"
            ).splitlines()
            if line.strip()
        ]
        texts = _workbook_texts(workbook)
        if not isinstance(annotations, list) or not isinstance(normalized, list):
            raise ValueError(f"precomputed CV outputs must be lists: {run_dir}")
        if len(texts) != len(annotations):
            # A continue-on-error run may contain a full input workbook but only
            # successful rows in the final outputs. The extractor's stable
            # document IDs encode the one-based input row (cv_000001 -> row 1),
            # so select the corresponding raw texts instead of rejecting the
            # whole run.
            selected: list[str] = []
            for extraction in annotations:
                document_id = str(extraction.get("document_id", ""))
                prefix, _, suffix = document_id.rpartition("_")
                if prefix != "cv" or not suffix.isdigit():
                    selected = []
                    break
                source_index = int(suffix)
                if source_index < 1 or source_index > len(texts):
                    selected = []
                    break
                selected.append(texts[source_index - 1])
            if len(selected) == len(annotations):
                texts = selected
        if len(texts) != len(annotations) or len(texts) != len(normalized):
            raise ValueError(
                f"workbook and precomputed result counts differ: {run_dir}"
            )
        flags_by_id: dict[str, list[dict[str, Any]]] = {}
        for flag in flags:
            document_id = flag.get("cv_id") or flag.get("document_id")
            if isinstance(document_id, str):
                flags_by_id.setdefault(document_id, []).append(flag)
        normalized_by_id = {str(row["document_id"]): row for row in normalized}
        positions_by_document: dict[str, list[dict[str, Any]]] = {}
        for feature in position_features:
            if feature.get("feature_type") != "role":
                continue
            structured = feature.get("structured_values")
            if not isinstance(structured, dict):
                continue
            if feature.get("taxonomy_version") != "position-taxonomy.v3.0.0":
                # cv50_fresh_20260809 predates the v3 position classifier. Its
                # role features are retained in the run artifacts, but they do
                # not contain safe v3 position codes for persistence.
                continue
            positions_by_document.setdefault(
                str(feature.get("document_id") or ""), []
            ).append(
                {
                    "source_object_id": feature.get("source_object_id"),
                    "source_scope": feature.get("source_scope"),
                    "raw_text": feature.get("raw_text"),
                    "taxonomy_version": feature.get("taxonomy_version"),
                    **structured,
                    "feature_evidence_refs": feature.get("evidence_refs") or [],
                }
            )
        for row_index, (raw_text, extraction) in enumerate(zip(texts, annotations, strict=True), start=2):
            old_id = str(extraction["document_id"])
            if old_id not in normalized_by_id:
                raise ValueError(f"normalized CV result is missing: {old_id}")
            records.append({
                "source_record_id": f"{workbook.name}:{row_index}",
                "old_id": old_id,
                "raw_text": raw_text,
                "extraction": extraction,
                "normalized": normalized_by_id[old_id],
                "review_flags": flags_by_id.get(old_id, []),
                "position_classifications": positions_by_document.get(old_id, []),
                "position_source_run_id": run_dir.name,
            })
    return records


def apply_position_classifications(
    records: list[dict[str, Any]],
    database_url: str,
    *,
    execute: bool,
) -> dict[str, int]:
    database = create_database(database_url)
    matched = updated = 0
    try:
        with database.session_factory() as session:
            for record in records:
                source = session.scalar(
                    select(SourceCV).where(
                        SourceCV.owner_id == OWNER_ID,
                        SourceCV.source_platform == "full_cv_workbook_import",
                        SourceCV.source_record_id == record["source_record_id"],
                    )
                )
                if source is None:
                    raise ValueError(
                        "historical CV source is missing: "
                        f"{record['source_record_id']}"
                    )
                versions = session.scalars(
                    select(SourceCVVersion)
                    .where(SourceCVVersion.source_cv_id == source.id)
                    .order_by(SourceCVVersion.created_at.desc())
                ).all()
                resumes = session.scalars(
                    select(Resume).where(
                        Resume.source_cv_version_id.in_(
                            [version.id for version in versions]
                        )
                    )
                ).all()
                if len(resumes) != 1:
                    raise ValueError(
                        "historical CV must resolve to exactly one Resume: "
                        f"{record['source_record_id']}"
                    )
                resume = resumes[0]
                version = next(
                    item for item in versions
                    if item.id == resume.source_cv_version_id
                )
                if version.raw_text != record["raw_text"]:
                    raise ValueError(
                        "historical CV text does not match v3 run: "
                        f"{record['source_record_id']}"
                    )
                matched += 1
                values = {
                    "taxonomy_version": "position-taxonomy.v3.0.0",
                    "classifications": record["position_classifications"],
                    "source_run_ids": [record["position_source_run_id"]],
                }
                current = session.get(CVPositionClassification, resume.id)
                if current is None:
                    session.add(
                        CVPositionClassification(resume_id=resume.id, **values)
                    )
                    updated += 1
                elif any(
                    getattr(current, key) != value
                    for key, value in values.items()
                ):
                    for key, value in values.items():
                        setattr(current, key, value)
                    updated += 1
            if execute:
                session.commit()
            else:
                session.rollback()
    finally:
        database.dispose()
    return {
        "precomputed_cv_count": len(records),
        "matched_existing_resumes": matched,
        "updated_position_classifications": updated,
        "model_api_calls": 0,
        "executed": int(execute),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Apply position-taxonomy.v3 results to existing CV resumes."
    )
    parser.add_argument("--run-dir", action="append", required=True, type=Path)
    parser.add_argument("--workbook", action="append", required=True, type=Path)
    parser.add_argument("--database-url")
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args(argv)

    records = _load_precomputed(args.run_dir, args.workbook)
    settings = Settings()
    result = apply_position_classifications(
        records,
        args.database_url or settings.DATABASE_URL,
        execute=args.execute,
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
