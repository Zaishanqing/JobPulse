"""Import successful precomputed CV outputs without calling an extraction model."""

from __future__ import annotations

import argparse
import copy
import json
import secrets
import sys
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import app.models  # noqa: E402,F401
from app.bootstrap.container import _build_application_container  # noqa: E402
from app.contexts.cv_ingestion import CVIngestionUseCases, CVReviewConfirmation  # noqa: E402
from app.contexts.data_validation import CVValidationPolicy, CVValidatorSet  # noqa: E402
from app.core.config import Settings  # noqa: E402
from app.core.database import create_database  # noqa: E402
from app.domain.accounts import AccountActor  # noqa: E402
from app.domain.json_types import FrozenJsonObject, freeze_json_object  # noqa: E402
from app.infrastructure.accounts import Pbkdf2PasswordAdapter  # noqa: E402
from app.infrastructure.cv_ingestion import (  # noqa: E402
    ApplicationResumeImporter,
    SqlAlchemyCVIngestionUnitOfWork,
)
from app.infrastructure.data_validation import SqlAlchemyValidationPortFactory  # noqa: E402
from app.models.user import User  # noqa: E402
from jobgraph_contracts.cv_extraction_http import CVExtractionResponseV3  # noqa: E402


OWNER_ID = "full-cv-import-owner"
SOURCE_PLATFORM = "full_cv_workbook_import"
PROVIDER = "precomputed_cv_output"
IMPORT_REVISION = "v2"
MAIN_SKILL_TAXONOMY_VERSION = "skill-taxonomy-catalog-current"
DEFAULT_POSITION_CATALOG = ROOT / "config" / "position_taxonomy_catalog.v3.json"


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _workbook_texts(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    texts: list[str] = []
    for row_index, row in enumerate(rows[1:], start=2):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if len(values) != 1:
            raise ValueError(f"CV row must contain exactly one raw-text cell: {path}:{row_index}")
        texts.append(values[0])
    return texts


def _classification(
    feature: Mapping[str, Any], position_catalog: Mapping[str, Mapping[str, Any]]
) -> dict[str, Any]:
    structured = dict(feature.get("structured_values") or {})
    schema_version = structured.pop("classification_schema_version", None)
    structured.pop("role_kind", None)
    if schema_version != "job-position-classification.v3":
        raise ValueError(
            f"CV role feature has unsupported classification schema: {schema_version!r}"
        )
    if structured.get("classification_status") in {"resolved", "manually_confirmed"}:
        position_code = str(structured.get("position_code") or "")
        catalog_item = position_catalog.get(position_code)
        if catalog_item is None:
            raise ValueError(f"resolved CV position is absent from catalog: {position_code}")
        structured["position_name"] = catalog_item["position_name"]
        structured["family_code"] = catalog_item["family_code"]
        structured["family_name"] = catalog_item["family_name"]
    return {
        "schema_version": schema_version,
        "taxonomy_version": feature.get("taxonomy_version"),
        "source_title": feature.get("raw_text"),
        **structured,
    }


def _position_classifications(
    document_id: str,
    features: list[Mapping[str, Any]],
    position_catalog: Mapping[str, Mapping[str, Any]],
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for feature in features:
        if feature.get("document_id") != document_id or feature.get("feature_type") != "role":
            continue
        structured_values = feature.get("structured_values") or {}
        if structured_values.get("classification_schema_version") != "job-position-classification.v3":
            # Older CV runs contain only raw role mentions. Keep the resume
            # importable, but do not project those mentions as v3 positions.
            continue
        role_kind = structured_values["role_kind"]
        source_object_id = str(feature["source_object_id"])
        if role_kind == "expected":
            feature_id = "role_personal_info_expected_position"
            source_object_id = "personal_info"
            source_scope = "personal_info.expected_position"
        else:
            feature_id = f"role_{source_object_id}_position"
            source_scope = "work_experience.position"
        result.append(
            {
                "feature_id": feature_id,
                "source_object_id": source_object_id,
                "source_scope": source_scope,
                "role_kind": role_kind,
                "job_classification": _classification(feature, position_catalog),
            }
        )
    return result


def _normalized_skill(item: Mapping[str, Any]) -> dict[str, Any]:
    classifications = item.get("classifications") or []
    primary_concept = next(
        (
            entry
            for entry in classifications
            if entry.get("facet") == "concept_class" and entry.get("is_primary") is True
        ),
        None,
    )
    primary_kind = next(
        (
            entry
            for entry in classifications
            if entry.get("facet") == "technology_kind" and entry.get("is_primary") is True
        ),
        None,
    )
    resolved = item.get("identity_resolution_status") == "resolved"
    return {
        "source_item_id": item["source_item_id"],
        "source_scope": item["source_scope"],
        "source_name": item["source_name"],
        "skill_id": item.get("skill_id"),
        "canonical_name": item.get("canonical_name"),
        "category_code": (primary_concept["code"] if primary_concept is not None else "unknown"),
        "subcategory_code": (primary_kind["code"] if primary_kind is not None else None),
        "resolution_status": "resolved" if resolved else "unresolved",
        "normalization_confidence": 1.0 if resolved else None,
        "resolution_source": "explicit_mapping" if resolved else "unresolved",
    }


def _skill_taxonomy(normalized_skills: list[Mapping[str, Any]]) -> dict[str, Any]:
    skills: dict[str, dict[str, Any]] = {}
    for item in normalized_skills:
        skill_id = item.get("skill_id")
        canonical_name = item.get("canonical_name")
        classifications = item.get("classifications") or []
        if not skill_id or not canonical_name or not classifications:
            continue
        candidate = {
            "skill_id": skill_id,
            "canonical_name": canonical_name,
            "classifications": classifications,
        }
        previous = skills.setdefault(str(skill_id), candidate)
        if previous != candidate:
            raise ValueError(f"inconsistent skill taxonomy projection: {skill_id}")
    return {
        "schema_version": "skill-taxonomy-projection.v1",
        "taxonomy_version": MAIN_SKILL_TAXONOMY_VERSION,
        "skills": list(skills.values()),
    }


def _bind_document_id(value: Any, document_id: str) -> Any:
    if isinstance(value, dict):
        rebound = {key: _bind_document_id(child, document_id) for key, child in value.items()}
        if "source_id" in rebound and "quote" in rebound:
            rebound["source_document_id"] = document_id
        return rebound
    if isinstance(value, list):
        return [_bind_document_id(child, document_id) for child in value]
    return value


def build_response(
    record: Mapping[str, Any],
    normalized: Mapping[str, Any],
    role_features: list[Mapping[str, Any]],
    manifest: Mapping[str, Any],
    *,
    document_id: str,
    position_catalog: Mapping[str, Mapping[str, Any]] | None = None,
) -> dict[str, Any]:
    old_id = str(record["cv_id"])
    extraction = _bind_document_id(copy.deepcopy(record["annotation"]), document_id)
    extraction["document_id"] = document_id
    source_skills = list(normalized.get("normalized_skills") or [])
    normalized_result = {
        "document_id": document_id,
        "normalized_skills": [_normalized_skill(item) for item in source_skills],
        "unresolved_items": list(normalized.get("unresolved_items") or []),
        "position_classifications": _position_classifications(
            old_id, role_features, position_catalog or {}
        ),
    }
    review_flags = [
        {**dict(flag), "cv_id": document_id} for flag in record.get("review_flags") or []
    ]
    payload = {
        "contract_version": "cv-extraction-http.v3",
        "document_id": document_id,
        "execution": {
            "mode": "demo_snapshot",
            "provider": PROVIDER,
            "model": str(manifest.get("model") or "precomputed"),
            "prompt_version": f"precomputed:{manifest.get('run_id', 'unknown')}",
            "schema_version": str(manifest.get("extraction_schema_version") or "2.4"),
            "normalization_version": str(manifest.get("normalization_taxonomy_version") or "2.0"),
            "taxonomy_version": MAIN_SKILL_TAXONOMY_VERSION,
            "latency_ms": 0,
            "is_demo": False,
            "dataset_version": str(manifest.get("run_id") or "precomputed"),
        },
        "extraction_result": extraction,
        "normalized_result": normalized_result,
        "review_flags": review_flags,
        "skill_taxonomy": _skill_taxonomy(source_skills),
    }
    return CVExtractionResponseV3.model_validate(payload).model_dump(mode="json")


def _load_position_catalog(path: Path) -> dict[str, dict[str, Any]]:
    payload = _read_json(path)
    families = {str(item["code"]): str(item["name"]) for item in payload.get("families") or []}
    return {
        str(item["code"]): {
            "position_name": item["name"],
            "family_code": item["family_code"],
            "family_name": families[str(item["family_code"])],
        }
        for item in payload.get("positions") or []
    }


def load_records(
    run_dir: Path,
    workbook: Path,
    position_catalog_path: Path = DEFAULT_POSITION_CATALOG,
) -> list[dict[str, Any]]:
    final = run_dir / "final"
    manifest = _read_json(run_dir / "manifest.json")
    normalized_rows = _read_json(final / "normalized_annotations.json")
    normalized_by_id = {str(item["document_id"]): item for item in normalized_rows}
    role_features: list[dict[str, Any]] = []
    for line in (final / "match_features.jsonl").read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        feature = json.loads(line)
        if feature.get("feature_type") == "role":
            role_features.append(feature)
    texts = _workbook_texts(workbook)
    position_catalog = _load_position_catalog(position_catalog_path)
    records: list[dict[str, Any]] = []
    for path in sorted((run_dir / "records" / "success").glob("*.json")):
        record = _read_json(path)
        index = int(record["index"])
        old_id = str(record["cv_id"])
        normalized = normalized_by_id.get(old_id)
        if normalized is None:
            raise ValueError(f"normalized CV result is missing: {old_id}")
        if index < 1 or index > len(texts):
            raise ValueError(f"CV row index is outside workbook: {path}:{index}")
        records.append(
            {
                "source_record_id": f"{workbook.name}:{index + 1}",
                "resume_id": old_id,
                "raw_text": texts[index - 1],
                "record": record,
                "normalized": normalized,
                "role_features": role_features,
                "manifest": manifest,
                "source_version": str(manifest.get("run_id") or run_dir.name),
                "position_catalog": position_catalog,
            }
        )
    return records


def summarize_failed_records(run_dir: Path) -> dict[str, Any]:
    """Summarize extraction failures without retrying an external model service."""
    failures: list[dict[str, str | int | None]] = []
    by_reason: dict[str, int] = {}
    for path in sorted((run_dir / "records" / "failed").glob("*.json")):
        payload = _read_json(path)
        failure = payload.get("failed_case") or {}
        error_type = str(failure.get("error_type") or "unknown")
        stage = str(failure.get("stage") or "unknown")
        key = f"{error_type}:{stage}"
        by_reason[key] = by_reason.get(key, 0) + 1
        failures.append(
            {
                "cv_id": str(payload.get("cv_id") or failure.get("cv_id") or ""),
                "row_index": failure.get("row_index") or payload.get("index"),
                "error_type": error_type,
                "stage": stage,
                "error_message": str(failure.get("error_message") or ""),
            }
        )
    return {
        "failed_precomputed_cv_count": len(failures),
        "failure_reasons": dict(sorted(by_reason.items())),
        "failed_cases": failures,
    }


class PrecomputedCVProvider:
    def __init__(self, item: Mapping[str, Any]) -> None:
        self._item = item
        self.request_id = (
            f"{PROVIDER}:{IMPORT_REVISION}:{item['source_version']}:{item['resume_id']}"
        )

    def extract(
        self, *, document_id: str, raw_text: str, progress_callback=None
    ) -> FrozenJsonObject:
        if raw_text != self._item["raw_text"]:
            raise ValueError("precomputed CV raw text does not match imported source")
        response = build_response(
            self._item["record"],
            self._item["normalized"],
            self._item["role_features"],
            self._item["manifest"],
            document_id=document_id,
            position_catalog=self._item["position_catalog"],
        )
        return freeze_json_object(response, field="precomputed_cv_response")


def _ensure_owner(database) -> None:
    with database.session_factory() as session:
        owner = session.get(User, OWNER_ID)
        if owner is None:
            session.add(
                User(
                    id=OWNER_ID,
                    username=OWNER_ID,
                    hashed_password=Pbkdf2PasswordAdapter().hash(secrets.token_urlsafe(32)),
                    role="personal_user",
                )
            )
            session.commit()
        elif owner.role != "personal_user":
            raise ValueError("CV import owner has an incompatible role")


def import_records(records: list[dict[str, Any]], settings: Settings) -> dict[str, Any]:
    database = create_database(settings.DATABASE_URL)
    actor = AccountActor(OWNER_ID, "personal_user")
    imported: list[dict[str, str]] = []
    try:
        _ensure_owner(database)
        container = _build_application_container(settings, database)
        for item in records:
            provider = PrecomputedCVProvider(item)
            use_cases = CVIngestionUseCases(
                lambda: SqlAlchemyCVIngestionUnitOfWork(database.session_factory),
                provider,
                ApplicationResumeImporter(container.resumes),
                CVValidatorSet(
                    CVValidationPolicy(
                        version=settings.CV_EXTRACTION_VALIDATION_POLICY_VERSION,
                        provider=PROVIDER,
                        model=str(item["manifest"].get("model") or "precomputed"),
                        prompt_version=f"precomputed:{item['source_version']}",
                        schema_version=str(
                            item["manifest"].get("extraction_schema_version") or "2.4"
                        ),
                        normalization_version=str(
                            item["manifest"].get("normalization_taxonomy_version") or "2.0"
                        ),
                    ),
                    SqlAlchemyValidationPortFactory(database.session_factory).current_catalog,
                ),
                enabled=True,
                max_attempts=settings.CV_EXTRACTION_MAX_ATTEMPTS,
            )
            scheduled = use_cases.import_and_schedule(
                actor,
                source_record_id=item["source_record_id"],
                raw_text=item["raw_text"],
                source_platform=SOURCE_PLATFORM,
                source_version=item["source_version"],
            )
            task = use_cases.run(actor, scheduled.cv_extraction_task_id)
            if task.status != "succeeded" or task.review_id is None:
                raise RuntimeError(
                    f"precomputed CV import did not reach review: {item['resume_id']}"
                )
            confirmed = use_cases.confirm(
                actor,
                task.task_id,
                CVReviewConfirmation(
                    expected_review_id=task.review_id,
                    idempotency_key=f"{provider.request_id}:confirm",
                    field_decisions=(),
                    display_name=str(
                        (item["record"].get("annotation") or {})
                        .get("personal_info", {})
                        .get("name")
                        or item["resume_id"]
                    ),
                ),
                resume_id=item["resume_id"],
            )
            imported.append(
                {
                    "resume_id": confirmed.resume_id,
                    "snapshot_id": confirmed.snapshot_id,
                    "task_id": task.task_id,
                }
            )
    finally:
        database.dispose()
    return {
        "precomputed_cv_count": len(records),
        "imported_cv_count": len(imported),
        "model_api_calls": 0,
        "items": imported,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-dir", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args(argv)
    records = load_records(args.run_dir, args.workbook)
    for item in records:
        build_response(
            item["record"],
            item["normalized"],
            item["role_features"],
            item["manifest"],
            document_id=f"check-{item['resume_id']}",
            position_catalog=item["position_catalog"],
        )
    result: dict[str, Any] = {
        "precomputed_cv_count": len(records),
        "validated_contract_count": len(records),
        "model_api_calls": 0,
        "executed": False,
    }
    result.update(summarize_failed_records(args.run_dir))
    if args.execute:
        result = {**import_records(records, Settings()), "executed": True}
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
