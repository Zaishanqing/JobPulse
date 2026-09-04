from base64 import b64decode
from datetime import date
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError

from tests.runtime_database import reset_database_data
from tests.runtime_database import SessionLocal
from app.application.jd import JDApplicationError, JDUseCases, JDTextCreateCommand
from app.infrastructure.jd_repository import SqlAlchemyJDUoW
from app.infrastructure.jd_schema import VersionedJDSchemaAdapter
from app.main import app
from app.models.jd import JobDescription
from app.models.jd_parse_result import JDParseResult
from app.models.skill import Skill
from app.models.standard_position import StandardPosition
from app.models.task_record import TaskRecord
from app.models.review_task import ReviewTask
from app.ports.jd_repository import Actor
from tests.user_factory import create_internal_user
from app.api.contracts.jd import JDExtractionResult


client = TestClient(app)


@pytest.fixture(autouse=True)
def reset_database():
    reset_database_data()
    with SessionLocal() as session:
        session.add_all(
            [
                Skill(
                    id=skill_id,
                    skill_name=name,
                    category=category,
                )
                for skill_id, name, category in (
                    (
                        "skill_python",
                        "Python",
                        "programming_language",
                    ),
                    ("skill_java", "Java", "programming_language"),
                    ("skill_spring_boot", "Spring Boot", "framework"),
                    ("skill_rag", "RAG", "methodology"),
                    ("skill_docker", "Docker", "tool"),
                    ("skill_kubernetes", "Kubernetes", "platform"),
                    ("skill_multi_agent", "多智能体", "methodology"),
                    ("skill_fastapi", "FastAPI", "framework"),
                )
            ]
        )
        session.commit()
    yield
    reset_database_data()


def _register_and_login(username: str, role: str) -> str:
    create_internal_user(username, role)
    client.post(
        "/api/v1/auth/register",
        json={
            "role": role,
            "username": username,
            "password": "password123",
            "email": f"{username}@example.com",
            "phone": "13800000000",
        },
    )
    response = client.post(
        "/api/v1/auth/login",
        json={"username": username, "password": "password123"},
    )
    token = response.json()["data"]["access_token"]
    if role == "enterprise_user":
        enterprise_response = client.post(
            "/api/v1/enterprises",
            json={
                "enterprise_name": f"{username}企业",
                "industry": "软件",
                "scale": "1-20人",
                "location": "武汉",
                "description": "JD 权限测试企业",
            },
            headers=_auth_headers(token),
        )
        assert enterprise_response.status_code == 200
    return token


def _auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _jd_payload(title: str = "初级大模型应用开发工程师") -> dict:
    return {
        "source_type": "enterprise_upload",
        "source_name": "企业上传",
        "enterprise_id": None,
        "title": title,
        "raw_text": "岗位职责：负责 Python RAG 应用开发，使用 Docker 交付服务。本科优先。",
        "publish_date": "2026-07-01",
        "url": "",
    }


def _create_jd(token: str, title: str = "初级大模型应用开发工程师") -> str:
    response = client.post(
        "/api/v1/jds/text",
        json=_jd_payload(title=title),
        headers=_auth_headers(token),
    )
    assert response.status_code == 200
    return response.json()["data"]["jd_id"]


@pytest.mark.parametrize("role", ["enterprise_user", "admin", "developer"])
def test_jd_business_capability_allows_create_and_parse(role: str):
    token = _register_and_login(f"jd_capability_{role}", role)
    jd_id = _create_jd(token, title=f"{role} JD")

    response = client.post(
        f"/api/v1/jds/{jd_id}/parse",
        json={"extraction_mode": "rule"},
        headers=_auth_headers(token),
    )

    assert response.status_code == 200


def test_reviewer_without_jd_business_capability_cannot_create_or_parse():
    reviewer_token = _register_and_login("jd_capability_reviewer", "reviewer")
    admin_token = _register_and_login("jd_capability_admin_owner", "admin")
    jd_id = _create_jd(admin_token, title="reviewer forbidden JD")

    create_response = client.post(
        "/api/v1/jds/text",
        json=_jd_payload(title="reviewer forbidden JD"),
        headers=_auth_headers(reviewer_token),
    )
    parse_response = client.post(
        f"/api/v1/jds/{jd_id}/parse",
        json={"extraction_mode": "rule"},
        headers=_auth_headers(reviewer_token),
    )

    assert create_response.status_code == 403
    assert parse_response.status_code == 403


def _bind_position(token: str, jd_id: str) -> None:
    parse_result = client.get(
        f"/api/v1/jds/{jd_id}/parse-result",
        headers=_auth_headers(token),
    ).json()["data"]
    with SessionLocal() as session:
        position = StandardPosition(
            position_code="BACKEND_ENGINEER",
            position_name="Backend Engineer",
            taxonomy_family_code="SOFTWARE_ENGINEERING",
            taxonomy_family_name="软件研发",
            skill_domain_codes=["software_engineering"],
            core_responsibilities=[],
            required_skills=[],
            bonus_skills=[],
            industry_scenarios=[],
            status="existing",
        )
        session.add(position)
        session.commit()
        position_id = position.id
    response = client.post(
        (
            f"/api/v1/jd-parse-results/{parse_result['parse_result_id']}"
            "/position-catalog-mapping"
        ),
        json={"target_position_id": position_id},
        headers=_auth_headers(token),
    )
    assert response.status_code == 200


def test_create_jd_text():
    token = _register_and_login("enterprise_jd001", "enterprise_user")

    response = client.post(
        "/api/v1/jds/text",
        json=_jd_payload(),
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 0
    assert payload["data"]["jd_id"]
    assert payload["data"]["parse_status"] == "pending"


def test_create_jd_text_stores_cleaned_text():
    token = _register_and_login("enterprise_jd011", "enterprise_user")
    payload = _jd_payload()
    payload["cleaned_text"] = "岗位职责:负责 Python RAG 应用开发,使用 Docker 交付服务。本科优先。"

    response = client.post(
        "/api/v1/jds/text",
        json=payload,
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    jd_id = response.json()["data"]["jd_id"]
    db = SessionLocal()
    try:
        row = db.query(JobDescription).filter(JobDescription.id == jd_id).first()
        assert row is not None
        assert row.raw_text == payload["raw_text"]
        assert row.cleaned_text == payload["cleaned_text"]
    finally:
        db.close()


def test_jd_use_case_rolls_back_real_database_when_commit_fails():
    class FailingCommitUoW(SqlAlchemyJDUoW):
        def commit(self) -> None:
            self.rollback()
            raise RuntimeError("simulated commit failure")

    use_cases = JDUseCases(
        lambda: FailingCommitUoW(SessionLocal), object(), VersionedJDSchemaAdapter()
    )

    with pytest.raises(RuntimeError):
        use_cases.create_text(
            Actor(id="admin-id", role="admin"),
            JDTextCreateCommand(
                **{**_jd_payload(), "publish_date": date(2026, 7, 1)}
            ),
        )

    db = SessionLocal()
    try:
        assert db.query(JobDescription).count() == 0
    finally:
        db.close()


def test_jd_permission_is_checked_before_any_database_write():
    token = _register_and_login("personal_jd_forbidden", "personal_user")

    response = client.post(
        "/api/v1/jds/text",
        json=_jd_payload(),
        headers=_auth_headers(token),
    )

    assert response.status_code == 403
    db = SessionLocal()
    try:
        assert db.query(JobDescription).count() == 0
    finally:
        db.close()


def test_get_jd_list():
    token = _register_and_login("enterprise_jd002", "enterprise_user")
    jd_id = _create_jd(token)

    response = client.get("/api/v1/jds", headers=_auth_headers(token))

    assert response.status_code == 200
    data = response.json()["data"]
    assert len(data) == 1
    assert data[0]["jd_id"] == jd_id


def test_edit_jd_raw_text():
    token = _register_and_login("enterprise_jd003", "enterprise_user")
    jd_id = _create_jd(token)

    response = client.put(
        f"/api/v1/jds/{jd_id}/raw",
        json={"raw_text": "岗位职责：负责 Java Spring Boot 后端服务开发。"},
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert "Java Spring Boot" in data["raw_text"]
    assert data["parse_status"] == "pending"


def test_start_jd_rule_parse_uses_quality_score_and_requires_review():
    token = _register_and_login("enterprise_jd004", "enterprise_user")
    jd_id = _create_jd(token)

    response = client.post(
        f"/api/v1/jds/{jd_id}/parse",
        json={"model": "default", "extraction_mode": "rule"},
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["jd_id"] == jd_id
    assert data["status"] == "completed"
    quality_score = data["parse_result"]["parse_confidence"]
    assert 0.0 <= quality_score <= 1.0
    assert quality_score != 0.5
    assert data["parse_result"]["need_review"] is True
    with SessionLocal() as session:
        review_task = (
            session.query(ReviewTask)
            .filter(
                ReviewTask.object_id
                == data["parse_result"]["parse_result_id"]
            )
            .one()
        )
        assert review_task.priority == "normal"


def test_llm_parse_failure_creates_no_success_result_or_review_task():
    token = _register_and_login("enterprise_jd_llm_failure", "enterprise_user")
    jd_id = _create_jd(token)

    response = client.post(
        f"/api/v1/jds/{jd_id}/parse",
        json={"extraction_mode": "llm"},
        headers=_auth_headers(token),
    )

    assert response.status_code == 409
    assert "extraction_provider_not_configured" in str(response.json())
    with SessionLocal() as session:
        assert session.query(JDParseResult).filter_by(jd_id=jd_id).count() == 0
        assert session.query(ReviewTask).count() == 0
        assert session.query(TaskRecord).filter_by(task_type="jd_parse").count() == 0


def test_get_jd_parse_result():
    token = _register_and_login("enterprise_jd005", "enterprise_user")
    jd_id = _create_jd(token)
    client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))

    response = client.get(
        f"/api/v1/jds/{jd_id}/parse-result",
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["jd_id"] == jd_id
    assert data["position_title"] == "初级大模型应用开发工程师"
    assert any(skill["raw_skill"] == "Python" for skill in data["required_skills"])
    assert any(skill["raw_skill"] == "RAG" for skill in data["required_skills"])
    assert data["compatibility"] == {
        "legacy_fields": [
            "position_title",
            "responsibilities",
            "required_skills",
            "bonus_skills",
            "education",
            "experience",
            "industry",
            "tools",
            "business_scenarios",
        ],
        "source": "versioned_domain_adapter",
    }


def test_edit_jd_parse_result():
    token = _register_and_login("enterprise_jd006", "enterprise_user")
    jd_id = _create_jd(token)
    client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))

    legacy_response = client.put(
        f"/api/v1/jds/{jd_id}/parse-result",
        json={
            "position_title": "Python 后端开发工程师",
            "required_skills": [
                {
                    "raw_skill": "Python",
                    "normalized_skill_id": "skill_python",
                    "confidence": 0.98,
                }
            ],
            "parse_confidence": 0.95,
        },
        headers=_auth_headers(token),
    )

    assert legacy_response.status_code == 422
    assert legacy_response.json()["code"] == 422
    assert legacy_response.json()["message"] == (
        "Legacy compatibility fields are read-only; edit extraction_result "
        "or normalized_result instead: position_title, required_skills"
    )
    before = client.get(
        f"/api/v1/jds/{jd_id}/parse-result", headers=_auth_headers(token)
    ).json()["data"]
    contract_response = client.put(
        f"/api/v1/jds/{jd_id}/parse-result",
        json={"extraction_result": before["extraction_result"]},
        headers=_auth_headers(token),
    )
    assert contract_response.status_code == 200
    data = contract_response.json()["data"]
    assert data["extraction_result"] == before["extraction_result"]
    assert data["schema_version"] == "v2"


def test_put_versioned_parse_result_creates_without_reextracting():
    token = _register_and_login("enterprise_jd610", "enterprise_user")
    source_id = _create_jd(token)
    client.post(f"/api/v1/jds/{source_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))
    source = client.get(
        f"/api/v1/jds/{source_id}/parse-result",
        headers=_auth_headers(token),
    ).json()["data"]
    target_id = _create_jd(token)
    extraction = source["extraction_result"]
    normalized = source["normalized_result"]
    extraction["document_id"] = target_id
    normalized["document_id"] = target_id

    response = client.put(
        f"/api/v1/jds/{target_id}/parse-result",
        json={
            "extraction_result": extraction,
            "normalized_result": normalized,
        },
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["extraction_result"] == extraction
    assert data["normalized_result"]["document_id"] == target_id
    def normalized_skill_ids(result):
        skill_ids = set()
        for requirement in result["normalized_requirements"]:
            if "normalized_skills" in requirement:
                skill_ids.update(
                    skill["skill_id"]
                    for skill in requirement["normalized_skills"]
                    if skill.get("skill_id")
                )
            elif requirement.get("skill_id"):
                skill_ids.add(requirement["skill_id"])
        return skill_ids

    assert normalized_skill_ids(data["normalized_result"]) == normalized_skill_ids(
        normalized
    )
    assert data["need_review"] is True


def test_confirm_jd_parse_result():
    owner_token = _register_and_login("enterprise_jd007", "enterprise_user")
    reviewer_token = _register_and_login("reviewer_jd007", "reviewer")
    jd_id = _create_jd(owner_token)
    client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(owner_token))
    _bind_position(reviewer_token, jd_id)

    denied = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/confirm",
        headers=_auth_headers(owner_token),
    )
    response = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/confirm",
        headers=_auth_headers(reviewer_token),
    )

    assert denied.status_code == 403
    assert response.status_code == 200
    assert response.json()["data"]["need_review"] is False


def test_jd_publish_gate_requires_review_before_publication():
    owner_token = _register_and_login("enterprise_jd_publish_gate", "enterprise_user")
    admin_token = _register_and_login("admin_jd_publish_gate", "admin")
    jd_id = _create_jd(owner_token)
    parse_response = client.post(
        f"/api/v1/jds/{jd_id}/parse",
        json={"extraction_mode": "rule"},
        headers=_auth_headers(owner_token),
    )
    assert parse_response.status_code == 200

    publish_response = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/publish",
        headers=_auth_headers(admin_token),
    )

    assert publish_response.status_code == 409
    db = SessionLocal()
    try:
        result = db.query(JDParseResult).filter(JDParseResult.jd_id == jd_id).first()
        assert result.workflow_status == "draft"
        assert result.need_review is True
    finally:
        db.close()


def test_rule_parse_metadata_is_deterministic_and_publish_requires_review():
    owner_token = _register_and_login("enterprise_rule_meta", "enterprise_user")
    reviewer_token = _register_and_login("reviewer_rule_meta", "reviewer")
    admin_token = _register_and_login("admin_rule_meta", "admin")
    jd_id = _create_jd(owner_token)

    parse_response = client.post(
        f"/api/v1/jds/{jd_id}/parse",
        json={"extraction_mode": "rule"},
        headers=_auth_headers(owner_token),
    )
    assert parse_response.status_code == 200
    data = parse_response.json()["data"]
    assert data["execution_mode"] == "rule"
    assert data["capability_implementation_status"] == "deterministic_rule_jd_parse"
    assert data.get("review_only") is True
    assert data.get("mock") is None
    assert "mock_keyword_jd_parse" not in str(data)
    assert data["result_payload"].get("mock") is None
    assert data["result_payload"]["implementation_status"] == (
        "deterministic_rule_jd_parse"
    )

    # Rule results must not bypass Review/Publication gate.
    blocked_publish = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/publish",
        headers=_auth_headers(admin_token),
    )
    assert blocked_publish.status_code == 409

    _bind_position(reviewer_token, jd_id)
    confirmed = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/confirm",
        headers=_auth_headers(reviewer_token),
    )
    assert confirmed.status_code == 200
    published = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/publish",
        headers=_auth_headers(admin_token),
    )
    assert published.status_code == 200
    assert published.json()["data"]["workflow_status"] == "published"


def test_jd_parse_task_and_result_are_atomic(monkeypatch):
    from app.infrastructure import jd_repository

    def fail_task_creation(self, *args, **kwargs):
        raise RuntimeError("task persistence failed")

    monkeypatch.setattr(
        jd_repository.SqlAlchemyTaskRepository,
        "create_succeeded_task",
        fail_task_creation,
    )
    token = _register_and_login("enterprise_jd_atomic_task", "enterprise_user")
    jd_id = _create_jd(token)

    with pytest.raises(RuntimeError):
        client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))

    db = SessionLocal()
    try:
        jd = db.query(JobDescription).filter(JobDescription.id == jd_id).first()
        assert jd.parse_status == "pending"
        assert db.query(JDParseResult).filter(JDParseResult.jd_id == jd_id).count() == 0
        assert db.query(TaskRecord).filter(TaskRecord.task_type == "jd_parse").count() == 0
    finally:
        db.close()


def test_jd_parse_task_query_uses_application_permissions_without_writes():
    owner_token = _register_and_login("enterprise_jd_task_owner", "enterprise_user")
    other_token = _register_and_login("enterprise_jd_task_other", "enterprise_user")
    jd_id = _create_jd(owner_token)
    parsed = client.post(
        f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(owner_token)
    )
    task_id = parsed.json()["data"]["task_id"]

    allowed = client.get(
        f"/api/v1/jds/parse-tasks/{task_id}", headers=_auth_headers(owner_token)
    )
    with SessionLocal() as before_db:
        before = (
            before_db.query(JobDescription).filter_by(id=jd_id).one().parse_status,
            before_db.query(JDParseResult).filter_by(jd_id=jd_id).count(),
            before_db.query(TaskRecord).filter_by(id=task_id).one().status,
            before_db.query(TaskRecord).count(),
        )

    denied = client.get(
        f"/api/v1/jds/parse-tasks/{task_id}", headers=_auth_headers(other_token)
    )
    missing = client.get(
        "/api/v1/jds/parse-tasks/missing-task",
        headers=_auth_headers(owner_token),
    )
    with SessionLocal() as after_db:
        after = (
            after_db.query(JobDescription).filter_by(id=jd_id).one().parse_status,
            after_db.query(JDParseResult).filter_by(jd_id=jd_id).count(),
            after_db.query(TaskRecord).filter_by(id=task_id).one().status,
            after_db.query(TaskRecord).count(),
        )

    assert allowed.status_code == 200
    assert allowed.json()["code"] == 0
    assert allowed.json()["message"] == "success"
    assert allowed.json()["data"]["task_id"] == task_id
    assert allowed.json()["data"]["canonical_status"] == "succeeded"
    assert denied.status_code == 403
    assert missing.status_code == 404
    assert after == before


def test_jd_duplicate_check_batch_rolls_back_every_item_when_second_fails(monkeypatch):
    token = _register_and_login("admin_jd_duplicate_atomic", "admin")
    jd_ids = [_create_jd(token, "重复检测一"), _create_jd(token, "重复检测二")]
    original = JDUseCases._duplicate_check
    calls = 0

    def fail_second(self, uow, jd):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("second duplicate check failed")
        return original(self, uow, jd)

    monkeypatch.setattr(JDUseCases, "_duplicate_check", fail_second)
    with pytest.raises(RuntimeError):
        client.post(
            "/api/v1/jds/duplicate-check-batch",
            json=jd_ids,
            headers=_auth_headers(token),
        )

    with SessionLocal() as verify_db:
        rows = verify_db.query(JobDescription).filter(JobDescription.id.in_(jd_ids)).all()
        assert len(rows) == 2
        assert all(row.copy_risk_score is None for row in rows)


def test_jd_inflation_check_batch_rolls_back_every_item_when_second_fails(monkeypatch):
    token = _register_and_login("admin_jd_inflation_atomic", "admin")
    jd_ids = [_create_jd(token, "通胀检测一"), _create_jd(token, "通胀检测二")]
    for jd_id in jd_ids:
        response = client.post(
            f"/api/v1/jds/{jd_id}/parse",
            json={"extraction_mode": "rule"},
            headers=_auth_headers(token),
        )
        assert response.status_code == 200
    original = JDUseCases._inflation_check_for_jd
    calls = 0

    def fail_second(self, uow, actor, jd):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("second inflation check failed")
        return original(self, uow, actor, jd)

    monkeypatch.setattr(JDUseCases, "_inflation_check_for_jd", fail_second)
    with pytest.raises(RuntimeError):
        client.post(
            "/api/v1/jds/inflation-check-batch",
            json=jd_ids,
            headers=_auth_headers(token),
        )

    with SessionLocal() as verify_db:
        rows = verify_db.query(JobDescription).filter(JobDescription.id.in_(jd_ids)).all()
        assert len(rows) == 2
        assert all(row.inflation_score is None and row.parse_status == "completed" for row in rows)
        assert verify_db.query(JDParseResult).filter(JDParseResult.jd_id.in_(jd_ids)).count() == 2
        assert verify_db.query(TaskRecord).filter(TaskRecord.task_type == "jd_parse").count() == 2


def test_jd_batch_permissions_are_rejected_before_any_database_write():
    admin_token = _register_and_login("admin_jd_batch_owner", "admin")
    personal_token = _register_and_login("personal_jd_batch_denied", "personal_user")
    jd_ids = [_create_jd(admin_token, "批量权限一"), _create_jd(admin_token, "批量权限二")]

    duplicate = client.post(
        "/api/v1/jds/duplicate-check-batch",
        json=jd_ids,
        headers=_auth_headers(personal_token),
    )
    inflation = client.post(
        "/api/v1/jds/inflation-check-batch",
        json=jd_ids,
        headers=_auth_headers(personal_token),
    )

    with SessionLocal() as verify_db:
        rows = verify_db.query(JobDescription).filter(JobDescription.id.in_(jd_ids)).all()
        assert len(rows) == 2
        assert all(
            row.copy_risk_score is None
            and row.inflation_score is None
            and row.parse_status == "pending"
            for row in rows
        )
        assert verify_db.query(JDParseResult).filter(JDParseResult.jd_id.in_(jd_ids)).count() == 0
        assert verify_db.query(TaskRecord).count() == 0

    assert duplicate.status_code == 403
    assert inflation.status_code == 403


def test_jd_batch_checks_preserve_requested_order_and_response_fields():
    token = _register_and_login("admin_jd_batch_contract", "admin")
    first = _create_jd(token, "批量契约一")
    second = _create_jd(token, "批量契约二")
    requested_order = [second, first]
    for jd_id in requested_order:
        response = client.post(
            f"/api/v1/jds/{jd_id}/parse",
            json={"extraction_mode": "rule"},
            headers=_auth_headers(token),
        )
        assert response.status_code == 200

    duplicate = client.post(
        "/api/v1/jds/duplicate-check-batch",
        json=requested_order,
        headers=_auth_headers(token),
    )
    inflation = client.post(
        "/api/v1/jds/inflation-check-batch",
        json=requested_order,
        headers=_auth_headers(token),
    )

    assert duplicate.status_code == 200
    assert duplicate.json()["data"]["checked_count"] == 2
    assert [item["jd_id"] for item in duplicate.json()["data"]["items"]] == requested_order
    assert {
        "jd_id",
        "copy_risk_score",
        "similar_jds",
        "recommended_action",
        "reason",
    } <= duplicate.json()["data"]["items"][0].keys()
    assert inflation.status_code == 200
    assert inflation.json()["data"]["checked_count"] == 2
    assert [item["jd_id"] for item in inflation.json()["data"]["items"]] == requested_order
    assert {
        "jd_id",
        "inflation_score",
        "abnormal_skills",
        "recommended_action",
    } <= inflation.json()["data"]["items"][0].keys()


def test_jd_abnormal_skill_permission_is_enforced_by_application_before_uow():
    entered_uow = False

    def forbidden_factory():
        nonlocal entered_uow
        entered_uow = True
        raise AssertionError("UoW must not be opened before permission is accepted")

    use_cases = JDUseCases(forbidden_factory, object(), VersionedJDSchemaAdapter())
    with pytest.raises(JDApplicationError) as exc_info:
        use_cases.mark_parse_skill_abnormal(
            Actor(id="personal-id", role="personal_user"),
            "jd-id",
            "skill-id",
            abnormal=True,
            reason="not allowed",
        )

    assert exc_info.value.error_code == "forbidden"
    assert entered_uow is False


@pytest.mark.parametrize(
    ("abnormal", "reason", "message"),
    [
        ("true", None, "abnormal must be a boolean"),
        (True, 123, "reason must be a string"),
    ],
)
def test_jd_abnormal_skill_business_input_rules_are_in_application(
    abnormal, reason, message
):
    entered_uow = False

    def invalid_factory():
        nonlocal entered_uow
        entered_uow = True
        raise AssertionError("invalid business input must be rejected before UoW")

    use_cases = JDUseCases(invalid_factory, object(), VersionedJDSchemaAdapter())
    with pytest.raises(JDApplicationError) as exc_info:
        use_cases.mark_parse_skill_abnormal(
            Actor(id="admin-id", role="admin"),
            "jd-id",
            "skill-id",
            abnormal=abnormal,
            reason=reason,
        )

    assert exc_info.value.error_code == "invalid"
    assert exc_info.value.detail == message
    assert entered_uow is False


def test_admin_can_persist_and_clear_abnormal_parse_skill_mark():
    admin_token = _register_and_login("admin_jd_abnormal", "admin")
    jd_id = _create_jd(admin_token)
    client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(admin_token))
    _bind_position(admin_token, jd_id)

    mark_response = client.put(
        f"/api/v1/jds/{jd_id}/skills/skill_python/mark-abnormal",
        json={"abnormal": True, "reason": "需要复核"},
        headers=_auth_headers(admin_token),
    )
    result_response = client.get(
        f"/api/v1/jds/{jd_id}/parse-result",
        headers=_auth_headers(admin_token),
    )
    review_with_abnormal_skill_response = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/confirm",
        headers=_auth_headers(admin_token),
    )
    clear_response = client.put(
        f"/api/v1/jds/{jd_id}/skills/skill_python/mark-abnormal",
        json={"abnormal": False},
        headers=_auth_headers(admin_token),
    )
    cleared_result = client.get(
        f"/api/v1/jds/{jd_id}/parse-result",
        headers=_auth_headers(admin_token),
    )
    confirmed_response = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/confirm",
        headers=_auth_headers(admin_token),
    )
    published_response = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/publish",
        headers=_auth_headers(admin_token),
    )
    missing_response = client.put(
        f"/api/v1/jds/{jd_id}/skills/missing/mark-abnormal",
        headers=_auth_headers(admin_token),
    )

    assert mark_response.status_code == 200
    assert mark_response.json()["data"]["abnormal"] is True
    assert mark_response.json()["data"]["implementation_status"] == (
        "domain_review_flag_persisted_via_contract"
    )
    python_flag = next(
        item for item in result_response.json()["data"]["normalized_result"]["unresolved_items"]
        if item["source_value"] == "Python"
    )
    assert python_flag["severity"] == "blocking"
    assert python_flag["reason"] == "需要复核"
    assert review_with_abnormal_skill_response.status_code == 200
    assert (
        review_with_abnormal_skill_response.json()["data"]["workflow_status"]
        == "reviewed"
    )
    assert clear_response.json()["data"]["abnormal"] is False
    assert not any(
        item.get("source_value") == "Python"
        and item.get("code") == "skill_abnormal"
        for item in cleared_result.json()["data"]["normalized_result"][
            "unresolved_items"
        ]
    )
    assert confirmed_response.json()["data"]["workflow_status"] == "reviewed"
    assert published_response.json()["data"]["workflow_status"] == "published"
    assert missing_response.status_code == 409


def test_duplicate_check_returns_copy_risk_score():
    token = _register_and_login("enterprise_jd008", "enterprise_user")
    jd_id = _create_jd(token)

    response = client.post(
        f"/api/v1/jds/{jd_id}/duplicate-check",
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["jd_id"] == jd_id
    assert "copy_risk_score" in data
    assert isinstance(data["copy_risk_score"], float)


def test_inflation_check_returns_inflation_score():
    token = _register_and_login("enterprise_jd009", "enterprise_user")
    jd_id = _create_jd(token)
    client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))

    response = client.post(
        f"/api/v1/jds/{jd_id}/inflation-check",
        headers=_auth_headers(token),
    )

    assert response.status_code == 200
    data = response.json()["data"]
    assert data["jd_id"] == jd_id
    assert "inflation_score" in data
    assert data["inflation_score"] >= 0


def test_downweight_jd_sets_is_downweighted_true():
    enterprise_token = _register_and_login("enterprise_jd010", "enterprise_user")
    admin_token = _register_and_login("admin_jd010", "admin")
    jd_id = _create_jd(enterprise_token)

    response = client.put(
        f"/api/v1/jds/{jd_id}/downweight",
        headers=_auth_headers(admin_token),
    )

    assert response.status_code == 200
    assert response.json()["data"]["is_downweighted"] is True


def test_enterprise_user_cannot_access_another_enterprise_jd():
    owner_token = _register_and_login("enterprise_jd_owner011", "enterprise_user")
    other_token = _register_and_login("enterprise_jd_other011", "enterprise_user")
    jd_id = _create_jd(owner_token)

    list_response = client.get("/api/v1/jds", headers=_auth_headers(other_token))
    detail_response = client.get(
        f"/api/v1/jds/{jd_id}",
        headers=_auth_headers(other_token),
    )
    update_response = client.put(
        f"/api/v1/jds/{jd_id}/raw",
        json={"raw_text": "越权修改"},
        headers=_auth_headers(other_token),
    )

    assert list_response.status_code == 200
    assert list_response.json()["data"] == []
    assert detail_response.status_code == 403
    assert update_response.status_code == 403


def test_v2_pipeline_separates_extraction_normalization_and_exports_nine_sheets():
    token = _register_and_login("enterprise_jd_v2", "enterprise_user")
    reviewer_token = _register_and_login("reviewer_jd_v2", "reviewer")
    admin_token = _register_and_login("admin_jd_v2", "admin")
    payload = _jd_payload(title="Python 后端开发工程师")
    payload["raw_text"] = (
        "招聘岗位：Python 后端开发工程师。岗位职责：负责 FastAPI 服务开发与交付。"
        "要求 Python、FastAPI，3-5年经验，本科及以上。行业：人工智能。"
        "工作地点：武汉。薪资：15k-25k。五险一金，带薪培训。"
    )
    created = client.post("/api/v1/jds/text", json=payload, headers=_auth_headers(token))
    jd_id = created.json()["data"]["jd_id"]

    parsed = client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))
    result = client.get(
        f"/api/v1/jds/{jd_id}/parse-result", headers=_auth_headers(token)
    ).json()["data"]
    _bind_position(admin_token, jd_id)

    assert parsed.status_code == 200
    extraction = result["extraction_result"]
    normalized = result["normalized_result"]
    assert all(item["kind"] == "task" for item in extraction["responsibilities"])
    assert all(item["kind"] != "task" for item in extraction["requirements"])
    assert all(
        item["evidence"]["alignment"] == "exact"
        for group in (
            extraction["responsibilities"],
            extraction["requirements"],
            extraction["company_facts"],
            extraction["employment_facts"],
        )
        for item in group
    )
    forbidden = {"normalized", "skill_id", "job_family", "category_code", "subcategory_code"}
    assert forbidden.isdisjoint(str(extraction).replace("'", '"').split('"'))
    fastapi = next(
        item for item in normalized["normalized_requirements"] if item["source_name"] == "FastAPI"
    )
    assert fastapi["resolution_status"] == "resolved"
    assert fastapi["skill_id"] == "skill_fastapi"
    assert "salary" in {item["kind"] for item in extraction["employment_facts"]}
    employment_text = " ".join(
        item["value"] for item in extraction["employment_facts"]
    )
    assert "武汉" in employment_text
    assert "五险一金" in employment_text
    assert "带薪培训" in employment_text
    assert all(item["value"] != "带薪培训" for item in extraction["company_facts"])

    confirmed = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/confirm",
        headers=_auth_headers(reviewer_token),
    )
    published = client.post(
        f"/api/v1/jds/{jd_id}/parse-result/publish",
        headers=_auth_headers(admin_token),
    )
    exported = client.get(
        f"/api/v1/jds/{jd_id}/parse-result/export", headers=_auth_headers(token)
    )
    assert confirmed.json()["data"]["workflow_status"] == "reviewed"
    assert published.json()["data"]["workflow_status"] == "published"
    export_data = exported.json()["data"]
    workbook = load_workbook(BytesIO(b64decode(export_data["content_base64"])))
    assert workbook.sheetnames == [
        "document_summary", "responsibilities", "requirements", "skills",
        "company_facts", "employment_facts", "skill_normalization", "salary",
        "review_flags",
    ]


def test_v2_schema_rejects_extraction_normalization_field_leakage():
    with pytest.raises(ValidationError):
        JDExtractionResult.model_validate(
            {
                "document_id": "jd-1",
                "responsibilities": [],
                "requirements": [
                    {
                        "requirement_id": "req-1",
                        "kind": "skill",
                        "modality": "required",
                        "evidence": {
                            "source_id": "jd-1", "quote": "Python", "start": 0,
                            "end": 6, "alignment": "exact", "occurrence_index": 0,
                        },
                        "items": [{"name": "Python", "item_type": "programming_language"}],
                        "skill_id": "fabricated",
                    }
                ],
            }
        )


def test_non_exact_evidence_cannot_replace_or_publish_formal_result():
    token = _register_and_login("enterprise_jd_nonexact", "enterprise_user")
    jd_id = _create_jd(token)
    client.post(f"/api/v1/jds/{jd_id}/parse", json={"extraction_mode": "rule"}, headers=_auth_headers(token))
    result = client.get(
        f"/api/v1/jds/{jd_id}/parse-result", headers=_auth_headers(token)
    ).json()["data"]
    extraction = result["extraction_result"]
    extraction["requirements"][0]["evidence"]["alignment"] = "normalized_exact"

    edited = client.put(
        f"/api/v1/jds/{jd_id}/parse-result",
        json={"extraction_result": extraction},
        headers=_auth_headers(token),
    )

    assert edited.status_code == 409
    assert "Only exact evidence" in edited.json()["message"]
