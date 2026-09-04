import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from multi_company_scraper.collector import JobCollector

# Characters illegal in Excel cells (control chars except tab, newline, carriage return)
_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize(value):
    if isinstance(value, str):
        return _ILLEGAL_RE.sub("", value)
    return value


class ExcelWriter:
    COLUMNS = [
        ("company_name", "公司名称", 20),
        ("job_title", "职位名称", 30),
        ("job_id", "职位ID", 20),
        ("department", "所属部门", 20),
        ("city", "工作城市", 12),
        ("district", "区/县", 12),
        ("job_type", "工作类型", 10),
        ("salary_desc", "薪资原文", 20),
        ("jd_text", "JD全文", 80),
        ("source_url", "原始链接", 50),
        ("source_platform", "来源平台", 14),
        ("crawl_time", "爬取时间", 22),
        # --- new raw fields (task 02) ---
        ("experience_raw", "原始经验要求", 16),
        ("education_raw", "原始学历要求", 16),
        ("benefits_raw", "原始福利", 20),
        ("raw_text_status", "全文状态", 14),
        # --- deprecated semantic fields (kept for backward compat) ---
        ("experience", "[旧]经验要求", 10),
        ("education", "[旧]学历要求", 10),
        ("salary_min", "[旧]最低月薪(K)", 12),
        ("salary_max", "[旧]最高月薪(K)", 12),
        ("jd_responsibility", "[旧]岗位职责", 60),
        ("jd_requirement", "[旧]任职要求", 60),
        ("skill_tags", "[旧]技能标签", 30),
        ("benefits", "[旧]福利待遇", 30),
        ("publish_date", "发布时间", 14),
    ]

    @classmethod
    def write(cls, collector: JobCollector, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "全部职位"

        # Header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, (_, header, width) in enumerate(cls.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, job in enumerate(collector.get_all(), 2):
            d = job.to_dict()
            for col_idx, (field, _, _) in enumerate(cls.COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_sanitize(d.get(field, "")))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Freeze header
        ws.freeze_panes = "A2"
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cls.COLUMNS))}{collector.total() + 1}"

        # Summary sheet
        ws2 = wb.create_sheet("统计")
        stats = collector.stats()
        ws2.cell(1, 1, "总职位数").font = Font(bold=True)
        ws2.cell(1, 2, stats["total_jobs"])
        row = 3
        ws2.cell(row, 1, "各公司职位数").font = Font(bold=True)
        row += 1
        for company, count in stats["companies"].items():
            ws2.cell(row, 1, company)
            ws2.cell(row, 2, count)
            row += 1

        wb.save(output_path)

    @classmethod
    def write_text_only(cls, collector: JobCollector, output_path: str):
        """Single-column output — each row is one job's full jd_text block."""
        wb = Workbook()
        ws = wb.active
        ws.title = "职位原始文本"

        # Header
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        cell = ws.cell(row=1, column=1, value="原始文本")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions["A"].width = 120

        # Post-process: truncate at first noise marker
        _NOISE_CUT = (
            "猎聘温馨提示", "猜你喜欢", "公司信息", "推荐企业",
            "相关推荐", "相关公司", "热门城市", "热门招聘", "当前位置",
            "公司简介", "其他信息",
        )

        # Data rows — skip empty jd_text
        row_idx = 2
        for job in collector.get_all():
            text = _sanitize(job.jd_text)
            if not text or not text.strip():
                continue
            # Truncate at first noise marker (belt-and-suspenders)
            for marker in _NOISE_CUT:
                idx = text.find(marker)
                if idx > 100:  # only if marker is well into the content
                    text = text[:idx].rstrip()
                    break
            cell = ws.cell(row=row_idx, column=1, value=text)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_idx += 1

        ws.freeze_panes = "A2"

        wb.save(output_path)
