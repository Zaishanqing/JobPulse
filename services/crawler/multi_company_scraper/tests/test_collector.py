import os
import tempfile
from multi_company_scraper.models.job_data import JobData
from multi_company_scraper.collector import JobCollector
from multi_company_scraper.excel_writer import ExcelWriter


def test_collector_add_and_get():
    c = JobCollector()
    jd = JobData(company_name="测试", job_title="职位", source_platform="test")
    c.add(jd)
    assert c.total() == 1
    assert c.get_all() == [jd]


def test_collector_add_batch():
    c = JobCollector()
    jds = [
        JobData(company_name=f"公司{i}", job_title="职位", source_platform="test")
        for i in range(10)
    ]
    c.add_batch(jds)
    assert c.total() == 10


def test_collector_stats():
    c = JobCollector()
    jds = [
        JobData(company_name="A", job_title="x", source_platform="moka", city="北京"),
        JobData(company_name="B", job_title="y", source_platform="feishu", city="上海"),
        JobData(company_name="A", job_title="z", source_platform="moka", city="北京"),
    ]
    c.add_batch(jds)
    stats = c.stats()
    assert stats["total_jobs"] == 3
    assert stats["companies"]["A"] == 2
    assert stats["companies"]["B"] == 1
    assert stats["cities"]["北京"] == 2
    assert stats["cities"]["上海"] == 1


def test_collector_clear():
    c = JobCollector()
    c.add(JobData(company_name="测试", job_title="x", source_platform="t"))
    c.clear()
    assert c.total() == 0


def test_excel_writer_write_and_read():
    c = JobCollector()
    jds = [
        JobData(
            company_name="字节跳动",
            job_title="后端工程师",
            source_platform="playwright",
            city="北京",
            salary_min=30,
            salary_max=60,
            experience="3-5年",
            education="本科",
        ),
        JobData(
            company_name="腾讯",
            job_title="前端工程师",
            source_platform="tencent",
            city="深圳",
            salary_min=25,
            salary_max=50,
            experience="1-3年",
            education="本科",
        ),
    ]
    c.add_batch(jds)

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "test_output.xlsx")
        ExcelWriter.write(c, path)
        assert os.path.exists(path)
        # 验证文件可被openpyxl打开
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        assert ws.title == "全部职位"
        assert ws.cell(1, 1).value == "公司名称"  # header
        assert ws.cell(2, 1).value == "字节跳动"
        assert ws.cell(3, 1).value == "腾讯"
        wb.close()
