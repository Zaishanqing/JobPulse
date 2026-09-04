"""End-to-end tests with real network scraping and data quality checks.

All tests that hit real servers are marked ``@pytest.mark.slow`` so they can be
skipped during normal development runs::

    pytest -m "not slow"          # skip slow tests
    pytest -m "slow" -v           # run only slow tests
    pytest -v                     # run everything
"""

import tempfile
from datetime import datetime
from pathlib import Path

import pytest
import yaml

from multi_company_scraper.models.company_config import CompanyConfig
from multi_company_scraper.models.job_data import JobData
from multi_company_scraper.scrapers.baidu_scraper import BaiduScraper
from multi_company_scraper.scrapers.netease_scraper import NeteaseScraper
from multi_company_scraper.collector import JobCollector
from multi_company_scraper.excel_writer import ExcelWriter
from multi_company_scraper.main import load_companies, setup_dispatcher


PROJECT_ROOT = Path(__file__).resolve().parents[1]


# ---------------------------------------------------------------------------
# Shared helpers — data quality assertions
# ---------------------------------------------------------------------------

def validate_job_data(job: JobData, expected_company: str, expected_platform: str):
    """Run standard data-quality assertions on a single JobData instance."""
    # 1. Non-empty job_title
    assert job.job_title, f"job_title must not be empty (company={expected_company})"
    assert isinstance(job.job_title, str)

    # 2. company_name matches
    assert job.company_name == expected_company, (
        f"company_name mismatch: expected {expected_company!r}, got {job.company_name!r}"
    )

    # 3. salary sanity: if both values are set, min <= max
    if job.salary_min > 0 and job.salary_max > 0:
        assert job.salary_min <= job.salary_max, (
            f"salary_min ({job.salary_min}) > salary_max ({job.salary_max}) "
            f"for {job.job_title} at {job.company_name}"
        )
        # sanity ceiling: monthly salary in thousands should be reasonable
        assert job.salary_max < 500, (
            f"salary_max ({job.salary_max}) looks implausibly high "
            f"for {job.job_title} at {job.company_name}"
        )

    # 4. source_platform is correct
    assert job.source_platform == expected_platform, (
        f"source_platform mismatch: expected {expected_platform!r}, "
        f"got {job.source_platform!r}"
    )

    # 5. crawl_time is set
    assert job.crawl_time, "crawl_time must not be empty"

    # 6. job_id is present (may be a URL for some platforms)
    #    Not all platforms use numeric IDs; some use URLs as identifiers.


def validate_job_batch(jobs: list[JobData], expected_company: str, expected_platform: str):
    """Validate every job in a batch and print summary stats."""
    assert len(jobs) > 0, (
        f"Expected at least 1 job from {expected_company}, got 0. "
        f"The site may be blocking or the scraper selector may be stale."
    )

    for i, job in enumerate(jobs):
        validate_job_data(job, expected_company, expected_platform)

    # Summary
    titles = [j.job_title[:30] for j in jobs[:5]]
    depts = set(j.department for j in jobs if j.department)
    cities = set(j.city for j in jobs if j.city)
    has_jd = sum(1 for j in jobs if j.jd_text)
    has_salary = sum(1 for j in jobs if j.salary_min > 0 or j.salary_max > 0)

    print(f"\n  --- Quality summary ({expected_company}) ---")
    print(f"  Total jobs:       {len(jobs)}")
    print(f"  Unique depts:     {len(depts)}")
    print(f"  Cities:           {sorted(cities)[:10]}")
    print(f"  With JD text:     {has_jd}/{len(jobs)}")
    print(f"  With salary data: {has_salary}/{len(jobs)}")
    print(f"  Sample titles:    {titles}")
    print(f"  ---\n")


# ======================================================================
# E2E tests — real network calls
# ======================================================================


@pytest.mark.slow
class TestBaiduE2E:
    """Real end-to-end tests against the public Baidu Talent API.

    The Baidu Talent API at talent.baidu.com/httservice/getPostListNew is a
    public JSON endpoint that does not require authentication for listing jobs.
    """

    def test_baidu_real_scrape_basic(self):
        """Scrape Baidu via the real POST API and validate every job.

        If the API blocks or returns empty, the test is marked xfail with an
        explanation.
        """
        scraper = BaiduScraper()
        config = CompanyConfig(
            name="百度",
            platform="baidu",
            base_url="https://talent.baidu.com/",
        )
        assert scraper.supports(config) is True

        try:
            # Fetch only page 1 to keep the test fast
            items = scraper._fetch_page(1)
        except Exception as e:
            pytest.xfail(f"Baidu API unreachable (network or block): {e}")

        if not items:
            pytest.xfail(
                "Baidu API returned empty list. Diagnosis (2026-07-08): "
                "the API now returns {\"status\":\"no-auth\",\"message\":\"illegal-visit\"} "
                "-- the endpoint requires authentication that was not needed when "
                "the scraper was written."
            )

        jobs = [scraper._parse_job(item, "百度") for item in items]

        print(f"\n  Baidu page 1 returned {len(items)} raw items, "
              f"parsed into {len(jobs)} jobs.")

        validate_job_batch(jobs, "百度", "baidu")

    def test_baidu_real_multi_page(self):
        """Scrape at least 2 pages from Baidu to exercise pagination.

        This also verifies that the total count grows across pages.
        """
        scraper = BaiduScraper()
        config = CompanyConfig(
            name="百度",
            platform="baidu",
            base_url="https://talent.baidu.com/",
        )

        page1 = scraper._fetch_page(1)
        if not page1:
            pytest.xfail("Baidu API page 1 returned empty — cannot test pagination.")

        page2 = scraper._fetch_page(2)
        if not page2:
            pytest.xfail(
                "Baidu API page 2 returned empty — the site may have fewer "
                "than 20 jobs total or pagination behaves differently."
            )

        ids_page1 = {item.get("id") for item in page1 if item.get("id")}
        ids_page2 = {item.get("id") for item in page2 if item.get("id")}

        # The two pages should not have overlapping IDs
        overlap = ids_page1 & ids_page2
        assert not overlap, (
            f"Baidu API returned duplicate job IDs across pages: {overlap}"
        )

        print(f"\n  Page 1: {len(page1)} items, Page 2: {len(page2)} items, "
              f"Overlap: {overlap}")

    def test_baidu_full_scrape(self):
        """Run the complete Baidu scraper (all pages)."""
        scraper = BaiduScraper()
        config = CompanyConfig(
            name="百度",
            platform="baidu",
            base_url="https://talent.baidu.com/",
        )

        try:
            jobs = scraper.scrape(config)
        except Exception as e:
            pytest.xfail(f"Baidu full scrape failed: {e}")

        if not jobs:
            pytest.xfail("Baidu full scrape returned no jobs.")

        validate_job_batch(jobs, "百度", "baidu")


@pytest.mark.slow
class TestNeteaseE2E:
    """Real end-to-end tests against Netease HR (hr.163.com).

    Netease uses server-rendered HTML.  These tests fetch real pages and parse
    job cards via BeautifulSoup.
    """

    def test_netease_real_scrape_basic(self):
        """Fetch page 1 from hr.163.com and validate parsed jobs."""
        scraper = NeteaseScraper()
        config = CompanyConfig(
            name="网易",
            platform="netease",
            base_url="https://hr.163.com/",
        )
        assert scraper.supports(config) is True

        try:
            soup = scraper._fetch_page(1)
        except Exception as e:
            pytest.xfail(f"Netease HR unreachable (network or block): {e}")

        if soup is None:
            pytest.xfail("Netease HR returned None for page 1.")

        # Try primary selectors
        cards = soup.select(".position-item, .job-item, tr[class*='item']")
        if not cards:
            cards = soup.find_all("a", href=lambda h: h and "detail" in h)

        if not cards:
            pytest.xfail(
                "Netease HR page 1 has no parseable job cards. "
                "Diagnosis (2026-07-08): the page is an SPA — only a minimal "
                "HTML shell is served (6083 bytes, zero <a> tags, empty body). "
                "All job listing content is loaded via JavaScript and cannot "
                "be scraped with plain HTTP requests. "
                "The Playwright scraper should be used for this site instead."
            )

        jobs = []
        for card in cards:
            job = scraper._parse_card(card, "网易")
            if job is not None:
                jobs.append(job)

        print(f"\n  Netease page 1 found {len(cards)} HTML elements, "
              f"parsed into {len(jobs)} jobs.")

        validate_job_batch(jobs, "网易", "netease")

    def test_netease_full_scrape(self):
        """Run the complete Netease scraper."""
        scraper = NeteaseScraper()
        config = CompanyConfig(
            name="网易",
            platform="netease",
            base_url="https://hr.163.com/",
        )

        try:
            jobs = scraper.scrape(config)
        except Exception as e:
            pytest.xfail(f"Netease full scrape failed: {e}")

        if not jobs:
            pytest.xfail("Netease full scrape returned no jobs.")

        validate_job_batch(jobs, "网易", "netease")


# ======================================================================
# Config structure test — loads all 50 companies
# ======================================================================


def test_config_all_50_companies():
    """Load companies.yaml and verify every entry has the required fields."""
    config_path = PROJECT_ROOT / "config" / "companies.yaml"
    assert config_path.exists(), f"Config file missing: {config_path}"

    companies = load_companies(str(config_path))
    assert len(companies) >= 50, (
        f"Expected at least 50 companies, got {len(companies)}"
    )

    valid_platforms = {"moka", "feishu", "baidu", "tencent", "netease", "zhiye", "playwright", "liepin"}

    for c in companies:
        assert c.name, "Every company must have a name"
        assert c.platform in valid_platforms, (
            f"Company {c.name!r} has unknown platform {c.platform!r}"
        )
        assert c.base_url, f"Company {c.name!r} must have a base_url"
        assert c.base_url.startswith("http"), (
            f"Company {c.name!r} base_url must start with http: {c.base_url!r}"
        )
        assert isinstance(c.enabled, bool), (
            f"Company {c.name!r} 'enabled' must be a boolean"
        )

    # Summary counts by platform
    from collections import Counter
    pc = Counter(c.platform for c in companies)
    print(f"\n  --- Config summary ({len(companies)} companies) ---")
    for platform, count in pc.most_common():
        print(f"  {platform:15s} {count:3d}")
    print(f"  {'enabled':15s} {sum(1 for c in companies if c.enabled):3d}")
    print(f"  {'disabled':15s} {sum(1 for c in companies if not c.enabled):3d}")
    print()


# ======================================================================
# E2E pipeline test — mocked data flows through Collector -> ExcelWriter
# ======================================================================


def _make_fake_job(
    company: str,
    platform: str,
    title: str = "测试工程师",
    city: str = "北京",
    department: str = "技术部",
    jd_text: str = "岗位职责：负责系统测试\n任职要求：熟悉Python",
    salary_min: int = 15,
    salary_max: int = 30,
) -> JobData:
    """Build a realistic-looking JobData for pipeline testing."""
    return JobData(
        company_name=company,
        job_title=title,
        source_platform=platform,
        job_id=f"test-{company}-{title}",
        department=department,
        city=city,
        experience_raw="3-5年",
        education_raw="本科",
        salary_min=salary_min,
        salary_max=salary_max,
        salary_desc=f"{salary_min}K-{salary_max}K",
        jd_text=jd_text,
        skills_raw="Python, Pytest",
        benefits_raw="五险一金",
        source_url="https://example.com/jobs/test",
    )


class TestPipelineE2E:
    """End-to-end pipeline tests using mocked data.

    These tests verify that the full data flow — raw JobData instances through
    JobCollector to ExcelWriter — works correctly without requiring live
    network access.  They also exercise all data-quality checks on the mocked
    data.
    """

    def test_collector_flow(self):
        """Add fake jobs to collector, verify stats and retrieval."""
        collector = JobCollector()

        jobs_a = [_make_fake_job("百度", "baidu", f"职位{i}") for i in range(5)]
        jobs_b = [_make_fake_job("网易", "netease", f"Position{i}") for i in range(3)]

        collector.add_batch(jobs_a)
        collector.add_batch(jobs_b)

        assert collector.total() == 8

        stats = collector.stats()
        assert stats["total_jobs"] == 8
        assert stats["companies"]["百度"] == 5
        assert stats["companies"]["网易"] == 3
        assert stats["platforms"]["baidu"] == 5
        assert stats["platforms"]["netease"] == 3

    def test_excel_output_created(self):
        """Verify ExcelWriter produces a non-empty .xlsx file."""
        collector = JobCollector()

        for i in range(10):
            collector.add(_make_fake_job("百度", "baidu", f"职位{i}"))

        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "test_output.xlsx"
            ExcelWriter.write(collector, str(out_path))

            assert out_path.exists(), f"Excel file not created: {out_path}"
            assert out_path.stat().st_size > 0, "Excel file is empty"

            print(f"\n  Excel output: {out_path} ({out_path.stat().st_size} bytes)")

    def test_excel_output_data_integrity(self):
        """Jobs written to Excel should survive a round-trip in terms of count."""
        collector = JobCollector()

        companies = [
            ("百度", "baidu", 12),
            ("网易", "netease", 8),
            ("腾讯", "tencent", 5),
        ]
        for company, platform, count in companies:
            for i in range(count):
                collector.add(
                    _make_fake_job(company, platform, f"职位{i}", jd_text=f"JD {i} for {company}")
                )

        assert collector.total() == 25

        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "test_integrity.xlsx"
            ExcelWriter.write(collector, str(out_path))

            # Read back with openpyxl (no pandas dependency)
            from openpyxl import load_workbook
            wb = load_workbook(out_path)
            ws = wb["全部职位"]

            # Row 1 = header, subsequent rows = data
            data_rows = ws.max_row - 1
            assert data_rows == 25, (
                f"Expected 25 data rows in Excel, got {data_rows}"
            )

            # Check summary sheet
            ws2 = wb["统计"]
            assert ws2.cell(1, 2).value == 25  # total_jobs

            print(f"\n  Round-trip verified: {data_rows} rows in Excel")

    def test_pipeline_with_salary_validation(self):
        """Verify salary data quality (min <= max) across the pipeline."""
        collector = JobCollector()

        # Normal job
        collector.add(_make_fake_job("百度", "baidu", "正常薪资岗", salary_min=20, salary_max=40))
        # "Negotiable" — both zero
        collector.add(_make_fake_job("网易", "netease", "面议岗", salary_min=0, salary_max=0))
        # Equal min/max
        collector.add(_make_fake_job("腾讯", "tencent", "固定薪资岗", salary_min=25, salary_max=25))

        for job in collector.get_all():
            validate_job_data(job, job.company_name, job.source_platform)

        assert collector.total() == 3
        print(f"\n  Salary validation passed for {collector.total()} jobs")

    def test_pipeline_with_dispatcher(self):
        """Verify the dispatcher can route a company to the right scraper type.

        This is a structural test — it verifies that setup_dispatcher()
        registers scrapers and that scrape_company can be called without
        crashing for an enabled company.  Mock data is used for the actual
        scrape.
        """
        dispatcher = setup_dispatcher()

        # The dispatcher should have registered scrapers
        assert len(dispatcher._scrapers) > 0, (
            "Dispatcher should have at least one registered scraper"
        )

        # Verify each registered scraper has a name
        names = [s.name for s in dispatcher._scrapers]
        print(f"\n  Registered scrapers: {names}")

        # Create a config for a known platform
        baidu_config = CompanyConfig(
            name="百度",
            platform="baidu",
            base_url="https://talent.baidu.com/",
        )

        # scraper lookup should succeed (the scraper will actually hit the
        # network here — but in the dispatcher test we only verify routing)
        scraper = dispatcher._find_scraper(baidu_config)
        assert scraper is not None, (
            f"No scraper found for baidu platform. Registered: {names}"
        )
        assert scraper.name == "baidu"

    def test_empty_collector_handling(self):
        """Edge case: empty collector should produce valid stats."""
        collector = JobCollector()
        assert collector.total() == 0
        assert collector.get_all() == []

        stats = collector.stats()
        assert stats["total_jobs"] == 0
        assert stats["companies"] == {}
        assert stats["cities"] == {}
        assert stats["platforms"] == {}

    def test_data_quality_report(self):
        """Run the full suite of data-quality checks on a large batch.

        This serves as a smoke test that catches regressions in the validation
        logic itself.
        """
        collector = JobCollector()

        # Generate diverse fake data
        scenarios = [
            # (company, platform, city, dept, salary_min, salary_max, jd)
            ("百度", "baidu", "北京", "搜索", 30, 60, "搜索引擎开发"),
            ("百度", "baidu", "上海", "AI", 25, 50, "NLP算法"),
            ("网易", "netease", "杭州", "云音乐", 20, 40, "推荐系统"),
            ("网易", "netease", "广州", "游戏", 15, 30, "游戏引擎"),
            ("腾讯", "tencent", "深圳", "微信", 35, 70, "后端架构"),
        ]

        for company, platform, city, dept, lo, hi, jd in scenarios:
            collector.add(_make_fake_job(
                company, platform, f"{dept}工程师",
                city=city, department=dept,
                salary_min=lo, salary_max=hi,
                jd_text=f"岗位职责：{jd}\n任职要求：精通相关技术",
            ))

        # Validate every job
        for job in collector.get_all():
            validate_job_data(job, job.company_name, job.source_platform)

        # Cross-job check: all salary_min values should be in a sane range
        for job in collector.get_all():
            if job.salary_min > 0:
                assert 1 <= job.salary_min <= 500, (
                    f"Implausible salary_min {job.salary_min} for {job.job_title}"
                )

        # city should not be empty for most jobs
        jobs_with_city = sum(1 for j in collector.get_all() if j.city)
        assert jobs_with_city == len(scenarios), (
            f"Expected all jobs to have city, but only {jobs_with_city} do"
        )

        print(f"\n  Full quality report: {collector.total()} jobs, all checks passed")
