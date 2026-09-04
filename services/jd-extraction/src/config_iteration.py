from __future__ import annotations

import json
import os
import re
import shutil
from hashlib import sha256
from collections import defaultdict
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .deepseek_client import DeepSeekClient
from .normalizer import (
    _normalization_key,
    load_normalization_map,
    lookup_skill_mapping,
    skill_mapping_candidates,
)


DEFAULT_POLICY = {
    "batch_size": 20,
    "min_document_count": 1,
    "min_create_new_document_count": 1,
    "max_evidence_samples": 3,
    "max_candidates_per_review": 500,
    "semantic_request_batch_size": 50,
    "pending_review_dir": "config/pending_review",
    "applied_review_dir": "config/pending_review/applied",
    "candidate_pool_path": "config/normalization_candidate_pool.json",
    "decision_ledger_path": "config/normalization_decision_ledger.json",
}
REVIEW_SHEET = "待审查"
REVIEW_COMPLETE_CELL = "B2"
REQUIRED_HEADERS = (
    "candidate_id", "source_name", "document_count", "item_type", "evidence_samples",
    "suggestion_action", "suggested_anchor_alias", "suggested_skill_id", "suggested_canonical_name",
    "suggested_category_code", "suggested_subcategory_code", "suggestion_reason", "enabled", "reviewer_note",
)
ALLOWED_CATEGORY_CODES = {
    "programming_language", "framework", "library", "database", "tool",
    "platform", "methodology", "domain_knowledge", "other",
}
PROPOSAL_VERSION = 4
DECISION_LEDGER_VERSION = 1
RESOLVED_ACTIONS = {"alias_existing", "alias_candidate", "create_new"}
DISPOSITION_ACTIONS = {
    "generic_unresolved",
    "ambiguous",
    "excluded_non_skill",
    "pending_review",
}
REVIEW_ACTIONS = RESOLVED_ACTIONS | DISPOSITION_ACTIONS
_ILLEGAL_EXCEL_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
SEMANTIC_CHECKPOINT_VERSION = 1
CATALOG_ENTRIES_PER_CANDIDATE = 12
MAX_RELEVANT_CATALOG_ENTRIES = 200


def load_iteration_policy(path: str | None = None) -> dict[str, Any]:
    policy = deepcopy(DEFAULT_POLICY)
    if path is None:
        return policy
    file_path = Path(path)
    if not file_path.exists():
        return policy
    payload = yaml.safe_load(file_path.read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        raise ValueError("Iteration policy must be a YAML object.")
    policy.update(payload)
    if not isinstance(policy["batch_size"], int) or policy["batch_size"] < 1:
        raise ValueError("Iteration policy batch_size must be a positive integer.")
    if not isinstance(policy["min_document_count"], int) or policy["min_document_count"] < 1:
        raise ValueError("Iteration policy min_document_count must be a positive integer.")
    if (
        not isinstance(policy["min_create_new_document_count"], int)
        or policy["min_create_new_document_count"] < policy["min_document_count"]
    ):
        raise ValueError(
            "Iteration policy min_create_new_document_count must be an integer "
            "not smaller than min_document_count."
        )
    if not isinstance(policy["max_candidates_per_review"], int) or policy["max_candidates_per_review"] < 1:
        raise ValueError("Iteration policy max_candidates_per_review must be a positive integer.")
    if (
        not isinstance(policy["semantic_request_batch_size"], int)
        or policy["semantic_request_batch_size"] < 1
        or policy["semantic_request_batch_size"] > policy["max_candidates_per_review"]
    ):
        raise ValueError(
            "Iteration policy semantic_request_batch_size must be a positive integer "
            "not greater than max_candidates_per_review."
        )
    return policy


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    decoder = json.JSONDecoder(strict=False)
    records: list[dict[str, Any]] = []
    offset = 0
    while offset < len(text):
        while offset < len(text) and text[offset].isspace():
            offset += 1
        if offset >= len(text):
            break
        record, offset = decoder.raw_decode(text, offset)
        if not isinstance(record, dict):
            raise ValueError(f"JSON stream record in {path} must be an object.")
        records.append(record)
    return records


def _candidate_id(run_id: str, source_name: str, item_type: str) -> str:
    del run_id
    material = f"{_normalization_key(source_name)}\x1f{item_type}".encode("utf-8")
    return f"candidate_{sha256(material).hexdigest()[:24]}"


def _candidate_key(source_name: str, item_type: str) -> str:
    material = f"{' '.join(source_name.split())}\x1f{item_type}".encode("utf-8")
    return material.decode("utf-8")[:120]


def _load_candidate_pool(path: str | Path) -> dict[str, Any]:
    pool_path = Path(path)
    if not pool_path.exists():
        return {"version": 2, "candidates": {}}
    payload = json.loads(pool_path.read_text(encoding="utf-8"))
    if (
        not isinstance(payload, dict)
        or payload.get("version") != 2
        or not isinstance(payload.get("candidates"), dict)
    ):
        raise ValueError("Normalization candidate pool must be a version 2 object.")
    return payload


def _input_identity(input_path: str) -> str:
    normalized = str(Path(input_path)).replace("\\", "/").casefold()
    return normalized[:80]


def _write_candidate_pool(path: str | Path, payload: dict[str, Any]) -> None:
    pool_path = Path(path)
    pool_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = pool_path.with_suffix(pool_path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp_path, pool_path)


def load_decision_ledger(path: str | Path) -> dict[str, Any]:
    ledger_path = Path(path)
    if not ledger_path.exists():
        return {"version": DECISION_LEDGER_VERSION, "decisions": {}}
    payload = json.loads(ledger_path.read_text(encoding="utf-8"))
    if (
        not isinstance(payload, dict)
        or payload.get("version") != DECISION_LEDGER_VERSION
        or not isinstance(payload.get("decisions"), dict)
    ):
        raise ValueError("Normalization decision ledger must be a version 1 object.")
    return payload


def _write_decision_ledger(path: str | Path, payload: dict[str, Any]) -> None:
    ledger_path = Path(path)
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = ledger_path.with_suffix(ledger_path.suffix + ".tmp")
    tmp_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(tmp_path, ledger_path)


def lookup_semantic_decision(
    ledger: dict[str, Any],
    source_name: str,
    item_type: str,
) -> dict[str, Any] | None:
    decision = ledger.get("decisions", {}).get(
        _candidate_id("", source_name, item_type)
    )
    return decision if isinstance(decision, dict) else None


def _excel_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    return _ILLEGAL_EXCEL_CHARACTERS.sub("", value)


def collect_unresolved_candidates(
    run_dir: str | Path,
    normalization_path: str,
    min_document_count: int,
    candidate_pool_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    run_path = Path(run_dir)
    manifest = json.loads((run_path / "manifest.json").read_text(encoding="utf-8"))
    input_identity = _input_identity(manifest["input_path"])
    normalized = _read_jsonl(run_path / "final" / "normalized_annotations.jsonl")
    annotations = _read_jsonl(run_path / "final" / "annotations.jsonl")
    norm_map = load_normalization_map(normalization_path)
    evidence_by_document: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for annotation in annotations:
        by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for requirement in annotation.get("requirements", []):
            if requirement.get("kind") != "skill":
                continue
            evidence = requirement.get("evidence") or {}
            for item in requirement.get("items", []):
                if isinstance(item.get("name"), str):
                    by_name[item["name"]].append(
                        {
                            "item_type": item.get("item_type"),
                            "quote": evidence.get("quote"),
                            "source_id": evidence.get("source_id"),
                        }
                    )
        evidence_by_document[str(annotation.get("document_id"))] = by_name

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for record in normalized:
        document_id = str(record.get("document_id"))
        for requirement in record.get("normalized_requirements", []):
            for skill in requirement.get("skills", []):
                source_name = skill.get("source_name")
                if (
                    skill.get("identity_resolution_status") != "unresolved"
                    or not isinstance(source_name, str)
                ):
                    continue
                evidence_items = evidence_by_document.get(document_id, {}).get(
                    source_name, []
                )
                for item_type in sorted(
                    {
                        evidence.get("item_type")
                        for evidence in evidence_items
                        if isinstance(evidence.get("item_type"), str)
                    }
                ):
                    key = (source_name, item_type)
                    entry = grouped.setdefault(
                        key,
                        {
                            "source_name": source_name,
                            "item_type": item_type,
                            "document_ids": set(),
                            "evidence": [],
                        },
                    )
                    observation_id = f"{input_identity}/{document_id}"
                    entry["document_ids"].add(observation_id)
                    for evidence in evidence_items:
                        if (
                            evidence.get("item_type") == item_type
                            and evidence not in entry["evidence"]
                        ):
                            entry["evidence"].append(
                                {"jd_id": observation_id, **evidence}
                            )

    if candidate_pool_path is not None:
        pool = _load_candidate_pool(candidate_pool_path)
        pooled = pool["candidates"]
        input_observation_prefix = f"{input_identity}/"
        for key, stored in pooled.items():
            stored["document_ids"] = [
                observation_id
                for observation_id in stored.get("document_ids", [])
                if not str(observation_id).startswith(input_observation_prefix)
            ]
            stored["evidence_samples"] = [
                evidence
                for evidence in stored.get("evidence_samples", [])
                if not str(evidence.get("jd_id", "")).startswith(input_observation_prefix)
            ]
        for entry in grouped.values():
            key = _candidate_key(entry["source_name"], entry["item_type"])
            stored = pooled.setdefault(
                key,
                {
                    "source_name": entry["source_name"],
                    "item_type": entry["item_type"],
                    "document_ids": [],
                    "evidence_samples": [],
                    "last_reviewed_document_count": 0,
                },
            )
            if stored.get("source_name") != entry["source_name"] or stored.get("item_type") != entry["item_type"]:
                raise ValueError("Normalization candidate pool contains a conflicting candidate key.")
            stored["document_ids"] = sorted(set(stored.get("document_ids", [])) | entry["document_ids"])
            known_evidence = {
                (item.get("jd_id"), item.get("source_id"), item.get("quote"))
                for item in stored.get("evidence_samples", [])
                if isinstance(item, dict)
            }
            for evidence in entry["evidence"]:
                evidence_key = (evidence.get("jd_id"), evidence.get("source_id"), evidence.get("quote"))
                if evidence_key not in known_evidence:
                    stored.setdefault("evidence_samples", []).append(evidence)
                    known_evidence.add(evidence_key)
        empty_keys = [key for key, stored in pooled.items() if not stored.get("document_ids")]
        for key in empty_keys:
            del pooled[key]
        resolved_keys = {
            key
            for key, entry in pooled.items()
            if lookup_skill_mapping(
                norm_map, str(entry.get("source_name", "")), str(entry.get("item_type", ""))
            ) is not None
        }
        for key in resolved_keys:
            del pooled[key]
        _write_candidate_pool(candidate_pool_path, pool)
        grouped = {
            (entry["source_name"], entry["item_type"]): {
                "source_name": entry["source_name"],
                "item_type": entry["item_type"],
                "document_ids": set(entry.get("document_ids", [])),
                "evidence": entry.get("evidence_samples", []),
                "last_reviewed_document_count": entry.get("last_reviewed_document_count", 0),
            }
            for entry in pooled.values()
        }

    candidates = []
    for entry in grouped.values():
        if len(entry["document_ids"]) < min_document_count:
            continue
        if len(entry["document_ids"]) <= entry.get("last_reviewed_document_count", 0):
            continue
        if lookup_skill_mapping(norm_map, entry["source_name"], entry["item_type"]) is not None:
            continue
        candidates.append(
            {
                "candidate_id": _candidate_id(manifest["run_id"], entry["source_name"], entry["item_type"]),
                "source_name": entry["source_name"],
                "document_count": len(entry["document_ids"]),
                "item_type": entry["item_type"],
                "evidence_samples": entry["evidence"],
            }
        )
    return sorted(candidates, key=lambda item: (-item["document_count"], item["source_name"]))


def _taxonomy_prompt() -> dict[str, Any]:
    return {
        "normalization_layer_responsibility": [
            "归一化层负责标准技能名称、标准技能 ID、受控 category/subcategory。",
            "归一化层可以扩展 config/normalization_map.yaml，但不得改写抽取层原文事实。",
            "未被人工审查同意前，建议只进入待审查表，不直接进入正式配置。",
        ],
        "item_type_definitions": {
            "programming_language": "编程、脚本或查询语言，例如 Python、Java、C++、SQL、Shell、Go。",
            "framework": "提供应用骨架、运行范式或完整开发框架，例如 Django、Spring、React、PyTorch、FastAPI。",
            "library": "被程序调用的库、SDK 或较轻量代码依赖，例如 NumPy、Pandas、cublas、cudnn。",
            "database": "数据库、向量库、缓存、搜索或消息存储系统，例如 MySQL、Redis、Milvus、Elasticsearch。",
            "tool": "具体开发、部署、协作、标注或 AI 辅助工具，例如 Git、Docker、Postman、Cursor。",
            "platform": "云平台、业务平台、操作系统、硬件计算平台或运行环境，例如 AWS、Linux、CUDA、Hugging Face。",
            "methodology": "算法、模型方法、训练方法、分析方法、工程策略，例如 RAG、LoRA、量化、Prompt Engineering。",
            "domain_knowledge": "理论、技术领域、行业或业务知识，例如 机器学习、NLP、数据结构、金融风控。",
            "other": "明确是技能实体但无法稳定归入以上类型。",
        },
        "boundary_rules": [
            "SQL 是 programming_language，MySQL 是 database。",
            "PyTorch 按当前 Schema 标为 framework。",
            "RAG、微调、量化、Prompt Engineering 是 methodology。",
            "Linux 是 platform，Shell 是 programming_language。",
            "Git 是 tool。",
            "CUDA 作为计算平台或运行环境时是 platform。",
            "Node.js 是 JavaScript 运行环境，归入 platform，不得归入 programming_language。",
            "Hugging Face 作为模型与工具生态平台时归入 platform。",
            "TensorRT-LLM 作为 LLM 推理框架时归入 framework。",
            "自然语言处理（NLP）、视觉语言模型（VLM）和多模态模型是可复用的 domain_knowledge 实体，不应仅因名称宽泛而 reject。",
            "上下文工程在证据明确指向 LLM 上下文设计方法时是 methodology 实体。",
            "普通语言模型不等于大语言模型；只有原文明示 LLM 或大语言模型时才能映射到大语言模型。",
            "不要把上下位概念映射为同一个标准项，例如 AI 不等于 LLM。",
            "INT4、INT8、FP16、BF16 等精度或数据格式是量化/训练方法的参数，不是独立技能实体。",
            "英文读写、沟通能力、代码实现能力等候选人能力短语不是稳定技能实体，应在抽取层进入软技能或其他要求。",
            "模型、协议、API、框架等具名技术不得使用 other；必须按实体本身选择最接近的合法类型。",
            "搭建、实现、优化、控制等动作短语不是标准技能名；只有能归约为行业稳定术语时才可新增。",
        ],
        "skill_id_rules": [
            "skill_id 使用大写 ASCII、数字和下划线。",
            "推荐前缀：LANG、FRAMEWORK、LIBRARY、DATABASE、TOOL、PLATFORM、METHOD、KNOWLEDGE、OTHER。",
            "同一个标准实体的不同别名必须使用完全相同的 skill_id、canonical_name、category_code、subcategory_code。",
            "不同标准实体不得复用同一个 skill_id。",
        ],
        "canonical_name_rules": [
            "优先采用项目或厂商官方大小写，例如 Go、CUDA、OpenCL、cuBLAS、cuDNN、pgvector。",
            "同义短写与完整名称不得各自创建标准实体；选择一个 canonical_name，其余候选作为别名。",
            "同一批候选中若一个候选应作为另一个新实体的别名，使用 alias_candidate，不要 reject 或重复 create_new。",
            "Triton、对齐等可能指向多个实体或领域的短名，必须结合证据消歧为完整标准名；证据不足时 reject。",
            "canonical_name 必须是可跨 JD 复用的稳定实体，不能只是删除候选词中的动作后得到一个不自然的中英文混合短语。",
        ],
    }


def _catalog_terms(value: str) -> set[str]:
    normalized = _normalization_key(value)
    words = set(re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]", normalized))
    compact = "".join(character for character in normalized if character.isalnum())
    if len(compact) >= 3:
        words.update(compact[index:index + 3] for index in range(len(compact) - 2))
    return words


def _select_relevant_catalog(
    candidates: list[dict[str, Any]],
    normalization_map: dict[str, Any],
) -> list[dict[str, Any]]:
    catalog = [
        {"alias": alias, **mapping}
        for alias, mapping in normalization_map["skills"].items()
        if isinstance(mapping, dict)
    ]
    selected: dict[str, tuple[int, dict[str, Any]]] = {}
    for candidate in candidates:
        source_name = str(candidate["source_name"])
        source_key = _normalization_key(source_name)
        source_terms = _catalog_terms(source_name)
        ranked: list[tuple[int, str, dict[str, Any]]] = []
        for entry in catalog:
            alias = str(entry["alias"])
            canonical_name = str(entry.get("canonical_name", ""))
            alias_key = _normalization_key(alias)
            canonical_key = _normalization_key(canonical_name)
            target_terms = _catalog_terms(f"{alias} {canonical_name}")
            overlap = len(source_terms & target_terms)
            score = overlap * 20
            if source_key and (
                source_key in alias_key
                or alias_key in source_key
                or source_key in canonical_key
                or canonical_key in source_key
            ):
                score += 100
            if entry.get("category_code") == candidate.get("item_type"):
                score += 2
            if score:
                ranked.append((score, alias.casefold(), entry))
        ranked.sort(key=lambda item: (-item[0], item[1]))
        for score, _, entry in ranked[:CATALOG_ENTRIES_PER_CANDIDATE]:
            alias = str(entry["alias"])
            previous = selected.get(alias)
            if previous is None or score > previous[0]:
                selected[alias] = (score, entry)
    ranked_selected = sorted(
        selected.values(),
        key=lambda item: (-item[0], str(item[1]["alias"]).casefold()),
    )
    return [
        entry
        for _, entry in ranked_selected[:MAX_RELEVANT_CATALOG_ENTRIES]
    ]


def _build_semantic_update_prompt(
    candidates: list[dict[str, Any]],
    normalization_map: dict[str, Any],
) -> tuple[str, str]:
    catalog = _select_relevant_catalog(candidates, normalization_map)
    system = """你是归一化词表扩展审查助手。你要根据候选词、证据和当前字段标准，为每个 candidate 给出一个可供人工审查的建议。

只能输出 JSON object，格式为：
{"suggestions":[{"candidate_id":str,"action":"alias_existing"|"alias_candidate"|"create_new"|"generic_unresolved"|"ambiguous"|"excluded_non_skill"|"pending_review","anchor_alias":str|null,"mapping":{"skill_id":str,"canonical_name":str,"category_code":str,"subcategory_code":str}|null,"reason":str}]}

规则：
1. 如果候选词与 existing_catalog 中某个 alias 是完全相同技能实体，使用 action=alias_existing，并填写 anchor_alias；mapping 为 null。candidate.item_type 只是抽取层粗类型，即使与正式 category_code 不同也必须以 existing_catalog 为准。
2. 如果候选词是值得纳入词表的新标准实体，使用 action=create_new，并填写完整 mapping。
3. 如果候选词是同一批某个 create_new 候选的别名，使用 action=alias_candidate，anchor_alias 填目标候选的 source_name；mapping 为 null。目标候选必须同为本批候选且 action=create_new；抽取层 item_type 不同不妨碍别名关系。
4. create_new 的 mapping.category_code 必须依据 taxonomy_standard 和证据独立判断；candidate.item_type 仅供参考，不得把明显错误的粗类型写入正式词表。
5. 同一实体不得在本批创建多个 skill_id，例如 Go/Golang、量化/模型量化应为 create_new + alias_candidate。
6. 如果多个 candidate 的 source_name 归一化后完全相同、只是 candidate.item_type 不同，只选择证据最明确的一行 create_new；其余同名行使用 pending_review，不能 alias_candidate 到同名候选。
7. canonical_name 必须采用官方或行业通用大小写，不得输出 cublas、cudnn、PGVector 这类非标准写法。
8. 泛化程度过高但仍可能是能力领域时使用 generic_unresolved；缩写或上下文不足导致无法唯一消歧时使用 ambiguous；明确不是技能实体时使用 excluded_non_skill；其余暂无法判定时使用 pending_review。
9. 不要因为相似、上下位、共同出现或模糊缩写把候选映射到已有 alias。
10. INT4、INT8、FP16、BF16 等精度标签只能作为量化或训练方法的参数，不得 create_new。
11. 对歧义短名必须依据 evidence_samples 消歧并给出完整 canonical_name；不能消歧时 reject。
12. 英文读写、沟通、代码实现等能力短语，以及搭建、实现、优化等任务动作，不得 create_new 为技能实体。
13. candidate.item_type=other 时，只有确实无法归入合法类型的稳定技术实体才可 create_new；具名模型、协议、API、框架不得沿用 other。
14. Node.js 是 JavaScript 运行环境而非编程语言；普通语言模型不得 alias 到大语言模型。自然语言处理、VLM、多模态模型以及证据明确的上下文工程不得仅以“宽泛”或“非稳定术语”为由 reject。
15. create_new 可用于单份或多份 JD 中 Evidence 已明确确认的稳定实体；频次只决定审查优先级，不得替代 identity 审查。
16. 每条建议都必须给出与 evidence_samples 和 taxonomy_standard 一致的非空 reason。
17. 不要输出除上述 JSON 之外的任何文字。"""
    user = json.dumps(
        {
            "taxonomy_standard": _taxonomy_prompt(),
            "candidates": candidates,
            "existing_catalog": catalog,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    return system, user


def _is_non_empty_string(value: Any) -> bool:
    return isinstance(value, str) and bool(value.strip())


def _validate_suggested_mapping(mapping: Any) -> dict[str, str]:
    if not isinstance(mapping, dict):
        raise ValueError("Suggested create_new mapping must be an object.")
    required = ("skill_id", "canonical_name", "category_code", "subcategory_code")
    normalized: dict[str, str] = {}
    for key in required:
        value = mapping.get(key)
        if not _is_non_empty_string(value):
            raise ValueError(f"Suggested create_new mapping field {key!r} must be a non-empty string.")
        normalized[key] = str(value).strip()
    if normalized["category_code"] not in ALLOWED_CATEGORY_CODES:
        raise ValueError("Suggested create_new mapping category_code is not a legal taxonomy category.")
    return normalized


def _deterministic_semantic_suggestions(
    candidates: list[dict[str, Any]],
    normalization_map: dict[str, Any],
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    suggestions: dict[str, dict[str, Any]] = {}
    unresolved: list[dict[str, Any]] = []
    catalog = normalization_map["skills"]
    for candidate in candidates:
        mappings = skill_mapping_candidates(
            normalization_map,
            str(candidate["source_name"]),
        )
        if len(mappings) != 1:
            unresolved.append(candidate)
            continue
        mapping = mappings[0]
        matching_aliases = [
            alias
            for alias, candidate_mapping in catalog.items()
            if isinstance(candidate_mapping, dict) and candidate_mapping == mapping
        ]
        source_key = _normalization_key(str(candidate["source_name"]))
        anchor_alias = next(
            (
                alias
                for alias in matching_aliases
                if _normalization_key(alias) == source_key
            ),
            None,
        )
        if anchor_alias is None:
            canonical_name = str(mapping.get("canonical_name", ""))
            anchor_alias = next(
                (
                    alias
                    for alias in matching_aliases
                    if _normalization_key(alias) == _normalization_key(canonical_name)
                ),
                matching_aliases[0] if matching_aliases else None,
            )
        if anchor_alias is None:
            unresolved.append(candidate)
            continue
        suggestions[candidate["candidate_id"]] = {
            "action": "alias_existing",
            "anchor_alias": anchor_alias,
            "mapping": None,
            "reason": "名称已唯一命中现有技能身份，仅来源 item_type 与正式分类不一致。",
        }
    return suggestions, unresolved


def request_semantic_suggestions(candidates: list[dict[str, Any]], normalization_path: str, model: str) -> dict[str, dict[str, Any]]:
    if not candidates:
        return {}
    normalization_map = load_normalization_map(normalization_path)
    result, unresolved = _deterministic_semantic_suggestions(
        candidates,
        normalization_map,
    )
    if not unresolved:
        return result
    system, user = _build_semantic_update_prompt(unresolved, normalization_map)
    relevant_catalog_count = len(
        json.loads(user)["existing_catalog"]
    )
    print(
        f"Candidate preprocessing completed: "
        f"{len(candidates) - len(unresolved)} deterministic, "
        f"{len(unresolved)} sent to {model} with "
        f"{relevant_catalog_count} relevant catalog aliases.",
        flush=True,
    )
    response = DeepSeekClient(model=model).extract(system, user).data
    print(
        f"Candidate model response received from {model}.",
        flush=True,
    )
    suggestions = response.get("suggestions") if isinstance(response, dict) else None
    if not isinstance(suggestions, list):
        raise ValueError("DeepSeek semantic iteration response must contain a suggestions list.")
    allowed = {candidate["candidate_id"]: candidate for candidate in unresolved}
    candidate_id_by_source = {
        (candidate["source_name"], candidate["item_type"]): candidate["candidate_id"]
        for candidate in unresolved
    }
    for item in suggestions:
        if not isinstance(item, dict) or item.get("candidate_id") not in allowed:
            continue
        reason = item.get("reason")
        if not _is_non_empty_string(reason):
            continue
        candidate = allowed[item["candidate_id"]]
        action = item.get("action")
        if action == "alias_existing":
            anchor_alias = item.get("anchor_alias")
            mapping = normalization_map["skills"].get(anchor_alias) if isinstance(anchor_alias, str) else None
            if not isinstance(mapping, dict):
                continue
            result[item["candidate_id"]] = {
                "action": "alias_existing",
                "anchor_alias": anchor_alias,
                "mapping": None,
                "reason": str(reason).strip(),
            }
            continue
        if action == "create_new":
            try:
                mapping = _validate_suggested_mapping(item.get("mapping"))
            except ValueError:
                continue
            result[item["candidate_id"]] = {
                "action": "create_new",
                "anchor_alias": None,
                "mapping": mapping,
                "reason": str(reason).strip(),
            }
            continue
        if action not in DISPOSITION_ACTIONS:
            continue
        result[item["candidate_id"]] = {
            "action": action,
            "anchor_alias": None,
            "mapping": None,
            "reason": str(reason).strip(),
        }
    for item in suggestions:
        if not isinstance(item, dict) or item.get("candidate_id") not in allowed:
            continue
        if item.get("action") != "alias_candidate":
            continue
        reason = item.get("reason")
        if not _is_non_empty_string(reason):
            continue
        candidate = allowed[item["candidate_id"]]
        anchor_alias = item.get("anchor_alias")
        if isinstance(anchor_alias, str) and _normalization_key(candidate["source_name"]) == _normalization_key(anchor_alias):
            result[item["candidate_id"]] = {
                "action": "pending_review",
                "anchor_alias": None,
                "mapping": None,
                "reason": "与同批已选候选使用相同归一化名称，正式配置无需重复 alias。",
            }
            continue
        anchor_candidate_id = next(
            (
                candidate_id
                for (source_name, _), candidate_id in candidate_id_by_source.items()
                if source_name == anchor_alias
            ),
            None,
        )
        anchor_suggestion = result.get(anchor_candidate_id or "")
        if not isinstance(anchor_suggestion, dict) or anchor_suggestion.get("action") != "create_new":
            continue
        result[item["candidate_id"]] = {
            "action": "alias_candidate",
            "anchor_alias": anchor_alias,
            "mapping": None,
            "reason": str(reason).strip(),
        }
    return result


def _semantic_checkpoint_fingerprint(
    candidates: list[dict[str, Any]],
    normalization_path: str,
    model: str,
) -> str:
    normalization_bytes = Path(normalization_path).read_bytes()
    payload = {
        "model": model,
        "normalization_sha256": sha256(normalization_bytes).hexdigest(),
        "candidates": candidates,
    }
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return sha256(encoded).hexdigest()


def _load_semantic_checkpoint(
    path: Path,
    fingerprint: str,
) -> tuple[set[str], dict[str, dict[str, Any]]]:
    if not path.exists():
        return set(), {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if (
        payload.get("version") != SEMANTIC_CHECKPOINT_VERSION
        or payload.get("fingerprint") != fingerprint
        or not isinstance(payload.get("completed_candidate_ids"), list)
        or not isinstance(payload.get("suggestions"), dict)
    ):
        return set(), {}
    completed = {
        candidate_id
        for candidate_id in payload["completed_candidate_ids"]
        if isinstance(candidate_id, str)
    }
    suggestions = {
        candidate_id: suggestion
        for candidate_id, suggestion in payload["suggestions"].items()
        if isinstance(candidate_id, str) and isinstance(suggestion, dict)
    }
    return completed, suggestions


def _write_semantic_checkpoint(
    path: Path,
    fingerprint: str,
    completed_candidate_ids: set[str],
    suggestions: dict[str, dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(
        json.dumps(
            {
                "version": SEMANTIC_CHECKPOINT_VERSION,
                "fingerprint": fingerprint,
                "completed_candidate_ids": sorted(completed_candidate_ids),
                "suggestions": suggestions,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    os.replace(tmp_path, path)


def write_review_workbook(path: str | Path, run_id: str, candidates: list[dict[str, Any]], suggestions: dict[str, dict[str, Any]]) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = REVIEW_SHEET
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "归一化语义更新人工审查"
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet.merge_cells("A1:N1")
    sheet["A2"] = "审查完成（勾选 TRUE 后，本文件才允许在下次抽取启动前应用）"
    sheet[REVIEW_COMPLETE_CELL] = False
    sheet["A3"] = f"run_id: {run_id}；可修改 action、anchor、mapping 和 enabled。create_new 新增标准项；alias_existing/alias_candidate 新增别名。"
    sheet.merge_cells("A3:N3")
    sheet["A3"].alignment = Alignment(wrap_text=True)
    headers = list(REQUIRED_HEADERS)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(5, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, candidate in enumerate(candidates, start=6):
        suggestion = suggestions.get(candidate["candidate_id"], {})
        action = suggestion.get("action", "pending_review")
        if action == "reject":
            action = "pending_review"
        anchor = suggestion.get("anchor_alias")
        mapping = suggestion.get("mapping") if isinstance(suggestion.get("mapping"), dict) else {}
        sheet.cell(row_index, 1, _excel_text(candidate["candidate_id"]))
        sheet.cell(row_index, 2, _excel_text(candidate["source_name"]))
        sheet.cell(row_index, 3, candidate["document_count"])
        sheet.cell(row_index, 4, _excel_text(candidate["item_type"]))
        sheet.cell(row_index, 5, _excel_text("\n".join(
            f"{item.get('jd_id')}: {item.get('quote')}" for item in candidate["evidence_samples"][:3]
        )))
        sheet.cell(row_index, 6, _excel_text(action))
        sheet.cell(row_index, 7, _excel_text(anchor))
        sheet.cell(row_index, 8, _excel_text(mapping.get("skill_id")))
        sheet.cell(row_index, 9, _excel_text(mapping.get("canonical_name")))
        sheet.cell(row_index, 10, _excel_text(mapping.get("category_code")))
        sheet.cell(row_index, 11, _excel_text(mapping.get("subcategory_code")))
        sheet.cell(row_index, 12, _excel_text(suggestion.get("reason", "")))
        sheet.cell(row_index, 13, action != "pending_review")
        sheet.cell(row_index, 14, "")
        for cell in sheet[row_index]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    last_row = max(6, 5 + len(candidates))
    bool_validation = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    sheet.add_data_validation(bool_validation)
    bool_validation.add(sheet[REVIEW_COMPLETE_CELL])
    bool_validation.add(f"M6:M{last_row}")
    action_validation = DataValidation(
        type="list",
        formula1='"alias_existing,alias_candidate,create_new,generic_unresolved,ambiguous,excluded_non_skill,pending_review"',
        allow_blank=False,
    )
    sheet.add_data_validation(action_validation)
    action_validation.add(f"F6:F{last_row}")
    sheet.conditional_formatting.add(
        f"M6:M{last_row}",
        FormulaRule(formula=["M6=FALSE"], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    sheet.freeze_panes = "A6"
    for column, width in {
        "A": 20, "B": 24, "C": 12, "D": 20, "E": 52, "F": 16, "G": 24,
        "H": 24, "I": 28, "J": 20, "K": 24, "L": 42, "M": 12, "N": 28,
    }.items():
        sheet.column_dimensions[column].width = width
    workbook.save(output_path)
    proposal_path = output_path.with_suffix(".proposal.json")
    proposal_path.write_text(
        json.dumps(
            {
                "proposal_version": PROPOSAL_VERSION,
                "run_id": run_id,
                "candidates": candidates,
                "suggestions": suggestions,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return output_path


def propose_from_run(
    run_dir: str | Path,
    normalization_path: str,
    pending_review_dir: str,
    model: str,
    min_document_count: int,
    max_evidence_samples: int = 3,
    candidate_pool_path: str | Path | None = None,
    max_candidates_per_review: int = 500,
    semantic_request_batch_size: int = 50,
) -> Path:
    run_path = Path(run_dir)
    manifest = json.loads((run_path / "manifest.json").read_text(encoding="utf-8"))
    output = Path(pending_review_dir) / f"normalization_review_{manifest['run_id']}.xlsx"
    candidates = collect_unresolved_candidates(
        run_path,
        normalization_path,
        min_document_count,
        candidate_pool_path=candidate_pool_path,
    )
    pending_keys: set[tuple[str, str]] = set()
    for proposal_path in Path(pending_review_dir).glob("normalization_review_*.proposal.json"):
        if proposal_path == output.with_suffix(".proposal.json"):
            continue
        workbook_path = proposal_path.with_name(proposal_path.name.removesuffix(".proposal.json") + ".xlsx")
        if not workbook_path.exists() or workbook_path.with_suffix(".applied.json").exists():
            continue
        proposal = json.loads(proposal_path.read_text(encoding="utf-8"))
        if proposal.get("proposal_version") != PROPOSAL_VERSION:
            continue
        for candidate in proposal.get("candidates", []):
            if isinstance(candidate, dict):
                pending_keys.add(
                    (_normalization_key(str(candidate.get("source_name"))), str(candidate.get("item_type")))
                )
    candidates = [
        candidate
        for candidate in candidates
        if (_normalization_key(candidate["source_name"]), candidate["item_type"]) not in pending_keys
    ]
    candidates = candidates[:max_candidates_per_review]
    for candidate in candidates:
        candidate["evidence_samples"] = candidate["evidence_samples"][:max_evidence_samples]
    checkpoint_path = output.with_suffix(".semantic-checkpoint.json")
    fingerprint = _semantic_checkpoint_fingerprint(
        candidates,
        normalization_path,
        model,
    )
    completed_candidate_ids, suggestions = _load_semantic_checkpoint(
        checkpoint_path,
        fingerprint,
    )
    remaining = [
        candidate
        for candidate in candidates
        if candidate["candidate_id"] not in completed_candidate_ids
    ]
    total_batches = (
        len(remaining) + semantic_request_batch_size - 1
    ) // semantic_request_batch_size
    if completed_candidate_ids:
        print(
            f"Resuming normalization candidate review: "
            f"{len(completed_candidate_ids)}/{len(candidates)} candidates completed.",
            flush=True,
        )
    for batch_number, offset in enumerate(
        range(0, len(remaining), semantic_request_batch_size),
        start=1,
    ):
        chunk = remaining[offset: offset + semantic_request_batch_size]
        print(
            f"Normalization candidate semantic batch {batch_number}/{total_batches}: "
            f"{len(chunk)} candidates.",
            flush=True,
        )
        chunk_suggestions = request_semantic_suggestions(
            chunk,
            normalization_path,
            model,
        )
        duplicate_ids = set(suggestions) & set(chunk_suggestions)
        if duplicate_ids:
            raise ValueError(
                "Semantic suggestion batches returned duplicate candidate ids: "
                + ", ".join(sorted(duplicate_ids))
            )
        suggestions.update(chunk_suggestions)
        completed_candidate_ids.update(
            candidate["candidate_id"] for candidate in chunk
        )
        _write_semantic_checkpoint(
            checkpoint_path,
            fingerprint,
            completed_candidate_ids,
            suggestions,
        )
        print(
            f"Normalization candidate semantic batch {batch_number}/{total_batches} "
            f"completed; checkpoint saved.",
            flush=True,
        )
    review_path = write_review_workbook(
        output,
        manifest["run_id"],
        candidates,
        suggestions,
    )
    checkpoint_path.unlink(missing_ok=True)
    return review_path


def _as_bool(value: Any) -> bool:
    return value is True or (isinstance(value, str) and value.strip().casefold() in {"true", "yes", "是"})


def _write_receipt(receipt_path: Path, workbook_path: Path, config_path: Path, updates: list[dict[str, str]]) -> None:
    config = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    receipt_path.write_text(
        json.dumps(
            {
                "review_file": str(workbook_path),
                "config_version": str(config.get("version", "unspecified")),
                "updates": updates,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def apply_review_workbook(
    path: str | Path,
    normalization_path: str,
    min_create_new_document_count: int = 3,
    decision_ledger_path: str | Path | None = None,
) -> list[str]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=True)
    if REVIEW_SHEET not in workbook.sheetnames:
        raise ValueError(f"Review workbook must contain sheet {REVIEW_SHEET!r}.")
    sheet = workbook[REVIEW_SHEET]
    if not _as_bool(sheet[REVIEW_COMPLETE_CELL].value):
        return []
    proposal_path = workbook_path.with_suffix(".proposal.json")
    if not proposal_path.exists():
        raise ValueError("Review workbook is missing its immutable proposal sidecar.")
    proposal = json.loads(proposal_path.read_text(encoding="utf-8"))
    if proposal.get("proposal_version") != PROPOSAL_VERSION:
        raise ValueError("Review proposal uses an obsolete candidate identity version and cannot be applied.")
    proposal_candidates = {item["candidate_id"]: item for item in proposal.get("candidates", [])}
    headers = tuple(sheet.cell(5, index).value for index in range(1, len(REQUIRED_HEADERS) + 1))
    if headers != REQUIRED_HEADERS:
        raise ValueError("Review workbook columns do not match the required contract.")
    config_path = Path(normalization_path)
    raw_config = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    before = deepcopy(raw_config)
    loaded = load_normalization_map(normalization_path)
    updates: list[dict[str, str]] = []
    decisions: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=6, max_col=len(REQUIRED_HEADERS), values_only=True):
        if not any(value is not None for value in values):
            continue
        (
            candidate_id, source_name, document_count, item_type, _, action, anchor_alias, skill_id,
            canonical_name, category_code, subcategory_code, _, enabled, _,
        ) = values
        if not _as_bool(enabled):
            continue
        if not isinstance(candidate_id, str) or not isinstance(source_name, str) or not isinstance(item_type, str):
            raise ValueError("Accepted review row is missing immutable candidate fields.")
        proposal_candidate = proposal_candidates.get(candidate_id)
        if (
            not isinstance(proposal_candidate, dict)
            or _excel_text(proposal_candidate.get("source_name")) != source_name
            or _excel_text(proposal_candidate.get("item_type")) != item_type
            or proposal_candidate.get("document_count") != document_count
        ):
            raise ValueError("Review workbook contains an altered immutable candidate field.")
        source_name = proposal_candidate["source_name"]
        item_type = proposal_candidate["item_type"]
        if action not in REVIEW_ACTIONS:
            raise ValueError(
                f"Accepted candidate {source_name!r} uses unsupported action {action!r}."
            )
        mapping = None
        if action == "create_new":
            if not isinstance(document_count, int) or document_count < min_create_new_document_count:
                raise ValueError(
                    f"Accepted create_new candidate {source_name!r} has document_count="
                    f"{document_count!r}; at least {min_create_new_document_count} independent "
                    "documents are required."
                )
            mapping = _validate_suggested_mapping(
                {
                    "skill_id": skill_id,
                    "canonical_name": canonical_name,
                    "category_code": category_code,
                    "subcategory_code": subcategory_code,
                },
            )
        decisions.append(
            {
                "candidate_id": candidate_id,
                "source_name": source_name,
                "item_type": item_type,
                "action": action,
                "anchor_alias": anchor_alias,
                "mapping": mapping,
                "document_count": document_count,
            }
        )

    decisions_by_key = {
        (_normalization_key(item["source_name"]), item["item_type"]): item for item in decisions
    }
    if len(decisions_by_key) != len(decisions):
        raise ValueError("Review workbook accepts conflicting duplicate source names.")
    new_mappings = {
        key: item["mapping"]
        for key, item in decisions_by_key.items()
        if item["action"] == "create_new"
    }
    new_mapping_keys_by_source: dict[str, list[tuple[str, str]]] = {}
    for key in new_mappings:
        new_mapping_keys_by_source.setdefault(key[0], []).append(key)
    mappings_by_skill_id: dict[str, tuple[str, str, str | None]] = {}
    for mapping in [*raw_config.get("skills", {}).values(), *new_mappings.values()]:
        if not isinstance(mapping, dict) or not isinstance(mapping.get("skill_id"), str):
            continue
        skill_id = mapping["skill_id"]
        identity_metadata = (
            str(mapping.get("canonical_name")),
            str(mapping.get("category_code")),
            mapping.get("subcategory_code"),
        )
        previous = mappings_by_skill_id.get(skill_id)
        if previous is not None and previous != identity_metadata:
            raise ValueError(f"Conflicting mappings reuse skill_id {skill_id!r}.")
        mappings_by_skill_id[skill_id] = identity_metadata
    accepted_keys: set[tuple[str, str]] = set()
    for decision in decisions:
        source_name = decision["source_name"]
        item_type = decision["item_type"]
        action = decision["action"]
        anchor_alias = decision["anchor_alias"]
        if action in DISPOSITION_ACTIONS:
            continue
        if action == "create_new":
            mapping = deepcopy(decision["mapping"])
        elif action == "alias_existing":
            if not isinstance(anchor_alias, str) or not isinstance(raw_config.get("skills", {}).get(anchor_alias), dict):
                raise ValueError(f"Accepted candidate {source_name!r} must use an existing anchor_alias.")
            mapping = deepcopy(raw_config["skills"][anchor_alias])
        else:
            if not isinstance(anchor_alias, str):
                raise ValueError(f"Accepted candidate {source_name!r} must name a same-workbook anchor candidate.")
            anchor_keys = new_mapping_keys_by_source.get(_normalization_key(anchor_alias), [])
            if len(anchor_keys) != 1:
                raise ValueError(
                    f"Accepted candidate {source_name!r} must resolve to exactly one enabled create_new row."
                )
            anchor_key = anchor_keys[0]
            anchor_decision = decisions_by_key.get(anchor_key)
            if not isinstance(anchor_decision, dict) or anchor_decision.get("action") != "create_new":
                raise ValueError(f"Accepted candidate {source_name!r} must anchor to an enabled create_new row.")
            mapping = deepcopy(new_mappings[anchor_key])
        mapping_type = mapping["category_code"]
        key = (_normalization_key(source_name), item_type)
        if lookup_skill_mapping(loaded, source_name, item_type) is not None:
            continue
        if key in accepted_keys:
            raise ValueError(f"Review workbook accepts conflicting duplicate source key {source_name!r}.")
        accepted_keys.add(key)
        raw_config["skills"][source_name] = mapping
        if item_type != mapping_type:
            overrides = raw_config.setdefault("skill_source_type_overrides", {})
            source_types = overrides.setdefault(source_name, [])
            if item_type not in source_types:
                source_types.append(item_type)
        updates.append(
            {
                "source_name": source_name,
                "action": action,
                "skill_id": mapping["skill_id"],
                "canonical_name": mapping["canonical_name"],
            }
        )
    if any(before["skills"].get(alias) != raw_config["skills"].get(alias) for alias in before["skills"]):
        raise ValueError("Configuration application may only add aliases.")
    if updates:
        raw_config["config_revision"] = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        tmp_path = config_path.with_suffix(config_path.suffix + ".tmp")
        tmp_path.write_text(yaml.safe_dump(raw_config, allow_unicode=True, sort_keys=False), encoding="utf-8")
        replayed = load_normalization_map(str(tmp_path))
        for decision in decisions:
            if decision["action"] not in RESOLVED_ACTIONS:
                continue
            source_name = decision["source_name"]
            item_type = decision["item_type"]
            if lookup_skill_mapping(replayed, source_name, item_type) is None:
                raise ValueError(f"Replay validation failed for accepted alias {source_name!r}.")
        os.replace(tmp_path, config_path)
    if decision_ledger_path is not None:
        ledger = load_decision_ledger(decision_ledger_path)
        reviewed_at = datetime.now(timezone.utc).isoformat()
        suggestions = proposal.get("suggestions", {})
        for decision in decisions:
            candidate = proposal_candidates[decision["candidate_id"]]
            suggestion = suggestions.get(candidate["candidate_id"], {})
            ledger["decisions"][candidate["candidate_id"]] = {
                "candidate_id": candidate["candidate_id"],
                "source_name": candidate["source_name"],
                "item_type": candidate["item_type"],
                "document_count": candidate["document_count"],
                "evidence_samples": candidate.get("evidence_samples", []),
                "action": decision["action"],
                "anchor_alias": decision["anchor_alias"],
                "mapping": decision["mapping"],
                "reason": suggestion.get("reason"),
                "reviewed_at": reviewed_at,
                "source_run_id": proposal.get("run_id"),
                "proposal_version": proposal.get("proposal_version"),
            }
        _write_decision_ledger(decision_ledger_path, ledger)
    receipt = workbook_path.with_suffix(".applied.json")
    _write_receipt(receipt, workbook_path, config_path, updates)
    return [update["source_name"] for update in updates]


def archive_applied_review(workbook_path: str | Path, archive_dir: str | Path) -> None:
    workbook = Path(workbook_path)
    archive_path = Path(archive_dir)
    archive_path.mkdir(parents=True, exist_ok=True)
    for source in (workbook, workbook.with_suffix(".proposal.json"), workbook.with_suffix(".applied.json")):
        if not source.exists():
            raise ValueError(f"Applied review artifact is missing: {source}")
        target = archive_path / source.name
        if target.exists():
            raise ValueError(f"Archived review artifact already exists: {target}")
        shutil.move(str(source), str(target))


def _mark_candidate_pool_reviewed(proposal_path: Path, candidate_pool_path: str | Path) -> None:
    pool = _load_candidate_pool(candidate_pool_path)
    proposal = json.loads(proposal_path.read_text(encoding="utf-8"))
    for candidate in proposal.get("candidates", []):
        if not isinstance(candidate, dict):
            continue
        key = _candidate_key(str(candidate.get("source_name")), str(candidate.get("item_type")))
        stored = pool["candidates"].get(key)
        if isinstance(stored, dict):
            stored["last_reviewed_document_count"] = max(
                int(stored.get("last_reviewed_document_count", 0)),
                int(candidate.get("document_count", 0)),
            )
    _write_candidate_pool(candidate_pool_path, pool)


def apply_completed_reviews(
    pending_review_dir: str,
    normalization_path: str,
    applied_review_dir: str | None = None,
    candidate_pool_path: str | Path | None = None,
    min_create_new_document_count: int = 3,
    decision_ledger_path: str | Path | None = None,
) -> dict[str, list[str]]:
    applied: dict[str, list[str]] = {}
    archive_dir = applied_review_dir or str(Path(pending_review_dir) / "applied")
    for workbook in sorted(Path(pending_review_dir).glob("normalization_review_*.xlsx")):
        if workbook.with_suffix(".applied.json").exists():
            if candidate_pool_path is not None:
                _mark_candidate_pool_reviewed(workbook.with_suffix(".proposal.json"), candidate_pool_path)
            archive_applied_review(workbook, archive_dir)
            continue
        updates = apply_review_workbook(
            workbook,
            normalization_path,
            min_create_new_document_count=min_create_new_document_count,
            decision_ledger_path=decision_ledger_path,
        )
        if workbook.with_suffix(".applied.json").exists():
            if candidate_pool_path is not None:
                _mark_candidate_pool_reviewed(workbook.with_suffix(".proposal.json"), candidate_pool_path)
            archive_applied_review(workbook, archive_dir)
        if updates:
            applied[str(workbook)] = updates
    return applied
