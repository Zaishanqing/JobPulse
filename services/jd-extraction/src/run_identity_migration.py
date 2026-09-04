from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .document_identity import build_offline_document_id
from .load_excel import load_excel_rows
from .preprocess import preprocess_row


def _replace_exact_ids(value: Any, mapping: dict[str, str]) -> Any:
    if isinstance(value, dict):
        return {key: _replace_exact_ids(item, mapping) for key, item in value.items()}
    if isinstance(value, list):
        return [_replace_exact_ids(item, mapping) for item in value]
    if isinstance(value, str):
        return mapping.get(value, value)
    return value


def _resolve_input_path(manifest_path: str, data_root: Path) -> Path:
    recorded = Path(manifest_path)
    if recorded.is_file():
        return recorded
    matches = list(data_root.rglob(recorded.name))
    if len(matches) != 1:
        raise ValueError(
            f"Expected exactly one local input named {recorded.name!r}; found {len(matches)}."
        )
    return matches[0]


def build_run_id_mapping(
    run_dir: str | Path,
    source_platform: str,
    data_root: str | Path,
) -> tuple[dict[str, str], Path]:
    run_path = Path(run_dir)
    manifest = json.loads((run_path / "manifest.json").read_text(encoding="utf-8"))
    input_path = _resolve_input_path(str(manifest["input_path"]), Path(data_root))
    rows = load_excel_rows(str(input_path))
    mapping: dict[str, str] = {}
    for record_path in sorted((run_path / "records" / "success").glob("*.json")):
        record = json.loads(record_path.read_text(encoding="utf-8"))
        row_index = int(record["row_index"])
        old_id = str(record["annotation"]["document_id"])
        jd_input, failure = preprocess_row(rows[row_index - 1], row_index)
        if failure is not None or jd_input is None:
            raise ValueError(f"Cannot reconstruct source row {row_index} in {run_path.name}.")
        new_id = build_offline_document_id(
            source_platform,
            input_path,
            row_index,
            str(jd_input["jd_text_original"]),
        )
        previous = mapping.get(old_id)
        if previous is not None and previous != new_id:
            raise ValueError(f"Legacy document ID {old_id!r} is duplicated within {run_path.name}.")
        mapping[old_id] = new_id
    if len(mapping) != len(set(mapping.values())):
        raise ValueError(f"New document IDs are not unique within {run_path.name}.")
    return mapping, input_path


def _rewrite_json_file(path: Path, mapping: dict[str, str]) -> None:
    payload = json.loads(path.read_text(encoding="utf-8"))
    rewritten = _replace_exact_ids(payload, mapping)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(rewritten, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp_path, path)


def _rewrite_jsonl_file(path: Path, mapping: dict[str, str]) -> None:
    records = [
        _replace_exact_ids(json.loads(line), mapping)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", encoding="utf-8") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")
    os.replace(tmp_path, path)


def _rewrite_xlsx_file(path: Path, mapping: dict[str, str]) -> None:
    workbook = load_workbook(path)
    changed = False
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in mapping:
                    cell.value = mapping[cell.value]
                    changed = True
    if changed:
        tmp_path = path.with_name(path.stem + ".tmp.xlsx")
        workbook.save(tmp_path)
        os.replace(tmp_path, path)


def migrate_run_document_ids(
    run_dir: str | Path,
    source_platform: str,
    data_root: str | Path,
) -> dict[str, Any]:
    run_path = Path(run_dir)
    mapping, input_path = build_run_id_mapping(run_path, source_platform, data_root)
    old_ids = set(mapping)

    for path in sorted(run_path.rglob("*.json")):
        _rewrite_json_file(path, mapping)
    for path in sorted(run_path.rglob("*.jsonl")):
        _rewrite_jsonl_file(path, mapping)
    for path in sorted(run_path.rglob("*.xlsx")):
        _rewrite_xlsx_file(path, mapping)

    rename_candidates = sorted(
        (
            path
            for path in run_path.rglob("*")
            if path.is_file() and any(old_id in path.name for old_id in old_ids)
        ),
        key=lambda path: len(path.parts),
        reverse=True,
    )
    for path in rename_candidates:
        name = path.name
        for old_id, new_id in mapping.items():
            name = name.replace(old_id, new_id)
        target = path.with_name(name)
        if target.exists() and target != path:
            raise ValueError(f"Refusing to overwrite migrated file: {target}")
        path.rename(target)

    manifest_path = run_path / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["source_platform"] = source_platform
    manifest["document_identity_scheme"] = "offline-source-version-v1"
    manifest["document_identity_input"] = str(input_path)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    annotations = [
        json.loads(line)
        for line in (run_path / "final" / "annotations.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    normalized = [
        json.loads(line)
        for line in (run_path / "final" / "normalized_annotations.jsonl").read_text(
            encoding="utf-8"
        ).splitlines()
        if line.strip()
    ]
    annotation_ids = [item["document_id"] for item in annotations]
    normalized_ids = [item["document_id"] for item in normalized]
    if set(annotation_ids) != set(mapping.values()) or annotation_ids != normalized_ids:
        raise ValueError(f"Migrated aggregate IDs do not match in {run_path.name}.")
    return {
        "run_id": run_path.name,
        "source_platform": source_platform,
        "input_path": str(input_path),
        "document_count": len(mapping),
        "mapping": mapping,
    }
