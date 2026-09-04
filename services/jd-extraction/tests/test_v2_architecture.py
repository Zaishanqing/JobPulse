import json
import shutil
from pathlib import Path
from time import perf_counter, sleep
from uuid import uuid4

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook
from pydantic import ValidationError

import src.config_iteration as config_iteration
from src.deepseek_client import DeepSeekResult
from src.config_iteration import (
    REVIEW_COMPLETE_CELL,
    REVIEW_SHEET,
    _candidate_id,
    _build_semantic_update_prompt,
    apply_review_workbook,
    archive_applied_review,
    collect_unresolved_candidates,
    load_decision_ledger,
    write_review_workbook,
)
from src.audit import RunAudit
from src.deduplicator import deduplicate_extraction
from src.deterministic_fields import (
    canonicalize_authoritative_fields,
    infer_employment_kind,
    populate_deterministic_fields,
)
from src.exceptions import InputFormatError, SchemaValidationError, SemanticValidationError, SourceBindingError
from src.exporter import export_xlsx
from src.models import (
    JDExtractionResult,
    JDNormalizedResult,
    OtherRequirement,
    SkillRequirement,
)
from src.normalizer import (
    load_normalization_map,
    lookup_skill_mapping,
    normalize_extraction,
)
from src.pipeline import JDExtractionPipeline
from src.post_review import apply_annotation_decisions
from src.local_repair import apply_local_repair, plan_local_repair
from src.prompt_builder import (
    build_local_repair_prompt,
    build_model_output_schema,
    build_system_prompt,
    build_user_prompt,
    build_validation_retry_prompt,
)
from src.provenance import align_all_evidence, canonicalize_evidence_quotes
from src.report_generator import generate_report, summarize_run
from src.run_renormalizer import renormalize_run
from src.semantic_rules import compile_semantic_handbook, load_semantic_rules
from src.skill_taxonomy import (
    build_classification_records,
    iter_jd_skill_occurrences,
    load_skill_taxonomy_snapshot,
    write_unified_normalized_artifacts,
)
from src.validator import (
    reject_retryable_review_flags,
    validate_business_rules,
    validate_normalized_rules,
    validate_schema,
    validate_semantic_constraints,
    validate_skill_item_type_contract,
)


def extraction_payload() -> dict:
    return {
        "document_id": "jd_001",
        "job_title": None,
        "responsibilities": [
            {
                "requirement_id": "req_001",
                "kind": "task",
                "modality": "required",
                "action": "开发服务",
                "evidence": {"source_id": "src_1", "quote": "负责开发服务"},
            }
        ],
        "requirements": [
            {
                "requirement_id": "req_002",
                "kind": "skill",
                "modality": "required",
                "proficiency": "proficient",
                "items": [
                    {"name": "Python", "item_type": "programming_language"},
                    {"name": "Python", "item_type": "programming_language"},
                ],
                "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"},
            },
            {
                "requirement_id": "req_003",
                "kind": "experience",
                "modality": "preferred",
                "minimum_years": 3,
                "domain": "Python 开发",
                "evidence": {"source_id": "src_3", "quote": "3年以上 Python 开发经验优先"},
            },
        ],
        "company_facts": [
            {
                "fact_id": "company_001",
                "kind": "industry",
                "value": "软件",
                "evidence": {"source_id": "src_4", "quote": "软件公司"},
            }
        ],
        "employment_facts": [],
    }


def source_blocks() -> list[dict]:
    texts = ["负责开发服务", "熟练使用 Python", "3年以上 Python 开发经验优先", "软件公司"]
    offset = 0
    blocks = []
    for index, text in enumerate(texts, start=1):
        blocks.append({"source_id": f"src_{index}", "text": text, "start": offset, "end": offset + len(text)})
        offset += len(text) + 1
    return blocks


def taxonomy() -> dict:
    mapping = {
        "version": "2.0",
        "skills": {
            "Python": {
                "skill_id": "LANG_PYTHON",
                "canonical_name": "Python 语言",
                "category_code": "programming_language",
                "subcategory_code": "GENERAL_PURPOSE",
            }
        },
        "position_taxonomy_version": "position-taxonomy.v3.0.0",
    }
    python_mapping = mapping["skills"]["Python"]
    mapping["_skills_by_exact_typed_key"] = {
        ("Python", "programming_language"): python_mapping
    }
    mapping["_skill_mappings_by_exact_key"] = {"Python": [python_mapping]}
    mapping["_skills_by_typed_normalized_key"] = {
        ("python", "programming_language"): python_mapping
    }
    mapping["_skill_mappings_by_normalized_key"] = {"python": [python_mapping]}
    mapping["_skills_by_normalized_key"] = {"python": python_mapping}
    return mapping


def test_discriminated_model_separates_tasks_constraints_and_facts():
    result = JDExtractionResult.model_validate(extraction_payload())
    assert result.responsibilities[0].kind == "task"
    assert isinstance(result.requirements[0], SkillRequirement)
    assert result.company_facts[0].kind == "industry"


def test_old_monolithic_item_fields_are_rejected():
    payload = extraction_payload()
    payload["requirements"][0]["items"][0]["sub_category"] = "后端开发"
    with pytest.raises(ValidationError, match="sub_category"):
        JDExtractionResult.model_validate(payload)


def test_experience_range_is_strict():
    payload = extraction_payload()
    payload["requirements"][1]["minimum_years"] = 5
    payload["requirements"][1]["maximum_years"] = 3
    with pytest.raises(ValidationError, match="must not exceed"):
        JDExtractionResult.model_validate(payload)


def test_model_schema_omits_python_and_normalization_fields():
    schema = build_model_output_schema()
    serialized = json.dumps(schema, ensure_ascii=False)
    assert "document_id" not in schema["properties"]
    assert "requirement_id" not in serialized
    assert "fact_id" not in serialized
    assert "normalized_name" not in serialized
    assert "job_family" not in serialized
    assert "sub_category" not in serialized
    assert '"start"' not in serialized


def test_prompt_has_atomic_evidence_and_normalization_boundaries():
    system = build_system_prompt()
    user = build_user_prompt(
        {"jd_id": "jd_001", "job_title_raw": "工程师", "company": "示例", "source_blocks": source_blocks()}
    )
    assert "原子事实抽取器" in system
    assert "不负责归一化" in system
    assert "evidence.quote" in user
    assert '"source_id": "src_1"' in user
    assert "数组字段没有内容时只能输出 []" in system
    assert "不得增加 label" in system
    assert "来源平台插入词" not in system
    assert "不得再次输出为 responsibility" in system


def test_source_binding_retry_prompt_includes_exact_copy_contract():
    prompt = build_validation_retry_prompt(
        {"jd_id": "jd_001", "job_title_raw": "工程师", "company": "示例", "source_blocks": source_blocks()},
        "SourceBindingError",
        {"invalid_quote": "熟练掌握 Python", "exact_source_text": "熟练使用 Python"},
    )
    assert "一次列出本轮发现的全部错误对象" in prompt
    assert "保留所有空格" in prompt
    assert "删除该对象 value/action/domain 等语义字段中不受 exact_source_text 支持" in prompt
    assert "原文不支持该对象，则不要生成该对象" in prompt


def test_semantic_retry_prompt_contains_code_specific_corrections():
    history = [
        {
            "attempt": 1,
            "error_type": "SourceBindingError",
            "error_details": {"invalid_quote": "错误标点;", "exact_source_text": "正确标点。"},
        },
        {
            "attempt": 2,
            "error_type": "SemanticValidationError",
            "error_details": [
                {"code": "composite_skill_item", "name": "C/C++", "parts": ["C", "C++"]},
                {"code": "experience_phrase_in_skill_item", "name": "架构设计经验"},
            ],
        },
    ]
    prompt = build_validation_retry_prompt(
        {"jd_id": "jd_001", "job_title_raw": "工程师", "company": "示例", "source_blocks": source_blocks()},
        "SemanticValidationError",
        [
            {"code": "composite_skill_item", "name": "C/C++", "parts": ["C", "C++"]},
            {"code": "experience_phrase_in_skill_item", "name": "架构设计经验"},
        ],
        previous_invalid_output='{"items":[{"name":"C/C++"}]}',
        validation_history=history,
    )
    assert "按 violations.parts 将复合名称拆成多个独立 SkillItem" in prompt
    assert "例如 C/C++ 拆为 C 与 C++" in prompt
    assert "改为 ExperienceRequirement" in prompt
    assert "上一轮被拒绝的完整 JSON" in prompt
    assert '"name":"C/C++"' in prompt
    assert "本条 JD 累计校验历史" in prompt
    assert "正确标点。" in prompt
    assert "不得在修复当前错误时重新引入此前已修正的问题" in prompt


def test_empty_experience_retry_prompt_requires_structured_domain():
    prompt = build_validation_retry_prompt(
        {"jd_id": "jd_001", "job_title_raw": "工程师", "company": "示例", "source_blocks": source_blocks()},
        "SemanticValidationError",
        [{"issue_type": "empty_structured_constraint", "kind": "experience"}],
    )
    assert "必须从原文填写 domain、role、year 或 duration_text" in prompt
    assert "大模型研究、部署" in prompt


def test_conflicting_modality_retry_prefers_explicit_prose():
    prompt = build_validation_retry_prompt(
        {"jd_id": "jd_001", "job_title_raw": "工程师", "company": "示例", "source_blocks": source_blocks()},
        "SemanticValidationError",
        [{"issue_type": "conflicting_requirement_modality", "raw_text": "硕士优先"}],
    )
    assert "显式强度词的完整正文优先" in prompt
    assert "学历:硕士" in prompt


def test_local_repair_prompt_forbids_full_reextraction_and_generated_fields():
    prompt = build_local_repair_prompt(
        "SemanticValidationError",
        [{"code": "base_salary_in_employment_fact", "fact_id": "employment_001"}],
        [{"collection": "employment_facts", "index": 0, "current_value": {"kind": "other"}}],
        [{"source_id": "src_1", "text": "20-40K"}],
    )
    assert "不要重新抽取整条 JD" in prompt
    assert '"operations"' in prompt
    assert "禁止填写 document_id、requirement_id、fact_id" in prompt


def test_local_repair_prompt_requires_one_combined_operation_per_target():
    prompt = build_local_repair_prompt(
        "SchemaValidationError",
        [
            {"loc": ["requirements", 0, "education", "majors"], "type": "list_type"},
            {"loc": ["requirements", 0, "education", "school_constraints"], "type": "list_type"},
        ],
        [{"collection": "requirements", "index": 0, "current_value": {"kind": "education"}}],
        [{"source_id": "src_1", "text": "学历:本科"}],
    )
    assert "同一个 collection/index 最多只能出现一次" in prompt
    assert "全部修正合并到一个完整 replace value" in prompt


def test_local_repair_applies_only_authorized_object_and_preserves_other_payload():
    payload = extraction_payload()
    canonicalized = extraction_payload()
    plan = plan_local_repair(
        canonicalized,
        "SemanticValidationError",
        [{"code": "composite_skill_item", "requirement_id": "req_002", "name": "Python/PyTorch"}],
        source_blocks(),
    )
    assert plan is not None
    assert plan.targets[0].collection == "requirements"
    assert plan.targets[0].index == 0
    repaired = apply_local_repair(
        payload,
        {
            "operations": [
                {
                    "op": "replace",
                    "target": {"collection": "requirements", "index": 0},
                    "value": {
                        "kind": "skill",
                        "modality": "required",
                        "proficiency": "proficient",
                        "items": [{"name": "Python", "item_type": "programming_language"}],
                        "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"},
                    },
                }
            ]
        },
        plan,
    )
    assert repaired["responsibilities"] == payload["responsibilities"]
    assert repaired["requirements"][1] == payload["requirements"][1]
    assert repaired["requirements"][0]["items"] == [{"name": "Python", "item_type": "programming_language"}]


def test_candidate_requirement_can_move_from_responsibilities_to_requirements():
    payload = extraction_payload()
    plan = plan_local_repair(
        payload,
        "SemanticValidationError",
        [
            {
                "code": "candidate_requirement_in_responsibility",
                "requirement_id": "req_001",
                "action": "熟悉 Python",
            }
        ],
        source_blocks(),
    )
    assert plan is not None
    assert plan.append_collections == ("requirements",)
    repaired = apply_local_repair(
        payload,
        {
            "operations": [
                {"op": "remove", "target": {"collection": "responsibilities", "index": 0}},
                {
                    "op": "append",
                    "target": {"collection": "requirements"},
                    "value": {
                        "kind": "skill",
                        "modality": "required",
                        "items": [{"name": "Python", "item_type": "programming_language"}],
                        "evidence": {"source_id": "src_1", "quote": "负责开发服务"},
                    },
                },
            ]
        },
        plan,
    )
    assert repaired["responsibilities"] == []
    assert repaired["requirements"][-1]["kind"] == "skill"


def test_local_repair_rejects_operation_outside_error_scope():
    plan = plan_local_repair(
        extraction_payload(),
        "SemanticValidationError",
        [{"code": "composite_skill_item", "requirement_id": "req_002", "name": "Python/PyTorch"}],
        source_blocks(),
    )
    assert plan is not None
    with pytest.raises(ValueError, match="outside the authorized error scope"):
        apply_local_repair(
            extraction_payload(),
            {
                "operations": [
                    {
                        "op": "remove",
                        "target": {"collection": "company_facts", "index": 0},
                    }
                ]
            },
            plan,
        )


def test_local_repair_rejects_duplicate_target_operation():
    plan = plan_local_repair(
        extraction_payload(),
        "SemanticValidationError",
        [{"code": "composite_skill_item", "requirement_id": "req_002", "name": "Python/PyTorch"}],
        source_blocks(),
    )
    assert plan is not None
    operation = {"op": "remove", "target": {"collection": "requirements", "index": 0}}
    with pytest.raises(ValueError, match="same target more than once"):
        apply_local_repair(extraction_payload(), {"operations": [operation, operation]}, plan)


def test_local_repair_declines_non_object_schema_error():
    plan = plan_local_repair(
        extraction_payload(),
        "SchemaValidationError",
        [{"loc": ["job_title"], "type": "model_type", "msg": "invalid title"}],
        source_blocks(),
    )
    assert plan is None


def test_review_workbook_applies_only_after_completion_and_respects_unchecked_rows(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {
                    "PyTorch": {
                        "skill_id": "FRAMEWORK_PYTORCH",
                        "canonical_name": "PyTorch",
                        "category_code": "framework",
                        "subcategory_code": "DEEP_LEARNING",
                    }
                },
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            allow_unicode=True,
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    workbook_path = tmp_path / "normalization_review.xlsx"
    candidates = [
        {
            "candidate_id": "candidate-1",
            "source_name": "Torch",
            "document_count": 3,
            "item_type": "framework",
            "evidence_samples": [{"jd_id": "jd_001", "quote": "熟悉 Torch", "source_id": "src_1"}],
        }
    ]
    write_review_workbook(
        workbook_path,
        "run_001",
        candidates,
        {
            "candidate-1": {
                "action": "alias_existing",
                "anchor_alias": "PyTorch",
                "mapping": None,
                "reason": "明确同义写法",
            }
        },
    )
    assert apply_review_workbook(workbook_path, config_path) == []
    workbook = load_workbook(workbook_path)
    sheet = workbook[REVIEW_SHEET]
    sheet[REVIEW_COMPLETE_CELL] = True
    sheet["M6"] = False
    workbook.save(workbook_path)
    assert apply_review_workbook(workbook_path, config_path) == []
    workbook = load_workbook(workbook_path)
    workbook[REVIEW_SHEET]["M6"] = True
    workbook.save(workbook_path)
    assert apply_review_workbook(workbook_path, config_path) == ["Torch"]
    updated = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    assert updated["skills"]["Torch"] == updated["skills"]["PyTorch"]


def test_review_workbook_can_correct_candidate_type_create_entry_and_archive(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            allow_unicode=True,
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    workbook_path = tmp_path / "normalization_review_new.xlsx"
    mapping = {
        "skill_id": "FRAMEWORK_FASTAPI",
        "canonical_name": "FastAPI",
        "category_code": "framework",
        "subcategory_code": "WEB_BACKEND",
    }
    write_review_workbook(
        workbook_path,
        "run_002",
        [
            {
                "candidate_id": "candidate-2",
                "source_name": "FastAPI",
                "document_count": 3,
                "item_type": "tool",
                "evidence_samples": [{"jd_id": "jd_001", "quote": "熟悉 FastAPI", "source_id": "src_1"}],
            }
        ],
        {
            "candidate-2": {
                "action": "create_new",
                "anchor_alias": None,
                "mapping": mapping,
                "reason": "稳定框架实体，当前词表缺失",
            }
        },
    )
    workbook = load_workbook(workbook_path)
    workbook[REVIEW_SHEET][REVIEW_COMPLETE_CELL] = True
    workbook.save(workbook_path)

    assert apply_review_workbook(workbook_path, config_path) == ["FastAPI"]
    updated = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    assert updated["skills"]["FastAPI"] == mapping
    assert updated["skill_source_type_overrides"]["FastAPI"] == ["tool"]
    loaded = load_normalization_map(str(config_path))
    assert lookup_skill_mapping(loaded, "fastapi", "tool") == mapping

    archive_dir = tmp_path / "applied"
    archive_applied_review(workbook_path, archive_dir)
    assert not workbook_path.exists()
    assert not workbook_path.with_suffix(".proposal.json").exists()
    assert not workbook_path.with_suffix(".applied.json").exists()
    assert (archive_dir / workbook_path.name).exists()
    assert (archive_dir / workbook_path.with_suffix(".proposal.json").name).exists()
    assert (archive_dir / workbook_path.with_suffix(".applied.json").name).exists()


def test_review_workbook_enabled_defaults_to_true_only_for_applicable_suggestions(tmp_path):
    workbook_path = tmp_path / "normalization_review_default_true.xlsx"
    write_review_workbook(
        workbook_path,
        "run_003",
        [
            {
                "candidate_id": "candidate-3",
                "source_name": "unclear",
                "document_count": 2,
                "item_type": "other",
                "evidence_samples": [{"jd_id": "jd_001", "quote": "unclear", "source_id": "src_1"}],
            },
            {
                "candidate_id": "candidate-4",
                "source_name": "FastAPI",
                "document_count": 2,
                "item_type": "framework",
                "evidence_samples": [{"jd_id": "jd_002", "quote": "FastAPI", "source_id": "src_1"}],
            }
        ],
        {
            "candidate-3": {
                "action": "reject",
                "anchor_alias": None,
                "mapping": None,
                "reason": "不是稳定技能实体",
            },
            "candidate-4": {
                "action": "create_new",
                "anchor_alias": None,
                "mapping": {
                    "skill_id": "FRAMEWORK_FASTAPI",
                    "canonical_name": "FastAPI",
                    "category_code": "framework",
                    "subcategory_code": "WEB_BACKEND",
                },
                "reason": "稳定框架实体",
            }
        },
    )
    workbook = load_workbook(workbook_path, data_only=True)
    assert workbook[REVIEW_SHEET]["M6"].value is False
    assert workbook[REVIEW_SHEET]["M7"].value is True


def test_generated_candidate_id_is_excel_safe_and_round_trips(tmp_path):
    workbook_path = tmp_path / "normalization_review_excel_safe.xlsx"
    candidate_id = _candidate_id("run", "DDR", "domain_knowledge")
    candidates = [{
        "candidate_id": candidate_id,
        "source_name": "DDR",
        "document_count": 3,
        "item_type": "domain_knowledge",
        "evidence_samples": [],
    }]

    write_review_workbook(
        workbook_path,
        "run",
        candidates,
        {
            candidate_id: {
                "action": "reject",
                "anchor_alias": None,
                "mapping": None,
                "reason": "test",
            }
        },
    )

    assert "\x1f" not in candidate_id
    workbook = load_workbook(workbook_path, data_only=True)
    assert workbook[REVIEW_SHEET]["A6"].value == candidate_id
    proposal = json.loads(
        workbook_path.with_suffix(".proposal.json").read_text(encoding="utf-8")
    )
    assert proposal["candidates"][0]["candidate_id"] == candidate_id


def test_review_workbook_sanitizes_excel_cells_without_mutating_proposal(tmp_path):
    workbook_path = tmp_path / "normalization_review_illegal_text.xlsx"
    candidate_id = _candidate_id("run", "DDR", "domain_knowledge")
    candidates = [{
        "candidate_id": candidate_id,
        "source_name": "DDR\x1f",
        "document_count": 3,
        "item_type": "domain_knowledge",
        "evidence_samples": [{
            "jd_id": "jd_001",
            "quote": "DDR\x07 protocol",
            "source_id": "src_001",
        }],
    }]
    write_review_workbook(
        workbook_path,
        "run",
        candidates,
        {
            candidate_id: {
                "action": "ambiguous",
                "anchor_alias": None,
                "mapping": None,
                "reason": "evidence\x0bneeds review",
            }
        },
    )

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[REVIEW_SHEET]
    assert sheet["B6"].value == "DDR"
    assert sheet["E6"].value == "jd_001: DDR protocol"
    assert sheet["L6"].value == "evidenceneeds review"
    proposal = json.loads(
        workbook_path.with_suffix(".proposal.json").read_text(encoding="utf-8")
    )
    assert proposal["candidates"][0]["source_name"] == "DDR\x1f"


def test_review_disposition_is_recorded_without_changing_extraction_or_map(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    ledger_path = tmp_path / "normalization_decision_ledger.json"
    workbook_path = tmp_path / "normalization_review_disposition.xlsx"
    candidate_id = _candidate_id("run", "UE", "tool")
    write_review_workbook(
        workbook_path,
        "run_disposition",
        [{
            "candidate_id": candidate_id,
            "source_name": "UE",
            "document_count": 4,
            "item_type": "tool",
            "evidence_samples": [{
                "jd_id": "jd_001",
                "quote": "熟悉 UE",
                "source_id": "src_001",
            }],
        }],
        {
            candidate_id: {
                "action": "ambiguous",
                "anchor_alias": None,
                "mapping": None,
                "reason": "可能指 Unreal Engine 或其他缩写",
            }
        },
    )
    workbook = load_workbook(workbook_path)
    workbook[REVIEW_SHEET][REVIEW_COMPLETE_CELL] = True
    workbook.save(workbook_path)

    assert apply_review_workbook(
        workbook_path,
        config_path,
        decision_ledger_path=ledger_path,
    ) == []
    assert yaml.safe_load(config_path.read_text(encoding="utf-8"))["skills"] == {}
    decision = load_decision_ledger(ledger_path)["decisions"][candidate_id]
    assert decision["action"] == "ambiguous"
    assert decision["source_name"] == "UE"


def test_review_workbook_allows_human_mapping_edits_and_same_batch_alias(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    workbook_path = tmp_path / "normalization_review_alias_candidate.xlsx"
    candidates = [
        {
            "candidate_id": "go",
            "source_name": "Go",
            "document_count": 3,
            "item_type": "programming_language",
            "evidence_samples": [],
        },
        {
            "candidate_id": "golang",
            "source_name": "Golang",
            "document_count": 3,
            "item_type": "domain_knowledge",
            "evidence_samples": [],
        },
    ]
    write_review_workbook(
        workbook_path,
        "run_alias",
        candidates,
        {
            "go": {
                "action": "create_new",
                "anchor_alias": None,
                "mapping": {
                    "skill_id": "LANG_GO",
                    "canonical_name": "GO",
                    "category_code": "programming_language",
                    "subcategory_code": "GENERAL_PURPOSE",
                },
                "reason": "new",
            },
            "golang": {"action": "reject", "anchor_alias": None, "mapping": None, "reason": "alias"},
        },
    )
    workbook = load_workbook(workbook_path)
    sheet = workbook[REVIEW_SHEET]
    sheet[REVIEW_COMPLETE_CELL] = True
    sheet["I6"] = "Go"
    sheet["F7"] = "alias_candidate"
    sheet["G7"] = "Go"
    sheet["M7"] = True
    workbook.save(workbook_path)

    assert apply_review_workbook(workbook_path, config_path) == ["Go", "Golang"]
    skills = yaml.safe_load(config_path.read_text(encoding="utf-8"))["skills"]
    assert skills["Go"] == skills["Golang"]
    assert skills["Go"]["canonical_name"] == "Go"


def test_review_workbook_rejects_low_coverage_new_identity(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    workbook_path = tmp_path / "normalization_review_low_coverage.xlsx"
    write_review_workbook(
        workbook_path,
        "run_low_coverage",
        [{
            "candidate_id": "candidate-low",
            "source_name": "OneOffFramework",
            "document_count": 2,
            "item_type": "framework",
            "evidence_samples": [],
        }],
        {
            "candidate-low": {
                "action": "create_new",
                "anchor_alias": None,
                "mapping": {
                    "skill_id": "FRAMEWORK_ONE_OFF",
                    "canonical_name": "OneOffFramework",
                    "category_code": "framework",
                    "subcategory_code": "OTHER",
                },
                "reason": "insufficient evidence",
            }
        },
    )
    workbook = load_workbook(workbook_path)
    workbook[REVIEW_SHEET][REVIEW_COMPLETE_CELL] = True
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="at least 3 independent documents"):
        apply_review_workbook(workbook_path, config_path)


def test_candidate_pool_deduplicates_same_input_across_runs(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    pool_path = tmp_path / "candidate_pool.json"

    def write_run(run_id: str, document_id: str, input_path: str) -> Path:
        run_dir = tmp_path / run_id
        final = run_dir / "final"
        final.mkdir(parents=True)
        (run_dir / "manifest.json").write_text(
            json.dumps({"run_id": run_id, "input_path": input_path}),
            encoding="utf-8",
        )
        annotation = {
            "document_id": document_id,
            "requirements": [{
                "requirement_id": "req_001",
                "kind": "skill",
                "evidence": {"source_id": "src_001", "quote": "RareSkill"},
                "items": [{"name": "RareSkill", "item_type": "tool"}],
            }],
        }
        normalized = {
            "document_id": document_id,
            "normalized_requirements": [{
                "requirement_id": "req_001",
                "skills": [{
                    "source_name": "RareSkill",
                    "identity_resolution_status": "unresolved",
                    "classification_resolution_status": "unresolved",
                    "classification_unresolved_reason": "identity_unresolved",
                    "classifications": [],
                }],
            }],
        }
        (final / "annotations.jsonl").write_text(json.dumps(annotation) + "\n", encoding="utf-8")
        (final / "normalized_annotations.jsonl").write_text(json.dumps(normalized) + "\n", encoding="utf-8")
        return run_dir

    first = collect_unresolved_candidates(
        write_run("run_1", "jd_001", "data/chunk_01.xlsx"), config_path, 2, pool_path
    )
    second = collect_unresolved_candidates(
        write_run("run_2", "jd_001", "data/chunk_01.xlsx"), config_path, 2, pool_path
    )
    assert first == []
    assert second == []
    third = collect_unresolved_candidates(
        write_run("run_3", "jd_001", "data/chunk_02.xlsx"), config_path, 2, pool_path
    )
    assert len(third) == 1
    assert third[0]["source_name"] == "RareSkill"
    assert third[0]["document_count"] == 2


def test_candidate_pool_replaces_observations_when_same_run_is_rebuilt(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    pool_path = tmp_path / "candidate_pool.json"

    def write_run(document_ids: list[str]) -> Path:
        run_dir = tmp_path / "same_run"
        final = run_dir / "final"
        final.mkdir(parents=True, exist_ok=True)
        (run_dir / "manifest.json").write_text(
            json.dumps({"run_id": "same_run", "input_path": "data/chunk.xlsx"}),
            encoding="utf-8",
        )
        annotations = []
        normalized = []
        for document_id in document_ids:
            annotations.append({
                "document_id": document_id,
                "requirements": [{
                    "requirement_id": "req_001", "kind": "skill",
                    "evidence": {"source_id": "src_001", "quote": "RareSkill"},
                    "items": [{"name": "RareSkill", "item_type": "tool"}],
                }],
            })
            normalized.append({
                "document_id": document_id,
                "normalized_requirements": [{
                    "requirement_id": "req_001",
                    "skills": [{
                        "source_name": "RareSkill",
                        "identity_resolution_status": "unresolved",
                        "classification_resolution_status": "unresolved",
                        "classification_unresolved_reason": "identity_unresolved",
                        "classifications": [],
                    }],
                }],
            })
        (final / "annotations.jsonl").write_text(
            "".join(json.dumps(item) + "\n" for item in annotations), encoding="utf-8"
        )
        (final / "normalized_annotations.jsonl").write_text(
            "".join(json.dumps(item) + "\n" for item in normalized), encoding="utf-8"
        )
        return run_dir

    first = collect_unresolved_candidates(write_run(["jd_001", "jd_002"]), config_path, 1, pool_path)
    second = collect_unresolved_candidates(write_run(["jd_001"]), config_path, 1, pool_path)

    assert first[0]["document_count"] == 2
    assert second[0]["document_count"] == 1
    evidence_ids = [item["jd_id"] for item in second[0]["evidence_samples"]]
    assert len(evidence_ids) == 1
    assert evidence_ids[0].endswith("/jd_001")


def test_candidate_pool_preserves_review_count_when_only_source_run_is_rebuilt(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "position_taxonomy_version": "position-taxonomy.v2.0.0",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )
    pool_path = tmp_path / "candidate_pool.json"
    run_dir = tmp_path / "same_run"
    final = run_dir / "final"
    final.mkdir(parents=True)
    (run_dir / "manifest.json").write_text(
        json.dumps({"run_id": "same_run", "input_path": "data/chunk.xlsx"}), encoding="utf-8"
    )
    annotation = {
        "document_id": "jd_001",
        "requirements": [{
            "requirement_id": "req_001", "kind": "skill",
            "evidence": {"source_id": "src_001", "quote": "RareSkill"},
            "items": [{"name": "RareSkill", "item_type": "tool"}],
        }],
    }
    normalized = {
        "document_id": "jd_001",
        "normalized_requirements": [{
            "requirement_id": "req_001",
            "skills": [{
                "source_name": "RareSkill",
                "identity_resolution_status": "unresolved",
                "classification_resolution_status": "unresolved",
                "classification_unresolved_reason": "identity_unresolved",
                "classifications": [],
            }],
        }],
    }
    (final / "annotations.jsonl").write_text(json.dumps(annotation) + "\n", encoding="utf-8")
    (final / "normalized_annotations.jsonl").write_text(json.dumps(normalized) + "\n", encoding="utf-8")

    first = collect_unresolved_candidates(run_dir, config_path, 1, pool_path)
    pool = json.loads(pool_path.read_text(encoding="utf-8"))
    stored = next(iter(pool["candidates"].values()))
    stored["last_reviewed_document_count"] = first[0]["document_count"]
    pool_path.write_text(json.dumps(pool), encoding="utf-8")

    second = collect_unresolved_candidates(run_dir, config_path, 1, pool_path)
    refreshed = next(iter(json.loads(pool_path.read_text(encoding="utf-8"))["candidates"].values()))
    assert second == []
    assert refreshed["last_reviewed_document_count"] == 1


def test_propose_from_run_caps_candidates_before_semantic_request(tmp_path, monkeypatch):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    (run_dir / "manifest.json").write_text(
        json.dumps({"run_id": "bounded_run", "input_path": "data/chunk.xlsx"}),
        encoding="utf-8",
    )
    candidates = [
        {
            "candidate_id": f"candidate_{index:03d}",
            "source_name": f"Skill {index:03d}",
            "document_count": 2,
            "item_type": "tool",
            "evidence_samples": [],
        }
        for index in range(75)
    ]
    captured: dict[str, int] = {}
    normalization_path = tmp_path / "normalization.yaml"
    normalization_path.write_text("skills: {}\n", encoding="utf-8")

    monkeypatch.setattr(config_iteration, "collect_unresolved_candidates", lambda *args, **kwargs: candidates)

    def fake_request(selected, normalization_path, model):
        captured["requested"] = len(selected)
        return {}

    monkeypatch.setattr(config_iteration, "request_semantic_suggestions", fake_request)
    monkeypatch.setattr(
        config_iteration,
        "write_review_workbook",
        lambda path, run_id, selected, suggestions: Path(path),
    )

    config_iteration.propose_from_run(
        run_dir,
        normalization_path,
        str(tmp_path / "pending"),
        "model",
        2,
        candidate_pool_path=tmp_path / "pool.json",
        max_candidates_per_review=50,
    )
    assert captured["requested"] == 50


def test_propose_from_run_chunks_large_review_without_dropping_candidates(
    tmp_path, monkeypatch
):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    (run_dir / "manifest.json").write_text(
        json.dumps({"run_id": "chunked_run", "input_path": "data/chunk.xlsx"}),
        encoding="utf-8",
    )
    candidates = [
        {
            "candidate_id": f"candidate_{index:03d}",
            "source_name": f"Skill {index:03d}",
            "document_count": 1,
            "item_type": "tool",
            "evidence_samples": [],
        }
        for index in range(125)
    ]
    requested_sizes: list[int] = []
    captured: dict[str, int] = {}
    normalization_path = tmp_path / "normalization.yaml"
    normalization_path.write_text("skills: {}\n", encoding="utf-8")

    monkeypatch.setattr(
        config_iteration,
        "collect_unresolved_candidates",
        lambda *args, **kwargs: candidates,
    )

    def fake_request(selected, normalization_path, model):
        requested_sizes.append(len(selected))
        return {
            candidate["candidate_id"]: {
                "action": "pending_review",
                "anchor_alias": None,
                "mapping": None,
                "reason": "test",
            }
            for candidate in selected
        }

    monkeypatch.setattr(
        config_iteration,
        "request_semantic_suggestions",
        fake_request,
    )

    def fake_workbook(path, run_id, selected, suggestions):
        captured["candidate_count"] = len(selected)
        captured["suggestion_count"] = len(suggestions)
        return Path(path)

    monkeypatch.setattr(
        config_iteration,
        "write_review_workbook",
        fake_workbook,
    )

    config_iteration.propose_from_run(
        run_dir,
        normalization_path,
        str(tmp_path / "pending"),
        "model",
        1,
        candidate_pool_path=tmp_path / "pool.json",
        max_candidates_per_review=125,
        semantic_request_batch_size=50,
    )

    assert requested_sizes == [50, 50, 25]
    assert captured == {"candidate_count": 125, "suggestion_count": 125}


def test_semantic_prompt_uses_only_relevant_catalog_entries():
    candidates = [{
        "candidate_id": "candidate_cuda",
        "source_name": "CUDA Toolkit",
        "document_count": 1,
        "item_type": "platform",
        "evidence_samples": [],
    }]
    skills = {
        f"Unrelated Tool {index}": {
            "skill_id": f"TOOL_{index}",
            "canonical_name": f"Unrelated Tool {index}",
            "category_code": "tool",
            "subcategory_code": "development_tool",
        }
        for index in range(30)
    }
    skills["CUDA"] = {
        "skill_id": "PLATFORM_CUDA",
        "canonical_name": "CUDA",
        "category_code": "platform",
        "subcategory_code": "compute_platform",
    }

    _, user_prompt = _build_semantic_update_prompt(
        candidates,
        {"skills": skills},
    )
    payload = json.loads(user_prompt)

    assert len(payload["existing_catalog"]) <= 12
    assert any(
        entry["alias"] == "CUDA"
        for entry in payload["existing_catalog"]
    )
    assert len(payload["existing_catalog"]) < len(skills)


def test_deterministic_identity_match_skips_remote_model():
    mapping = {
        "skill_id": "PLATFORM_NODEJS",
        "canonical_name": "Node.js",
        "category_code": "platform",
        "subcategory_code": "runtime",
    }
    normalization_map = {
        "skills": {"Node.js": mapping},
        "_skill_mappings_by_exact_key": {"Node.js": [mapping]},
        "_skill_mappings_by_normalized_key": {"node.js": [mapping]},
    }
    candidate = {
        "candidate_id": "candidate_nodejs",
        "source_name": "Node.js",
        "document_count": 1,
        "item_type": "programming_language",
        "evidence_samples": [],
    }

    suggestions, unresolved = (
        config_iteration._deterministic_semantic_suggestions(
            [candidate],
            normalization_map,
        )
    )

    assert unresolved == []
    assert suggestions["candidate_nodejs"]["action"] == "alias_existing"
    assert suggestions["candidate_nodejs"]["anchor_alias"] == "Node.js"


def test_propose_from_run_resumes_semantic_checkpoint(tmp_path, monkeypatch):
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    (run_dir / "manifest.json").write_text(
        json.dumps({"run_id": "resume_run", "input_path": "data/chunk.xlsx"}),
        encoding="utf-8",
    )
    normalization_path = tmp_path / "normalization.yaml"
    normalization_path.write_text("skills: {}\n", encoding="utf-8")
    candidates = [
        {
            "candidate_id": f"candidate_{index}",
            "source_name": f"Skill {index}",
            "document_count": 1,
            "item_type": "tool",
            "evidence_samples": [],
        }
        for index in range(3)
    ]
    monkeypatch.setattr(
        config_iteration,
        "collect_unresolved_candidates",
        lambda *args, **kwargs: candidates,
    )
    calls: list[list[str]] = []

    def interrupted_request(selected, normalization_path, model):
        calls.append([candidate["candidate_id"] for candidate in selected])
        if len(calls) == 2:
            raise RuntimeError("interrupted")
        return {
            candidate["candidate_id"]: {
                "action": "pending_review",
                "anchor_alias": None,
                "mapping": None,
                "reason": "test",
            }
            for candidate in selected
        }

    monkeypatch.setattr(
        config_iteration,
        "request_semantic_suggestions",
        interrupted_request,
    )
    pending_dir = tmp_path / "pending"
    with pytest.raises(RuntimeError, match="interrupted"):
        config_iteration.propose_from_run(
            run_dir,
            normalization_path,
            str(pending_dir),
            "model",
            1,
            candidate_pool_path=tmp_path / "pool.json",
            max_candidates_per_review=3,
            semantic_request_batch_size=2,
        )

    resumed_calls: list[list[str]] = []

    def resumed_request(selected, normalization_path, model):
        resumed_calls.append(
            [candidate["candidate_id"] for candidate in selected]
        )
        return {
            candidate["candidate_id"]: {
                "action": "pending_review",
                "anchor_alias": None,
                "mapping": None,
                "reason": "test",
            }
            for candidate in selected
        }

    captured: dict[str, int] = {}
    monkeypatch.setattr(
        config_iteration,
        "request_semantic_suggestions",
        resumed_request,
    )

    def fake_workbook(path, run_id, selected, suggestions):
        captured["candidate_count"] = len(selected)
        captured["suggestion_count"] = len(suggestions)
        return Path(path)

    monkeypatch.setattr(
        config_iteration,
        "write_review_workbook",
        fake_workbook,
    )
    config_iteration.propose_from_run(
        run_dir,
        normalization_path,
        str(pending_dir),
        "model",
        1,
        candidate_pool_path=tmp_path / "pool.json",
        max_candidates_per_review=3,
        semantic_request_batch_size=2,
    )

    assert calls == [
        ["candidate_0", "candidate_1"],
        ["candidate_2"],
    ]
    assert resumed_calls == [["candidate_2"]]
    assert captured == {"candidate_count": 3, "suggestion_count": 3}
    assert not (
        pending_dir / "normalization_review_resume_run.semantic-checkpoint.json"
    ).exists()


def test_manual_review_flags_cover_silent_document_and_taxonomy_gaps():
    payload = extraction_payload()
    payload["company_facts"][0]["kind"] = "other"
    payload["requirements"][0]["items"][0]["item_type"] = "other"
    result = JDExtractionResult.model_validate(payload)
    issue_types = {flag["issue_type"] for flag in validate_business_rules(result)}
    assert "missing_job_title" in issue_types
    assert "company_fact_other_requires_review" in issue_types
    assert "skill_item_other_requires_review" in issue_types

    normalized = JDNormalizedResult.model_validate(
        {
            "document_id": result.document_id,
                "job_classification": {
                    "taxonomy_version": "position-taxonomy.v3.0.0",
                    "classification_status": "ambiguous",
                    "classification_policy_version": "position-classifier.v3.0",
                },
        }
    )
    assert validate_normalized_rules(normalized)[0]["issue_type"] == "job_classification_not_resolved"


def test_deterministic_fields_populate_only_ids_and_evidence_state():
    payload = extraction_payload()
    for requirement in [*payload["responsibilities"], *payload["requirements"]]:
        requirement.pop("requirement_id")
    payload["company_facts"][0].pop("fact_id")
    payload.pop("document_id")
    populated = populate_deterministic_fields(payload, "jd_final")
    assert populated["document_id"] == "jd_final"
    assert populated["responsibilities"][0]["requirement_id"] == "req_001"
    assert populated["requirements"][0]["requirement_id"] == "req_002"
    assert populated["company_facts"][0]["fact_id"] == "company_001"
    assert populated["requirements"][0]["evidence"]["alignment"] == "unresolved"


def test_all_evidence_is_exactly_aligned():
    result = align_all_evidence(JDExtractionResult.model_validate(extraction_payload()), source_blocks())
    assert result.requirements[0].evidence.alignment == "exact"
    assert result.requirements[0].evidence.start is not None


def test_non_exact_evidence_is_hard_failure():
    result = JDExtractionResult.model_validate(extraction_payload())
    result.requirements[0].evidence.quote = "熟练掌握 Python"
    with pytest.raises(SourceBindingError, match="failed exact source binding") as exc_info:
        align_all_evidence(result, source_blocks())
    assert exc_info.value.details["invalid_quote"] == "熟练掌握 Python"
    assert exc_info.value.details["exact_source_text"] == "熟练使用 Python"
    assert exc_info.value.details["object_path"] == "requirements[0]"


def test_source_binding_reports_all_invalid_evidence_in_one_pass():
    result = JDExtractionResult.model_validate(extraction_payload())
    result.requirements[0].evidence.quote = "熟练掌握 Python"
    result.requirements[1].evidence.quote = "三年以上 Python 经验"
    with pytest.raises(SourceBindingError) as exc_info:
        align_all_evidence(result, source_blocks())
    assert isinstance(exc_info.value.details, list)
    assert [item["object_path"] for item in exc_info.value.details] == ["requirements[0]", "requirements[1]"]


def test_evidence_quote_canonicalization_restores_only_unique_source_span():
    payload = extraction_payload()
    payload["requirements"][0]["evidence"]["quote"] = "熟练使用Python"

    canonicalized, corrections = canonicalize_evidence_quotes(payload, source_blocks())

    assert canonicalized["requirements"][0]["evidence"]["quote"] == "熟练使用 Python"
    assert corrections == [{
        "path": "requirements[0].evidence.quote",
        "from": "熟练使用Python",
        "to": "熟练使用 Python",
        "authority": "unique_source_span",
    }]


def test_evidence_quote_canonicalization_does_not_guess_ambiguous_span():
    payload = extraction_payload()
    payload["requirements"][0]["evidence"] = {"source_id": "src_x", "quote": "Python!"}

    canonicalized, corrections = canonicalize_evidence_quotes(
        payload,
        [{"source_id": "src_x", "text": "Python，Python"}],
    )

    assert canonicalized["requirements"][0]["evidence"]["quote"] == "Python!"
    assert corrections == []


def test_pre_schema_canonicalizers_leave_invalid_collection_shape_for_schema_error():
    payload = extraction_payload()
    payload["requirements"] = None

    evidence_payload, evidence_corrections = canonicalize_evidence_quotes(payload, source_blocks())
    canonicalized, authoritative_corrections = canonicalize_authoritative_fields(
        evidence_payload, taxonomy(), source_blocks()
    )

    assert evidence_corrections == []
    assert canonicalized["requirements"] is None
    assert authoritative_corrections == []
    with pytest.raises(SchemaValidationError, match="Schema validation failed"):
        validate_schema(populate_deterministic_fields(canonicalized, "jd_invalid"))


def test_duplicate_skills_are_flagged_then_deduplicated():
    result = JDExtractionResult.model_validate(extraction_payload())
    flags = validate_business_rules(result)
    assert any(flag["issue_type"] == "duplicate_skill_in_requirement" for flag in flags)
    result = deduplicate_extraction(result)
    assert len(result.requirements[0].items) == 1


def test_retryable_review_issues_are_rejected_before_export():
    flags = validate_business_rules(JDExtractionResult.model_validate(extraction_payload()))
    with pytest.raises(SemanticValidationError) as exc_info:
        reject_retryable_review_flags(flags)
    assert any(item["issue_type"] == "duplicate_skill_in_requirement" for item in exc_info.value.violations)


def test_title_like_first_source_block_retries_missing_job_title():
    flag = {"issue_type": "missing_job_title", "raw_text": ""}
    blocks = [
        {"source_id": "src_0001", "text": "大模型算法工程师"},
        {"source_id": "src_0002", "text": "20-30K·14薪"},
    ]
    with pytest.raises(SemanticValidationError):
        reject_retryable_review_flags([flag], blocks)


def test_section_heading_does_not_retry_missing_job_title():
    flag = {"issue_type": "missing_job_title", "raw_text": ""}
    blocks = [
        {"source_id": "src_0001", "text": "岗位职责："},
        {"source_id": "src_0002", "text": "20-30K·14薪"},
    ]
    reject_retryable_review_flags([flag], blocks)


def test_bundle_style_title_first_block_retries_missing_job_title():
    flag = {"issue_type": "missing_job_title", "raw_text": ""}
    blocks = [
        {"source_id": "src_0001", "text": "售后质量管理"},
        {"source_id": "src_0002", "text": "【source.description】"},
    ]
    with pytest.raises(SemanticValidationError):
        reject_retryable_review_flags([flag], blocks)


def test_skill_item_other_is_retryable():
    flag = {
        "issue_type": "skill_item_other_requires_review",
        "requirement_id": "req_001",
        "item_id": "WebSocket",
        "raw_text": "熟悉 WebSocket",
    }
    with pytest.raises(SemanticValidationError):
        reject_retryable_review_flags([flag])


def test_retry_prompt_corrects_missing_title_and_other_skill_type():
    prompt = build_validation_retry_prompt(
        {
            "jd_id": "jd_001", "job_title_raw": "未提及", "company": "未提及",
            "source_blocks": [{"source_id": "src_0001", "text": "大模型算法工程师"}],
        },
        "SemanticValidationError",
        [
            {"issue_type": "missing_job_title"},
            {"issue_type": "skill_item_other_requires_review", "item_id": "WebSocket"},
        ],
    )
    assert "完整原文填入 job_title.value" in prompt
    assert "具名模型、协议、API" in prompt


def test_unknown_modality_remains_soft_review_only():
    payload = extraction_payload()
    payload["requirements"][1]["modality"] = "unknown"
    flags = validate_business_rules(JDExtractionResult.model_validate(payload))
    unknown_only = [flag for flag in flags if flag["issue_type"] == "unknown_modality"]
    reject_retryable_review_flags(unknown_only)


def test_structured_company_name_omission_is_retried():
    flags = [{"issue_type": "missing_company_name", "raw_text": ""}]
    blocks = [
        {"source_id": "src_0001", "text": "大模型工程师"},
        {"source_id": "src_0002", "text": "20-30K·14薪"},
        {"source_id": "src_0003", "text": "示例科技有限公司"},
    ]
    with pytest.raises(SemanticValidationError):
        reject_retryable_review_flags(flags, blocks)


def test_placeholder_company_name_is_not_retried():
    flags = [{"issue_type": "missing_company_name", "raw_text": ""}]
    blocks = [
        {"source_id": "src_0001", "text": "大模型工程师"},
        {"source_id": "src_0002", "text": "20-30K·14薪"},
        {"source_id": "src_0003", "text": "深圳市南山区名字..."},
    ]
    reject_retryable_review_flags(flags, blocks)


def test_unknown_modality_under_explicit_education_header_remains_review_only():
    flag = {
        "issue_type": "unknown_modality",
        "raw_text": "学历:本科",
    }
    reject_retryable_review_flags([flag])


def test_unknown_modality_with_duplicate_explicit_source_context_remains_review_only():
    flag = {"issue_type": "unknown_modality", "raw_text": "Java C++ 架构设计经验"}
    blocks = [
        {"source_id": "src_1", "text": "Java C++ 架构设计经验"},
        {"source_id": "src_2", "text": "技能要求:Java C++ 架构设计经验"},
    ]
    reject_retryable_review_flags([flag], blocks)


def test_authoritative_canonicalization_uses_config_and_structured_source_not_dataset_cases():
    payload = extraction_payload()
    payload["job_title"] = None
    payload["requirements"][0]["items"][0]["item_type"] = "tool"
    payload["company_facts"] = []
    payload["employment_facts"] = [
        {
            "fact_id": "employment_001",
            "kind": "other",
            "value": "交通补贴",
            "evidence": {"source_id": "src_4", "quote": "交通补贴"},
        }
    ]
    blocks = [
        {"source_id": "src_1", "text": "后端研发工程师"},
        {"source_id": "src_2", "text": "20-30K"},
        {"source_id": "src_3", "text": "示例科技有限公司"},
        {"source_id": "src_4", "text": "交通补贴"},
    ]

    canonicalized, corrections = canonicalize_authoritative_fields(payload, taxonomy(), blocks)

    assert canonicalized["job_title"]["value"] == "后端研发工程师"
    assert canonicalized["company_facts"][0]["kind"] == "company_name"
    assert canonicalized["requirements"][0]["items"][0]["item_type"] == "programming_language"
    assert canonicalized["employment_facts"][0]["kind"] == "allowance"
    assert {item["authority"] for item in corrections} == {
        "normalization_map", "employment_kind_ontology", "structured_source_header"
    }


def test_employment_kind_inference_is_conservative_when_multiple_categories_match():
    assert infer_employment_kind("住房补贴") is None
    assert infer_employment_kind("年度健康体检") == "health_check"


def test_unknown_modality_in_requirement_section_is_canonicalized_to_required():
    payload = extraction_payload()
    payload["requirements"][0]["modality"] = "unknown"
    payload["requirements"][0]["evidence"]["source_id"] = "src_2"
    blocks = [
        {"source_id": "src_1", "text": "What We Value"},
        {"source_id": "src_2", "text": "熟练使用 Python"},
    ]

    canonicalized, corrections = canonicalize_authoritative_fields(payload, taxonomy(), blocks)

    assert canonicalized["requirements"][0]["modality"] == "required"
    assert any(item["authority"] == "requirement_section_context" for item in corrections)


def test_unknown_modality_under_bundle_requirement_marker_is_required():
    payload = extraction_payload()
    payload["requirements"][0]["modality"] = "unknown"
    payload["requirements"][0]["evidence"]["source_id"] = "src_2"
    blocks = [
        {"source_id": "src_1", "text": "【source.requirement】"},
        {"source_id": "src_2", "text": "熟练使用 Python"},
    ]

    canonicalized, corrections = canonicalize_authoritative_fields(
        payload, taxonomy(), blocks
    )

    assert canonicalized["requirements"][0]["modality"] == "required"
    assert any(
        item["authority"] == "requirement_section_context"
        for item in corrections
    )


@pytest.mark.parametrize("quote", ["学历:硕士", "学历要求：本科", "经验不限", "经验要求:应届生"])
def test_unknown_modality_in_structured_requirement_label_is_canonicalized_to_required(quote):
    payload = extraction_payload()
    payload["requirements"][0]["modality"] = "unknown"
    payload["requirements"][0]["evidence"] = {"source_id": "src_2", "quote": quote}
    blocks = [{"source_id": "src_2", "text": quote}]

    canonicalized, corrections = canonicalize_authoritative_fields(payload, taxonomy(), blocks)

    assert canonicalized["requirements"][0]["modality"] == "required"
    assert any(item["authority"] == "structured_requirement_label" for item in corrections)


def test_unknown_modality_outside_requirement_section_is_not_guessed():
    payload = extraction_payload()
    payload["requirements"][0]["modality"] = "unknown"
    canonicalized, corrections = canonicalize_authoritative_fields(payload, taxonomy(), source_blocks())
    assert canonicalized["requirements"][0]["modality"] == "unknown"
    assert not any(item["authority"] == "requirement_section_context" for item in corrections)


def test_cross_requirement_duplicate_and_contained_evidence_are_flagged():
    payload = extraction_payload()
    payload["requirements"].extend(
        [
            {
                "requirement_id": "req_004", "kind": "other", "modality": "required",
                "label": "Python", "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"},
            },
            {
                "requirement_id": "req_005", "kind": "other", "modality": "required",
                "label": "编程语言", "value": "Python",
                "evidence": {"source_id": "src_2", "quote": "Python"},
            },
            {
                "requirement_id": "req_006", "kind": "other", "modality": "required",
                "label": "Python", "evidence": {"source_id": "src_3", "quote": "Python 开发经验"},
            },
            {
                "requirement_id": "req_007", "kind": "skill", "modality": "required",
                "items": [
                    {"name": "Python", "item_type": "programming_language"},
                    {"name": "SQL", "item_type": "programming_language"},
                ],
                "proficiency": "proficient",
                "evidence": {"source_id": "src_2", "quote": "Python"},
            },
        ]
    )
    flags = validate_business_rules(JDExtractionResult.model_validate(payload))
    issue_types = {flag["issue_type"] for flag in flags}
    assert "duplicate_requirement_semantics" in issue_types
    assert "overlapping_requirement_evidence" in issue_types
    duplicate = next(flag for flag in flags if flag["issue_type"] == "duplicate_requirement_semantics")
    assert duplicate["related_requirement_id"] == "req_004"


def test_same_semantics_with_different_modality_is_conflict():
    payload = extraction_payload()
    payload["requirements"].append(
        {
            "requirement_id": "req_004", "kind": "skill", "modality": "preferred",
            "items": [
                {"name": "Python", "item_type": "programming_language"},
                {"name": "Python", "item_type": "programming_language"},
            ],
            "proficiency": "proficient",
            "evidence": {"source_id": "src_3", "quote": "Python 开发经验优先"},
        }
    )
    flags = validate_business_rules(JDExtractionResult.model_validate(payload))
    conflict = next(flag for flag in flags if flag["issue_type"] == "conflicting_requirement_modality")
    assert conflict["related_requirement_id"] == "req_002"
    assert conflict["related_modality"] == "required"
    assert conflict["modality"] == "preferred"


def test_contained_evidence_with_disjoint_skill_items_is_not_overlap():
    payload = extraction_payload()
    payload["requirements"].append(
        {
            "requirement_id": "req_004", "kind": "skill", "modality": "required",
            "items": [{"name": "SQL", "item_type": "programming_language"}],
            "proficiency": "proficient",
            "evidence": {"source_id": "src_2", "quote": "Python"},
        }
    )
    flags = validate_business_rules(JDExtractionResult.model_validate(payload))
    assert not any(
        flag["issue_type"] == "overlapping_requirement_evidence" and flag["requirement_id"] == "req_004"
        for flag in flags
    )


def test_skill_semantic_fingerprint_ignores_item_order():
    payload = extraction_payload()
    payload["requirements"][0]["items"] = [
        {"name": "Python", "item_type": "programming_language"},
        {"name": "SQL", "item_type": "programming_language"},
    ]
    payload["requirements"].append(
        {
            "requirement_id": "req_004", "kind": "skill", "modality": "required",
            "items": [
                {"name": "SQL", "item_type": "programming_language"},
                {"name": "Python", "item_type": "programming_language"},
            ],
            "proficiency": "proficient",
            "evidence": {"source_id": "src_3", "quote": "Python 开发经验优先"},
        }
    )
    flags = validate_business_rules(JDExtractionResult.model_validate(payload))
    assert any(flag["issue_type"] == "duplicate_requirement_semantics" for flag in flags)


def test_different_kinds_can_share_evidence_without_duplicate_flag():
    flags = validate_business_rules(JDExtractionResult.model_validate(extraction_payload()))
    assert not any(flag["issue_type"] == "duplicate_requirement_semantics" for flag in flags)


def test_normalization_keys_are_case_and_unicode_stable():
    work = Path("pytest_artifacts") / f"normalization_stable_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        path = work / "normalization.yaml"
        path.write_text(
            """version: '2.0'\nposition_taxonomy_version: position-taxonomy.v2.0.0\nskills:\n  PyTorch: &p\n    skill_id: FRAMEWORK_PYTORCH\n    canonical_name: PyTorch\n    category_code: framework\n  pytorch: *p\n""",
            encoding="utf-8",
        )
        loaded = load_normalization_map(str(path))
        assert list(loaded["_skills_by_normalized_key"]) == ["pytorch"]
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_conflicting_normalization_aliases_are_rejected():
    work = Path("pytest_artifacts") / f"normalization_conflict_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        path = work / "normalization.yaml"
        path.write_text(
            """version: '2.0'\nposition_taxonomy_version: position-taxonomy.v2.0.0\nskills:\n  PyTorch:\n    skill_id: ONE\n  pytorch:\n    skill_id: TWO\n""",
            encoding="utf-8",
        )
        with pytest.raises(InputFormatError, match="Conflicting normalization aliases"):
            load_normalization_map(str(path))
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_same_normalized_name_can_resolve_by_item_type():
    work = Path("pytest_artifacts") / f"normalization_typed_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        path = work / "normalization.yaml"
        path.write_text(
            """version: '2.0'\nposition_taxonomy_version: position-taxonomy.v2.0.0\nskills:\n  React:\n    skill_id: FRAMEWORK_REACT\n    canonical_name: React\n    category_code: framework\n  ReAct:\n    skill_id: METHOD_REACT\n    canonical_name: ReAct\n    category_code: methodology\n""",
            encoding="utf-8",
        )
        loaded = load_normalization_map(str(path))
        payload = extraction_payload()
        payload["requirements"][0]["items"] = [
            {"name": "React", "item_type": "framework"},
            {"name": "ReAct", "item_type": "methodology"},
        ]
        result = JDExtractionResult.model_validate(payload)
        validate_skill_item_type_contract(result, loaded)
        normalized = normalize_extraction(result, loaded, "熟练使用 React 和 ReAct")
        skills = normalized.normalized_requirements[0].skills
        assert [(item.source_name, item.skill_id) for item in skills] == [
            ("React", "FRAMEWORK_REACT"),
            ("ReAct", "METHOD_REACT"),
        ]
        assert lookup_skill_mapping(loaded, "React", "methodology") is None
        assert lookup_skill_mapping(loaded, "ReAct", "framework") is None
        assert lookup_skill_mapping(loaded, "react", "framework") is None
        assert lookup_skill_mapping(loaded, "react", "methodology") is None
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_casefold_lookup_requires_matching_type_or_explicit_override():
    work = Path("pytest_artifacts") / f"normalization_type_guard_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        path = work / "normalization.yaml"
        path.write_text(
            """version: '2.0'\nposition_taxonomy_version: position-taxonomy.v2.0.0\nskills:\n  FastAPI:\n    skill_id: FRAMEWORK_FASTAPI\n    canonical_name: FastAPI\n    category_code: framework\n  ResNet:\n    skill_id: KNOWLEDGE_RESNET\n    canonical_name: ResNet\n    category_code: domain_knowledge\n  Unity引擎:\n    skill_id: PLATFORM_UNITY\n    canonical_name: Unity\n    category_code: platform\nskill_source_type_overrides:\n  Unity引擎:\n  - framework\n""",
            encoding="utf-8",
        )
        loaded = load_normalization_map(str(path))
        assert lookup_skill_mapping(loaded, "fastapi", "framework")["skill_id"] == "FRAMEWORK_FASTAPI"
        assert lookup_skill_mapping(loaded, "fastapi", "tool") is None
        assert lookup_skill_mapping(loaded, "resnet", "database") is None
        assert lookup_skill_mapping(loaded, "unity引擎", "framework")["skill_id"] == "PLATFORM_UNITY"
        assert lookup_skill_mapping(loaded, "unity引擎", "database") is None
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_normalization_rejects_canonical_name_owned_by_multiple_skill_ids():
    work = Path("pytest_artifacts") / f"canonical_name_conflict_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        path = work / "normalization.yaml"
        path.write_text(
            """version: '2.0'\nposition_taxonomy_version: position-taxonomy.v2.0.0\nskills:\n  数据库:\n    skill_id: KNOWLEDGE_DATABASE\n    canonical_name: 数据库原理\n    category_code: domain_knowledge\n  数据库原理:\n    skill_id: KNOWLEDGE_DATABASE_PRINCIPLES\n    canonical_name: 数据库原理\n    category_code: domain_knowledge\n""",
            encoding="utf-8",
        )
        with pytest.raises(InputFormatError, match="belongs to multiple skill ids"):
            load_normalization_map(str(path))
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_education_list_fields_reject_null():
    payload = extraction_payload()
    payload["requirements"].append(
        {
            "requirement_id": "req_004", "kind": "education", "modality": "required",
            "minimum_degree": "bachelor", "majors": None, "school_constraints": [],
            "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"},
        }
    )
    with pytest.raises(ValidationError, match="majors"):
        JDExtractionResult.model_validate(payload)


@pytest.mark.parametrize(
    ("employment_facts", "error_code"),
    [
        ([{"fact_id": "employment_001", "kind": "other", "value": "20-40K",
           "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}}],
         "base_salary_in_employment_fact"),
        ([{"fact_id": "employment_001", "kind": "other", "value": "五险一金 年终奖 带薪年假",
           "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}}],
         "non_atomic_employment_fact"),
        ([{"fact_id": "employment_001", "kind": "other", "value": "应届毕业生可投",
           "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}}],
         "candidate_constraint_in_employment_fact"),
        ([{"fact_id": "employment_001", "kind": "other", "value": "充足显卡资源，LLM API无限调用",
           "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}}],
         "company_fact_in_employment_fact"),
        ([{"fact_id": "employment_001", "kind": "training", "value": "项目机会多，快速成长",
           "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}}],
         "employment_kind_evidence_mismatch"),
        ([{"fact_id": "employment_001", "kind": "allowance", "value": "五险一金",
           "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}}],
         "employment_kind_evidence_mismatch"),
    ],
)
def test_deterministic_semantic_gate_rejects_invalid_employment_facts(employment_facts, error_code):
    payload = extraction_payload()
    payload["employment_facts"] = employment_facts
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == error_code for item in exc_info.value.violations)


def test_deterministic_semantic_gate_rejects_duplicate_facts():
    payload = extraction_payload()
    payload["employment_facts"] = [
        {"fact_id": "employment_001", "kind": "work_mode", "value": "远程办公",
         "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"}},
        {"fact_id": "employment_002", "kind": "work_mode", "value": "远程办公",
         "evidence": {"source_id": "src_3", "quote": "3年以上 Python 开发经验优先"}},
    ]
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert exc_info.value.violations[0]["code"] == "duplicate_fact_semantics"


def test_deterministic_semantic_gate_rejects_duplicate_soft_skills():
    payload = extraction_payload()
    payload["requirements"].append(
        {
            "requirement_id": "req_004",
            "kind": "soft_skill",
            "modality": "required",
            "skills": ["善于沟通", "善于沟通"],
            "evidence": {"source_id": "src_2", "quote": "熟练使用 Python"},
        }
    )
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "duplicate_soft_skill_in_requirement" for item in exc_info.value.violations)


def test_normalization_taxonomy_rejects_known_skill_type_drift():
    payload = extraction_payload()
    payload["requirements"][0]["items"] = [{"name": "Python", "item_type": "tool"}]
    result = JDExtractionResult.model_validate(payload)
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_skill_item_type_contract(result, taxonomy())
    assert exc_info.value.violations == [
        {
            "code": "skill_item_type_mismatch",
            "requirement_id": "req_002",
            "name": "Python",
            "item_type": "tool",
            "expected_item_type": "programming_language",
        }
    ]


def test_deterministic_semantic_gate_rejects_candidate_facing_company_fact():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "other",
            "value": "你将直接与技术决策层协作",
            "evidence": {"source_id": "src_1", "quote": "开发推荐系统"},
        }
    ]
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "candidate_facing_company_fact" for item in exc_info.value.violations)


def test_deterministic_semantic_gate_rejects_career_path_company_fact():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "other",
            "value": "双轨发展通道",
            "evidence": {"source_id": "src_1", "quote": "开发推荐系统"},
        }
    ]
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "candidate_facing_company_fact" for item in exc_info.value.violations)


def test_deterministic_semantic_gate_rejects_candidate_skill_as_responsibility():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "熟悉大模型"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "candidate_requirement_in_responsibility" for item in exc_info.value.violations)


def test_deterministic_semantic_gate_rejects_job_title_as_responsibility():
    payload = extraction_payload()
    payload["job_title"] = {
        "value": "LLM大模型文本改写",
        "evidence": {"source_id": "src_1", "quote": "LLM大模型文本改写"},
    }
    payload["responsibilities"][0] = {
        "requirement_id": "req_001",
        "kind": "task",
        "modality": "unknown",
        "action": "LLM大模型文本改写",
        "evidence": {"source_id": "src_1", "quote": "LLM大模型文本改写"},
    }
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "job_title_duplicate_responsibility" for item in exc_info.value.violations)


def test_deterministic_semantic_gate_rejects_platform_artifact_only_in_semantic_value():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "负责搭建kanzhun开源平台"
    payload["responsibilities"][0]["evidence"]["quote"] = "负责搭建kanzhun开源平台"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "platform_artifact_in_semantic_value" for item in exc_info.value.violations)


def test_evidence_platform_artifact_is_allowed_after_semantic_value_is_cleaned():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "负责搭建开源平台"
    payload["responsibilities"][0]["evidence"]["quote"] = "负责搭建kanzhun开源平台"
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


def test_non_boss_employer_rejects_boss_zhipin_in_semantic_value():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "company_name",
            "value": "芯动力",
            "evidence": {"source_id": "src_4", "quote": "芯动力"},
        }
    ]
    payload["responsibilities"][0]["action"] = "负责BOSS直聘基于C++的推理开发"
    payload["responsibilities"][0]["evidence"]["quote"] = "负责BOSS直聘基于C++的推理开发"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(
        item["code"] == "platform_artifact_in_semantic_value" and "BOSS直聘" in item["artifacts"]
        for item in exc_info.value.violations
    )


def test_boss_zhipin_employer_allows_brand_in_real_business_semantics():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "company_name",
            "value": "BOSS直聘",
            "evidence": {"source_id": "src_4", "quote": "BOSS直聘"},
        }
    ]
    payload["responsibilities"][0]["action"] = "负责BOSS直聘B/C端主搜场景算法优化"
    payload["responsibilities"][0]["evidence"]["quote"] = "负责BOSS直聘B/C端主搜场景算法优化"
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


def test_boss_zhipin_employer_still_rejects_inserted_boss_fragment():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "company_name",
            "value": "BOSS直聘",
            "evidence": {"source_id": "src_4", "quote": "BOSS直聘"},
        }
    ]
    payload["responsibilities"][0]["action"] = "负boss责B/C端主搜场景算法优化"
    payload["responsibilities"][0]["evidence"]["quote"] = "负boss责B/C端主搜场景算法优化"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(
        "boss（中文词内插入）" in item.get("artifacts", [])
        for item in exc_info.value.violations
    )


def test_boss_zhipin_employer_still_rejects_zhipin_inserted_inside_chinese_word():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "company_name",
            "value": "BOSS直聘",
            "evidence": {"source_id": "src_4", "quote": "BOSS直聘"},
        }
    ]
    payload["responsibilities"][0]["action"] = "负责智能直聘体应用开发"
    payload["responsibilities"][0]["evidence"]["quote"] = "负责智能直聘体应用开发"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(
        "直聘（词中插入）" in item.get("artifacts", [])
        for item in exc_info.value.violations
    )


def test_boss_homepage_is_allowed_as_a_complete_product_term():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "请先查看我公司在BOSS主页的介绍"
    payload["responsibilities"][0]["evidence"]["quote"] = "请先查看我公司在BOSS主页的介绍"
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


def test_missing_company_name_does_not_authorize_boss_zhipin_semantics():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "主要BOSS直聘参与大模型应用开发"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "platform_artifact_in_semantic_value" for item in exc_info.value.violations)


def test_non_boss_employer_rejects_zhipin_inserted_inside_chinese_word():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "负责智能直聘体应用开发"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(
        "直聘（词中插入）" in item.get("artifacts", [])
        for item in exc_info.value.violations
    )


def test_boss_fragment_inserted_inside_chinese_semantics_is_rejected():
    payload = extraction_payload()
    payload["responsibilities"][0]["action"] = "负责大语boss言模型开发"
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(
        "boss（中文词内插入）" in item.get("artifacts", [])
        for item in exc_info.value.violations
    )


def test_company_name_with_zhipin_or_boss_fragment_is_not_treated_as_watermark():
    payload = extraction_payload()
    payload["company_facts"] = [
        {
            "fact_id": "company_001",
            "kind": "company_name",
            "value": "浙江蚂蚁直聘Boss企业管理咨询有限公司",
            "evidence": {
                "source_id": "src_4",
                "quote": "浙江蚂蚁直聘Boss企业管理咨询有限公司",
            },
        }
    ]
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


def test_technical_sharing_is_valid_training_fact():
    payload = extraction_payload()
    payload["employment_facts"] = [{
        "fact_id": "employment_001",
        "kind": "training",
        "value": "定期技术分享",
        "evidence": {"source_id": "src_5", "quote": "定期技术分享"},
    }]
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


@pytest.mark.parametrize(
    ("label", "value", "quote"),
    [
        ("软件能力", "熟练使用办公软件", "熟练使用办公软件"),
        ("工具能力", "熟练使用AI工具", "熟练使用AI工具"),
    ],
)
def test_explicit_technical_capability_cannot_use_other_requirement(
    label, value, quote
):
    payload = extraction_payload()
    payload["requirements"] = [
        OtherRequirement(
            requirement_id="req_other",
            kind="other",
            modality="required",
            label=label,
            value=value,
            evidence={"source_id": "src_2", "quote": quote},
        ).model_dump(exclude_none=True)
    ]

    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))

    assert any(
        item["code"] == "technical_requirement_in_other"
        for item in exc_info.value.violations
    )


def test_candidate_identity_is_stable_across_runs():
    assert config_iteration._candidate_id("run_a", " Nginx ", "tool") == config_iteration._candidate_id(
        "run_b", "Nginx", "tool"
    )


@pytest.mark.parametrize(
    ("source_name", "item_type", "skill_id"),
    [
        ("C语言", "programming_language", "LANG_C"),
        ("PPT", "tool", "TOOL_POWERPOINT"),
        ("MAYA", "tool", "TOOL_MAYA"),
        ("8D报告", "methodology", "METHOD_8D"),
        ("IIC", "domain_knowledge", "KNOWLEDGE_I2C"),
        ("PCIe", "domain_knowledge", "KNOWLEDGE_PCIE"),
        ("STM32", "platform", "PLATFORM_STM32"),
    ],
)
def test_reviewed_bundle_normalization_aliases_resolve(
    source_name, item_type, skill_id
):
    normalization_map = load_normalization_map("config/normalization_map.yaml")
    mapping = lookup_skill_mapping(normalization_map, source_name, item_type)
    assert mapping is not None
    assert mapping["skill_id"] == skill_id


@pytest.mark.parametrize(
    ("source_name", "item_type"),
    [
        ("AI工具", "tool"),
        ("办公软件", "tool"),
        ("传感器", "domain_knowledge"),
        ("供应链", "domain_knowledge"),
        ("UE", "tool"),
        ("UE", "framework"),
    ],
)
def test_generic_or_ambiguous_bundle_terms_remain_unresolved(
    source_name, item_type
):
    normalization_map = load_normalization_map("config/normalization_map.yaml")
    assert lookup_skill_mapping(normalization_map, source_name, item_type) is None


def test_revoked_generic_mappings_are_recorded_in_ledger():
    normalization_map = load_normalization_map("config/normalization_map.yaml")
    ledger = load_decision_ledger("config/normalization_decision_ledger.json")
    terms = [
        ("办公软件", "tool"),
        ("传感器", "domain_knowledge"),
        ("供应链", "domain_knowledge"),
        ("UE", "tool"),
        ("UE", "framework"),
        ("UE", "platform"),
    ]
    for source_name, item_type in terms:
        assert lookup_skill_mapping(normalization_map, source_name, item_type) is None
        decision = config_iteration.lookup_semantic_decision(ledger, source_name, item_type)
        assert decision is not None
        assert decision["action"] == "generic_unresolved"
        assert isinstance(decision.get("previous_decision"), dict)


def test_post_review_replace_field_requires_exact_expected_value():
    payload = extraction_payload()
    decision = {
        "document_id": "jd_001",
        "action": "replace_field",
        "collection": "responsibilities",
        "object_id": "req_001",
        "field": "action",
        "expected_value": "开发服务",
        "new_value": "开发推理服务",
    }
    revised = apply_annotation_decisions(payload, [decision])
    assert revised["responsibilities"][0]["action"] == "开发推理服务"
    assert payload["responsibilities"][0]["action"] == "开发服务"
    decision["expected_value"] = "错误旧值"
    with pytest.raises(ValueError, match="Expected responsibilities.action"):
        apply_annotation_decisions(payload, [decision])


def test_post_review_remove_object_requires_expected_object_fields():
    payload = extraction_payload()
    decision = {
        "document_id": "jd_001",
        "action": "remove_object",
        "collection": "company_facts",
        "object_id": "company_001",
        "expected_object": {"kind": "industry", "value": "软件"},
    }
    revised = apply_annotation_decisions(payload, [decision])
    assert revised["company_facts"] == []
    decision["expected_object"]["value"] = "错误旧值"
    with pytest.raises(ValueError, match="Expected object fields"):
        apply_annotation_decisions(payload, [decision])


def test_post_review_replace_object_preserves_identity_and_changes_kind():
    payload = extraction_payload()
    original = payload["requirements"][0]
    replacement = {
        "requirement_id": original["requirement_id"],
        "kind": "experience",
        "modality": "preferred",
        "domain": "第三方平台开放接口对接",
        "experience_unlimited": False,
        "evidence": original["evidence"],
    }
    revised = apply_annotation_decisions(
        payload,
        [{
            "document_id": payload["document_id"],
            "action": "replace_object",
            "collection": "requirements",
            "object_id": original["requirement_id"],
            "expected_object": {"kind": "skill"},
            "new_object": replacement,
        }],
    )
    assert revised["requirements"][0] == replacement


def test_post_review_append_object_requires_exact_collection_size_and_unique_identity():
    payload = extraction_payload()
    decision = {
        "document_id": "jd_001",
        "action": "append_object",
        "collection": "company_facts",
        "expected_collection_size": 1,
        "new_object": {
            "fact_id": "company_002",
            "kind": "company_name",
            "value": "示例科技",
            "evidence": {"source_id": "src_5", "quote": "示例科技"},
        },
    }

    revised = apply_annotation_decisions(payload, [decision])

    assert revised["company_facts"][-1]["fact_id"] == "company_002"
    decision["expected_collection_size"] = 0
    with pytest.raises(ValueError, match="Expected company_facts size"):
        apply_annotation_decisions(payload, [decision])


def test_post_review_replaces_explicit_skill_item_types():
    payload = extraction_payload()
    payload["requirements"][0]["items"][1]["name"] = "PyTorch"
    decision = {
        "document_id": "jd_001",
        "action": "replace_skill_item_types",
        "items": [{
            "requirement_id": "req_002",
            "name": "Python",
            "expected_item_type": "programming_language",
            "new_item_type": "tool",
        }],
    }
    revised = apply_annotation_decisions(payload, [decision])
    assert {item["item_type"] for item in revised["requirements"][0]["items"]} == {
        "programming_language",
        "tool",
    }


def test_deterministic_semantic_gate_rejects_generic_degree_tag_conflicting_with_explicit_preference():
    payload = extraction_payload()
    payload["requirements"] = [
        {
            "requirement_id": "req_002",
            "kind": "education",
            "modality": "preferred",
            "minimum_degree": "master",
            "majors": ["计算机"],
            "school_constraints": [],
            "evidence": {"source_id": "src_2", "quote": "硕士优先"},
        },
        {
            "requirement_id": "req_003",
            "kind": "education",
            "modality": "required",
            "minimum_degree": "master",
            "majors": [],
            "school_constraints": [],
            "evidence": {"source_id": "src_3", "quote": "学历:硕士"},
        },
    ]
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == "conflicting_education_degree_modality" for item in exc_info.value.violations)


@pytest.mark.parametrize(
    ("name", "error_code"),
    [
        ("TensorFlow/PyTorch", "composite_skill_item"),
        ("Linux开发/部署经验", "experience_phrase_in_skill_item"),
        ("图像处理库(OpenCV等)", "category_with_parenthetical_examples_in_skill_item"),
        ("用英文和", "dangling_conjunction_skill_item"),
        ("或繁体中文写提示词", "dangling_conjunction_skill_item"),
        ("相关产品或框架的深入使用和工程理解", "descriptive_skill_item"),
        ("差异化系统架构与优化策略设计", "descriptive_skill_item"),
        ("模型微调能力", "descriptive_skill_item"),
    ],
)
def test_deterministic_semantic_gate_rejects_composite_or_experience_skill_items(name, error_code):
    payload = extraction_payload()
    payload["requirements"][0]["items"] = [{"name": name, "item_type": "other"}]
    with pytest.raises(SemanticValidationError) as exc_info:
        validate_semantic_constraints(JDExtractionResult.model_validate(payload))
    assert any(item["code"] == error_code for item in exc_info.value.violations)


def test_deterministic_semantic_gate_allows_ab_experiment_term():
    payload = extraction_payload()
    payload["requirements"][0]["items"] = [{"name": "A/B 实验", "item_type": "domain_knowledge"}]
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


def test_fixed_composite_skill_term_is_allowed():
    payload = extraction_payload()
    payload["requirements"][0]["items"] = [{"name": "CI/CD", "item_type": "methodology"}]
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


@pytest.mark.parametrize("name", ["I/O", "I/O网络编程", "I/O 网络编程"])
def test_io_skill_terms_are_allowed(name):
    payload = extraction_payload()
    payload["requirements"][0]["items"] = [{"name": name, "item_type": "domain_knowledge"}]
    validate_semantic_constraints(JDExtractionResult.model_validate(payload))


def test_semantic_iteration_prompt_covers_high_risk_taxonomy_boundaries():
    system, user = _build_semantic_update_prompt([], {"skills": {}})
    for expected in (
        "Node.js 是 JavaScript 运行环境",
        "普通语言模型不得 alias 到大语言模型",
        "自然语言处理、VLM、多模态模型",
        "每条建议都必须给出",
        "归一化后完全相同",
    ):
        assert expected in system
    assert "Node.js 是 JavaScript 运行环境" in user


def test_normalization_is_separate_from_extraction():
    extraction = JDExtractionResult.model_validate(extraction_payload())
    normalized = normalize_extraction(extraction, taxonomy(), "20-30K")
    skill = normalized.normalized_requirements[0].skills[0]
    assert skill.source_name == "Python"
    assert skill.canonical_name == "Python 语言"
    assert skill.skill_id == "LANG_PYTHON"
    assert "canonical_name" not in extraction.model_dump()["requirements"][0]["items"][0]
    assert normalized.salary.minimum == 20000


def test_existing_run_can_be_renormalized_without_model_extraction(tmp_path):
    input_path = tmp_path / "input.csv"
    raw_text = "\n".join(
        (
            "负责开发服务",
            "熟练使用 Python",
            "3年以上 Python 开发经验优先",
            "软件公司",
            "20-30K",
        )
    )
    pd.DataFrame([{"jd_id": "jd_001", "原始文本": raw_text}]).to_csv(
        input_path, index=False
    )
    config_path = tmp_path / "normalization.yaml"
    config = taxonomy()
    config = {key: value for key, value in config.items() if not key.startswith("_")}
    config["skills"]["Python"]["canonical_name"] = "Python"
    config_path.write_text(yaml.safe_dump(config, allow_unicode=True, sort_keys=False), encoding="utf-8")
    taxonomy_path = tmp_path / "skill_taxonomy.json"
    taxonomy_snapshot = json.loads(
        Path("config/skill_taxonomy_snapshot.json").read_text(encoding="utf-8")
    )
    taxonomy_snapshot["skills"] = {
        "LANG_PYTHON": taxonomy_snapshot["skills"]["LANG_PYTHON"]
    }
    taxonomy_path.write_text(json.dumps(taxonomy_snapshot), encoding="utf-8")

    run_dir = tmp_path / "output" / "runs" / "run_001"
    success_dir = run_dir / "records" / "success"
    final_dir = run_dir / "final"
    success_dir.mkdir(parents=True)
    final_dir.mkdir(parents=True)
    annotation = extraction_payload()
    for item in (
        *annotation["responsibilities"],
        *annotation["requirements"],
        *annotation["company_facts"],
    ):
        source_number = int(item["evidence"]["source_id"].removeprefix("src_"))
        item["evidence"]["source_id"] = f"src_{source_number:04d}"
    annotation = JDExtractionResult.model_validate(annotation).model_dump(
        exclude_none=True
    )
    (success_dir / "000001_jd_001.json").write_text(
        json.dumps({"row_index": 1, "annotation": annotation}, ensure_ascii=False),
        encoding="utf-8",
    )
    (final_dir / "review_flags.jsonl").write_text("", encoding="utf-8")
    (run_dir / "logs.jsonl").write_text("", encoding="utf-8")
    (run_dir / "manifest.json").write_text(
        json.dumps({
            "run_id": "run_001",
            "input_path": str(input_path),
            "success_count": 1,
            "normalization_path": str(config_path),
        }),
        encoding="utf-8",
    )

    result = renormalize_run(
        run_dir,
        str(config_path),
        skill_taxonomy_path=taxonomy_path,
    )

    assert result == {
        "documents": 1,
        "resolved_skills": 2,
        "total_skills": 2,
        "classification_occurrence_count": 2,
        "classification_resolved_count": 2,
        "classification_missing_count": 0,
        "classification_identity_unresolved_count": 0,
    }
    normalized = json.loads((final_dir / "normalized_annotations.json").read_text(encoding="utf-8"))
    assert normalized[0]["normalized_requirements"][0]["skills"][0]["skill_id"] == "LANG_PYTHON"
    annotations = [
        json.loads(line)
        for line in (final_dir / "annotations.jsonl").read_text(
            encoding="utf-8"
        ).splitlines()
        if line.strip()
    ]
    assert annotations[0]["document_id"] == "jd_001"
    checkpoint = json.loads(
        (success_dir / "000001_jd_001.json").read_text(encoding="utf-8")
    )
    assert checkpoint["normalized"]["document_id"] == "jd_001"
    manifest = json.loads((run_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["renormalized_at"]
    assert manifest["normalization_config_version"] == "2.0"
    resolution_summary = json.loads(
        (final_dir / "normalization_resolution_summary.json").read_text(
            encoding="utf-8"
        )
    )
    assert resolution_summary["counts"]["resolved_identity"] == 2
    assert resolution_summary["counts"]["unreviewed_unresolved"] == 0


def test_job_classification_requires_position_taxonomy_v3_batch_resolution():
    payload = extraction_payload()
    payload["job_title"] = {
        "value": "AI Agent后端开发工程师(LLM集成方向)",
        "evidence": {"source_id": "src_1", "quote": "负责开发服务"},
    }
    extraction = JDExtractionResult.model_validate(payload)
    normalized = normalize_extraction(extraction, taxonomy(), "20-30K")
    classification = normalized.job_classification
    assert classification.taxonomy_version == "position-taxonomy.v3.0.0"
    assert classification.source_title == "AI Agent后端开发工程师(LLM集成方向)"
    assert classification.classification_status == "catalog_gap"
    assert classification.review_reason_codes == ["CLASSIFICATION_NOT_RUN"]
    assert classification.position_code is None


def test_normalization_config_rejects_legacy_job_family_catalog(tmp_path):
    config_path = tmp_path / "normalization.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "2.0",
                "skills": {},
                "job_families": {
                    "SE_BACKEND": {
                        "family_name": "后端开发",
                        "priority": 100,
                        "title_keywords": ["Java"],
                    }
                },
            },
            allow_unicode=True,
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    with pytest.raises(InputFormatError, match="position_taxonomy_version"):
        load_normalization_map(str(config_path))


def test_export_layer_flattens_without_changing_extraction():
    work = Path("pytest_artifacts") / f"v2_export_{uuid4().hex}"
    try:
        extraction = align_all_evidence(JDExtractionResult.model_validate(extraction_payload()), source_blocks())
        normalized = normalize_extraction(extraction, taxonomy(), "20-30K")
        snapshot = load_skill_taxonomy_snapshot(
            Path("config/skill_taxonomy_snapshot.json")
        )
        classifications, _summary = build_classification_records(
            iter_jd_skill_occurrences([normalized]), snapshot
        )
        unified = write_unified_normalized_artifacts(
            [normalized], classifications, snapshot, work
        )
        export_xlsx([extraction], unified, [], str(work / "annotations.xlsx"))
        assert (work / "normalized_annotations.jsonl").exists()
        assert (work / "annotations.xlsx").exists()
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_semantic_config_matches_v2_architecture():
    rules = load_semantic_rules()
    handbook = compile_semantic_handbook()
    assert rules["version"] == "2.6"
    assert "task" in rules["requirement_kinds"]
    assert "模型不得输出 normalized" in handbook
    assert "技术名只出现在职责中时只建立 task" in handbook
    assert "普通基础薪资范围不生成 EmploymentFact" in handbook
    assert "proficiency 完全未提及时使用 null" in handbook
    assert "描述性能力短语不得作为技能实体" in handbook
    assert "OtherRequirement 只保存现有结构确实无法表达的准入条件" in handbook
    assert "不得放入 OtherRequirement" in handbook


def test_v2_annotation_handbook_covers_current_schema_and_boundaries():
    handbook_path = Path(__file__).resolve().parents[1] / "docs" / "annotation-standard.md"
    handbook = handbook_path.read_text(encoding="utf-8")
    for section in (
        "## 2. 总体标注原则",
        "## 5. Responsibilities 标注规则",
        "## 6. Candidate Requirements 标注规则",
        "## 7. CompanyFact 标注规则",
        "## 8. EmploymentFact 标注规则",
        "## 9. 易混边界规则",
        "## 12. AI 预标注校验清单",
    ):
        assert section in handbook
    for requirement_kind in ("SkillRequirement", "EducationRequirement", "ExperienceRequirement",
                             "CertificateRequirement", "SoftSkillRequirement", "OtherRequirement"):
        assert requirement_kind in handbook
    for item_type in ("programming_language", "framework", "library", "database", "tool", "platform",
                      "methodology", "domain_knowledge", "other"):
        assert f"`{item_type}`" in handbook
    assert "数组字段没有内容时只能输出 `[]`" in handbook
    assert "不得由模型猜测标准名称" in handbook
    assert "只建立 task；仅出现技术名不构成独立候选人技能要求" in handbook
    assert "只建立 experience；不能从经历要求派生独立 skill" in handbook
    assert "独立岗位标题列只作为定位 hint" in handbook
    assert "共同构成一个筛选条件的学历、专业和院校约束，合并" in handbook
    assert "普通基础薪资范围不生成 EmploymentFact" in handbook
    assert "原文完全没有熟练度表达" in handbook
    assert "同一年限直接修饰的 `domain` 或 `role` 必须保存在同一个 ExperienceRequirement" in handbook


class FakeClient:
    def __init__(self, model: str):
        self.model = model

    def extract(self, system_prompt: str, user_prompt: str) -> DeepSeekResult:
        payload = {
            "job_title": None,
            "responsibilities": [],
            "requirements": [
                {
                    "kind": "skill",
                    "modality": "required",
                    "items": [{"name": "Python", "item_type": "programming_language"}],
                    "proficiency": "proficient",
                    "evidence": {"source_id": "src_0001", "quote": "熟练使用 Python"},
                }
            ],
            "company_facts": [],
            "employment_facts": [],
        }
        return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))


class RetryThenValidClient(FakeClient):
    def __init__(self, model: str):
        super().__init__(model)
        self.calls = 0

    def extract(self, system_prompt: str, user_prompt: str) -> DeepSeekResult:
        self.calls += 1
        if self.calls == 1:
            payload = {
                "job_title": None,
                "responsibilities": [],
                "requirements": [],
                "company_facts": [],
                "employment_facts": [
                    {
                        "kind": "salary",
                        "value": "20-40K",
                        "evidence": {"source_id": "src_0001", "quote": "20-40K"},
                    }
                ],
            }
            return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))
        assert "局部校验修复任务" in user_prompt
        payload = {
            "operations": [
                {"op": "remove", "target": {"collection": "employment_facts", "index": 0}}
            ]
        }
        return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))


class EveryRowNeedsRepairClient(FakeClient):
    def __init__(self, model: str):
        super().__init__(model)
        self.calls = 0

    def extract(self, system_prompt: str, user_prompt: str) -> DeepSeekResult:
        self.calls += 1
        if "局部校验修复任务" in user_prompt:
            payload = {"operations": [{"op": "remove", "target": {"collection": "employment_facts", "index": 0}}]}
        else:
            payload = {
                "job_title": None,
                "responsibilities": [],
                "requirements": [],
                "company_facts": [],
                "employment_facts": [{
                    "kind": "salary",
                    "value": "20-40K",
                    "evidence": {"source_id": "src_0001", "quote": "20-40K"},
                }],
            }
        return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))


class SchemaValueErrorThenValidClient(FakeClient):
    def __init__(self, model: str):
        super().__init__(model)
        self.calls = 0

    def extract(self, system_prompt: str, user_prompt: str) -> DeepSeekResult:
        self.calls += 1
        if self.calls == 1:
            payload = {
                "job_title": None,
                "responsibilities": [],
                "requirements": [
                    {
                        "kind": "experience",
                        "modality": "required",
                        "minimum_years": 3,
                        "maximum_years": 1,
                        "domain": "Python开发",
                        "experience_unlimited": False,
                        "evidence": {"source_id": "src_0001", "quote": "Python开发经验"},
                    }
                ],
                "company_facts": [],
                "employment_facts": [],
            }
        else:
            payload = {
                "operations": [
                    {
                        "op": "replace",
                        "target": {"collection": "requirements", "index": 0},
                        "value": {
                            "kind": "experience",
                            "modality": "required",
                            "minimum_years": 1,
                            "maximum_years": 3,
                            "domain": "Python开发",
                            "experience_unlimited": False,
                            "evidence": {"source_id": "src_0001", "quote": "Python开发经验"},
                        },
                    }
                ]
            }
        return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))


class MalformedPatchThenFullRetryClient(FakeClient):
    def __init__(self, model: str):
        super().__init__(model)
        self.calls = 0

    def extract(self, system_prompt: str, user_prompt: str) -> DeepSeekResult:
        self.calls += 1
        if self.calls == 1:
            payload = {
                "job_title": None,
                "responsibilities": [],
                "requirements": [],
                "company_facts": [],
                "employment_facts": [
                    {
                        "kind": "salary",
                        "value": "20-40K",
                        "evidence": {"source_id": "src_0001", "quote": "20-40K"},
                    }
                ],
            }
        elif self.calls == 2:
            assert "局部校验修复任务" in user_prompt
            payload = {"operations": [{"op": "remove", "target": {"collection": "company_facts", "index": 0}}]}
        else:
            assert "上一轮输出未通过确定性校验" in user_prompt
            payload = {
                "job_title": None,
                "responsibilities": [],
                "requirements": [],
                "company_facts": [],
                "employment_facts": [],
            }
        return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))


def test_pipeline_retries_rejected_model_output_and_audits_attempts(monkeypatch):
    monkeypatch.setattr("src.pipeline.DeepSeekClient", RetryThenValidClient)
    work = Path("pytest_artifacts") / f"v2_retry_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame([{"jd_id": "jd_retry", "原始文本": "20-40K"}]).to_csv(input_path, index=False)
        pipeline = JDExtractionPipeline(
            model="fake", normalization_path="config/normalization_map.yaml",
            continue_on_error=False, run_id="retry", audit_sample_rate=0, max_workers=1,
            source_platform="test",
            semantic_retry_attempts=1,
        )
        pipeline.run(str(input_path), str(work / "output"))
        run = work / "output" / "runs" / "retry"
        manifest = json.loads((run / "manifest.json").read_text(encoding="utf-8"))
        assert manifest["success_count"] == 1
        assert manifest["failed_count"] == 0
        assert manifest["validation_retry_count"] == 1
        assert manifest["validation_retry_recovered_count"] == 1
        assert manifest["initial_pass_count"] == 0
        assert manifest["recovered_after_first_retry_count"] == 1
        assert manifest["recovered_after_second_retry_count"] == 0
        assert manifest["validation_retry_exhausted_count"] == 0
        audit_record = json.loads(next((run / "audit").glob("*.json")).read_text(encoding="utf-8"))
        assert audit_record["audit_reason"] == "validation_retry"
        assert [item["status"] for item in audit_record["extraction_attempts"]] == ["rejected", "passed"]
        assert [item["mode"] for item in audit_record["extraction_attempts"]] == ["initial", "local_repair"]
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_pipeline_retry_budget_is_independent_for_each_row(monkeypatch):
    monkeypatch.setattr("src.pipeline.DeepSeekClient", EveryRowNeedsRepairClient)
    work = Path("pytest_artifacts") / f"per_row_retry_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame([
            {"jd_id": "jd_retry_1", "原始文本": "20-40K"},
            {"jd_id": "jd_retry_2", "原始文本": "20-40K"},
        ]).to_csv(input_path, index=False)
        pipeline = JDExtractionPipeline(
            model="fake",
            normalization_path="config/normalization_map.yaml",
            continue_on_error=False,
            run_id="per-row-retry",
            source_platform="test",
            audit_sample_rate=0,
            max_workers=1,
            semantic_retry_attempts=1,
        )
        pipeline.run(str(input_path), str(work / "output"))
        manifest = json.loads(
            (work / "output" / "runs" / "per-row-retry" / "manifest.json").read_text(encoding="utf-8")
        )
        assert manifest["success_count"] == 2
        assert manifest["failed_count"] == 0
        assert manifest["validation_retry_count"] == 2
        assert manifest["local_repair_recovered_count"] == 2
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_pipeline_runs_each_rows_validation_retries_concurrently(monkeypatch):
    class SlowRepairClient:
        def __init__(self, model: str):
            self.model = model

        def extract(self, system_prompt: str, user_prompt: str) -> DeepSeekResult:
            sleep(0.25)
            if "局部校验修复任务" in user_prompt:
                payload = {
                    "operations": [
                        {"op": "remove", "target": {"collection": "employment_facts", "index": 0}}
                    ]
                }
            else:
                payload = {
                    "job_title": None,
                    "responsibilities": [],
                    "requirements": [],
                    "company_facts": [],
                    "employment_facts": [
                        {
                            "kind": "salary",
                            "value": "20-40K",
                            "evidence": {"source_id": "src_0001", "quote": "20-40K"},
                        }
                    ],
                }
            return DeepSeekResult(data=payload, raw_response=json.dumps(payload, ensure_ascii=False))

    monkeypatch.setattr("src.pipeline.DeepSeekClient", SlowRepairClient)
    work = Path("pytest_artifacts") / f"concurrent_retry_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame(
            [
                {"jd_id": "jd_retry_1", "原始文本": "20-40K"},
                {"jd_id": "jd_retry_2", "原始文本": "20-40K"},
            ]
        ).to_csv(input_path, index=False)
        pipeline = JDExtractionPipeline(
            model="fake",
            normalization_path="config/normalization_map.yaml",
            continue_on_error=False,
            run_id="concurrent-retry",
            source_platform="test",
            audit_sample_rate=0,
            max_workers=2,
            semantic_retry_attempts=1,
        )
        started = perf_counter()
        pipeline.run(str(input_path), str(work / "output"))
        elapsed = perf_counter() - started

        manifest = json.loads(
            (work / "output" / "runs" / "concurrent-retry" / "manifest.json").read_text(encoding="utf-8")
        )
        assert manifest["api_call_count"] == 4
        assert manifest["validation_retry_count"] == 2
        assert elapsed < 0.8
        logs = [
            json.loads(line)
            for line in (work / "output" / "runs" / "concurrent-retry" / "logs.jsonl").read_text(
                encoding="utf-8"
            ).splitlines()
            if line
        ]
        api_attempts = [event for event in logs if event["event_type"] == "api_attempt_finished"]
        assert len(api_attempts) == 4
        assert {event["mode"] for event in api_attempts} == {"initial", "local_repair"}
        summary = summarize_run(work / "output" / "runs" / "concurrent-retry")
        assert summary["integrity_checks"]["api_attempt_events_match_manifest"] is True
        assert summary["api_timing"]["retry_average_ms"] >= 200
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_pipeline_audits_pydantic_value_error_context_and_recovers(monkeypatch):
    monkeypatch.setattr("src.pipeline.DeepSeekClient", SchemaValueErrorThenValidClient)
    work = Path("pytest_artifacts") / f"schema_value_error_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame([{"jd_id": "jd_schema_error", "原始文本": "Python开发经验"}]).to_csv(
            input_path, index=False
        )
        pipeline = JDExtractionPipeline(
            model="fake",
            normalization_path="config/normalization_map.yaml",
            continue_on_error=False,
            run_id="schema-value-error",
            source_platform="test",
            audit_sample_rate=0,
            max_workers=1,
            semantic_retry_attempts=1,
        )
        pipeline.run(str(input_path), str(work / "output"))
        run = work / "output" / "runs" / "schema-value-error"
        manifest = json.loads((run / "manifest.json").read_text(encoding="utf-8"))
        audit_record = json.loads(next((run / "audit").glob("*.json")).read_text(encoding="utf-8"))
        assert manifest["success_count"] == 1
        assert manifest["failed_count"] == 0
        assert [item["status"] for item in audit_record["extraction_attempts"]] == ["rejected", "passed"]
        error = audit_record["extraction_attempts"][0]["error_details"][0]["ctx"]["error"]
        assert error["exception_type"] == "ValueError"
        assert "minimum_years must not exceed maximum_years" in error["message"]
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_pipeline_records_malformed_local_repair_then_uses_remaining_full_retry(monkeypatch):
    monkeypatch.setattr("src.pipeline.DeepSeekClient", MalformedPatchThenFullRetryClient)
    work = Path("pytest_artifacts") / f"malformed_patch_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame([{"jd_id": "jd_malformed_patch", "原始文本": "20-40K"}]).to_csv(input_path, index=False)
        pipeline = JDExtractionPipeline(
            model="fake", normalization_path="config/normalization_map.yaml",
            continue_on_error=False, run_id="malformed-patch", audit_sample_rate=0, max_workers=1,
            source_platform="test",
            semantic_retry_attempts=2,
        )
        pipeline.run(str(input_path), str(work / "output"))
        run = work / "output" / "runs" / "malformed-patch"
        manifest = json.loads((run / "manifest.json").read_text(encoding="utf-8"))
        audit_record = json.loads(next((run / "audit").glob("*.json")).read_text(encoding="utf-8"))
        assert manifest["api_call_count"] == 3
        assert manifest["validation_retry_count"] == 2
        assert manifest["local_repair_count"] == 1
        assert manifest["full_reextract_count"] == 1
        assert manifest["local_repair_protocol_rejected_count"] == 1
        assert [item["mode"] for item in audit_record["extraction_attempts"]] == [
            "initial", "local_repair", "full_reextract"
        ]
        assert [item["status"] for item in audit_record["extraction_attempts"]] == [
            "rejected", "rejected", "passed"
        ]
        assert audit_record["repair_response"] is not None
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_pipeline_writes_extraction_and_normalization_outputs(monkeypatch):
    monkeypatch.setattr("src.pipeline.DeepSeekClient", FakeClient)
    work = Path("pytest_artifacts") / f"v2_pipeline_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame([{"jd_id": "jd_001", "原始文本": "熟练使用 Python"}]).to_csv(input_path, index=False)
        pipeline = JDExtractionPipeline(
            model="fake", normalization_path="config/normalization_map.yaml",
            continue_on_error=False, run_id="v2", audit_sample_rate=0, max_workers=1,
            source_platform="test",
        )
        pipeline.run(str(input_path), str(work / "output"))
        final = work / "output" / "runs" / "v2" / "final"
        assert (final / "annotations.jsonl").exists()
        assert (final / "normalized_annotations.jsonl").exists()
        assert not (final / "skill_classifications.jsonl").exists()
        assert not (final / "skill_classifications.json").exists()
        annotation = json.loads((final / "annotations.jsonl").read_text(encoding="utf-8"))
        assert annotation["requirements"][0]["kind"] == "skill"
        normalized = json.loads(
            (final / "normalized_annotations.jsonl").read_text(encoding="utf-8")
        )
        skill = normalized["normalized_requirements"][0]["skills"][0]
        assert normalized["skill_taxonomy_version"] == "skill-taxonomy-snapshot.v1"
        assert skill["identity_resolution_status"] == "resolved"
        assert skill["classification_resolution_status"] == "resolved"
        assert skill["classifications"]
        assert "category_code" not in skill
        assert "subcategory_code" not in skill
        manifest = json.loads(
            (work / "output" / "runs" / "v2" / "manifest.json").read_text(
                encoding="utf-8"
            )
        )
        assert manifest["classification_occurrence_count"] >= 1
        assert manifest["classification_resolved_count"] >= 1
        summary = summarize_run(work / "output" / "runs" / "v2")
        assert all(summary["integrity_checks"].values())
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_pipeline_row_filter_preserves_original_row_index(monkeypatch):
    monkeypatch.setattr("src.pipeline.DeepSeekClient", FakeClient)
    work = Path("pytest_artifacts") / f"v2_row_filter_{uuid4().hex}"
    try:
        work.mkdir(parents=True)
        input_path = work / "input.csv"
        pd.DataFrame([{"原始文本": "忽略"}, {"原始文本": "熟练使用 Python"}]).to_csv(input_path, index=False)
        pipeline = JDExtractionPipeline(
            model="fake", normalization_path="config/normalization_map.yaml",
            continue_on_error=False, run_id="filtered", audit_sample_rate=0, max_workers=1,
            source_platform="test",
            row_indices={2},
        )
        pipeline.run(str(input_path), str(work / "output"))
        run = work / "output" / "runs" / "filtered"
        annotation = json.loads((run / "final" / "annotations.jsonl").read_text(encoding="utf-8"))
        manifest = json.loads((run / "manifest.json").read_text(encoding="utf-8"))
        assert annotation["document_id"] == "test:input.csv:row:2:2"
        assert manifest["selected_row_indices"] == [2]
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_reusing_run_id_keeps_run_dir_for_resume():
    work = Path("pytest_artifacts") / f"v2_run_replace_{uuid4().hex}"
    try:
        old_run = work / "runs" / "same"
        success = old_run / "records" / "success"
        success.mkdir(parents=True)
        (old_run / "old.txt").write_text("old", encoding="utf-8")
        (success / "000001_jd_1.json").write_text(
            json.dumps(
                {"run_id": "same", "jd_id": "jd_1", "status": "success"},
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        audit = RunAudit(work, "same", sample_rate=0)
        assert audit.run_dir.exists()
        assert audit.resume_jd_ids == set()
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_audit_serializes_nested_validation_exception_context():
    work = Path("pytest_artifacts") / f"audit_exception_{uuid4().hex}"
    try:
        audit = RunAudit(work, "exception", sample_rate=0)
        path = audit.write_jd_audit(
            "jd_error",
            1,
            {
                "extraction_attempts": [
                    {
                        "status": "rejected",
                        "error_details": [
                            {
                                "type": "value_error",
                                "ctx": {"error": ValueError("minimum_years cannot exceed maximum_years")},
                            }
                        ],
                    }
                ]
            },
        )
        payload = json.loads(path.read_text(encoding="utf-8"))
        serialized = payload["extraction_attempts"][0]["error_details"][0]["ctx"]["error"]
        assert serialized == {
            "exception_type": "ValueError",
            "message": "minimum_years cannot exceed maximum_years",
        }
    finally:
        shutil.rmtree(work, ignore_errors=True)


def test_report_uses_v2_counts():
    summary = {
        "manifest": {"total_rows": 1, "success_count": 1, "failed_count": 0,
                     "initial_pass_count": 1, "recovered_after_first_retry_count": 0,
                     "recovered_after_second_retry_count": 0, "validation_retry_exhausted_count": 0},
        "counts": {"responsibilities": 1, "requirements": 2, "skills": 1, "company_facts": 0, "employment_facts": 0,
                   "unresolved_items": 2, "flagged_documents": 1, "normalized_skills": 4, "resolved_skills": 2},
        "kinds": [("task", 1), ("skill", 1)], "modalities": [("required", 2)],
        "item_types": [("programming_language", 1)], "review_flags": [], "job_classifications": [("resolved", 1)],
        "failures": [], "integrity_checks": {"annotations_match_manifest_success": True},
    }
    report = generate_report(summary)
    assert "Responsibilities | 1" in report
    assert "Candidate requirements | 2" in report
    assert "流水线通过率" in report
    assert "Unresolved normalization items | 2" in report
    assert "Documents with review flags | 1" in report
    assert "Skill normalization coverage | 2/4 (50.0%)" in report
    assert "Initial-pass JDs | 1" in report
    assert "Recovered after first retry | 0" in report
    assert "Job classification status" in report
    assert "Artifact integrity" in report
    assert "annotations_match_manifest_success | PASS" in report
